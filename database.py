import os
import sqlite3
import json
import time
from functools import lru_cache
from contextlib import contextmanager
from pathlib import Path
from datetime import datetime, date, timedelta

try:
    import pandas as pd
except Exception:
    pd = None

try:
    import joblib
    import numpy as np
except Exception:
    joblib = None
    np = None

@lru_cache(maxsize=1)
def _load_reservoir_cos_model():
    """Load the approved RF model once per application process.

    The model file is deliberately external to the database so it can be versioned and
    replaced under controlled technical governance.
    """
    if joblib is None or np is None:
        raise RuntimeError("Reservoir CoS calculation requires joblib and numpy. Install the application requirements.")
    if not RF_MODEL_PATH.exists():
        raise RuntimeError(f"Reservoir CoS model is not available. Place RF_model.joblib at: {RF_MODEL_PATH}")
    return joblib.load(RF_MODEL_PATH)

def _model_float(value):
    if value is None or str(value).strip() == "":
        return np.nan
    try:
        return float(value)
    except (TypeError, ValueError):
        return np.nan

def _pull_up_model_value(value):
    """Map the user-facing Pull-up selection to the approved RF model encoding.

    No=0, Semi=1, Yes=2. Numeric legacy values remain valid for old records.
    """
    if value is None or str(value).strip() == "":
        return np.nan
    normalized = str(value).strip().lower()
    mapping = {"no": 0.0, "semi": 1.0, "yes": 2.0}
    if normalized in mapping:
        return mapping[normalized]
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("Pull-up must be selected as No, Semi, or Yes.") from exc


def calculate_reservoir_cos_rows(raw_rows):
    """Calculate Reservoir CoS for every row using [Pull-up, Amplitude Ratio, BTS].

    Model output is stored as a whole-number percentage string, e.g. ``44`` for 44%.
    Empty feature values are passed as np.nan exactly as specified by the model workflow.
    """
    if isinstance(raw_rows, str):
        try:
            rows = json.loads(raw_rows or "[]")
        except json.JSONDecodeError as exc:
            raise ValueError("Reservoir CoS rows must be valid data.") from exc
    else:
        rows = raw_rows or []
    if not isinstance(rows, list):
        raise ValueError("Reservoir CoS rows must be a list.")
    model = _load_reservoir_cos_model()
    normalized = []
    for index, item in enumerate(rows, start=1):
        row = dict(item or {})
        features = [[
            _pull_up_model_value(row.get("pull_up")),
            _model_float(row.get("amplitude_ratio")),
            _model_float(row.get("base_tight_sarah")),
        ]]
        try:
            probability = float(model.predict_proba(features)[0][1])
        except Exception as exc:
            raise ValueError(f"Reservoir CoS could not be calculated for row {index}: {exc}") from exc
        row["reservoir_cos_pct"] = str(int(round(probability * 100)))
        normalized.append(row)
    return json.dumps(normalized, separators=(",", ":"))


def _cos_probability(value, label):
    """Normalize a CoS entered/displayed as either 0-1 or 0-100 to a probability."""
    if value is None or str(value).strip() == "":
        raise ValueError(f"{label} is required to calculate Presence CoS.")
    try:
        numeric = float(str(value).strip().replace("%", ""))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} must be numeric.") from exc
    if numeric < 0 or numeric > 100:
        raise ValueError(f"{label} must be between 0 and 100%.")
    return numeric / 100.0 if numeric > 1 else numeric


def _task_field_value(conn, project_id, task_name, field_key):
    row = conn.execute("""
        SELECT tdf.field_value
        FROM project_tasks pt
        LEFT JOIN task_dynamic_fields tdf
          ON tdf.task_id = pt.task_id AND tdf.field_key = ?
        WHERE pt.project_id = ? AND pt.task_name = ?
        ORDER BY pt.task_id DESC
        LIMIT 1
    """, (field_key, project_id, task_name)).fetchone()
    return "" if not row or row["field_value"] is None else str(row["field_value"]).strip()


def _final_reservoir_cos_value(conn, project_id):
    """Return the last completed Reservoir CoS row, which is the final Reservoir CoS."""
    raw = _task_field_value(conn, project_id, "Reservoir CoS", "reservoir_cos_rows")
    if not raw:
        return ""
    try:
        rows = json.loads(raw)
    except (TypeError, json.JSONDecodeError):
        return ""
    if not isinstance(rows, list):
        return ""
    for row in reversed(rows):
        value = (row or {}).get("reservoir_cos_pct")
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    return ""


def calculate_presence_cos(conn, project_id):
    """Calculate Presence CoS from the final Reservoir, Trap and Seal CoS values.

    Each component can be stored as a decimal probability or a whole percentage.
    The stored/displayed Presence CoS is a whole percentage, e.g. ``18`` for 18%.
    """
    reservoir = _final_reservoir_cos_value(conn, project_id)
    trap = _task_field_value(conn, project_id, "Trap CoS", "trap_cos_pct")
    seal = _task_field_value(conn, project_id, "Seal CoS", "seal_cos_pct")
    values = {
        "presence_reservoir_cos_pct": reservoir,
        "presence_trap_cos_pct": trap,
        "presence_seal_cos_pct": seal,
    }
    if not reservoir or not trap or not seal:
        values["presence_cos"] = ""
        return values
    probability = (
        _cos_probability(reservoir, "Final Reservoir CoS")
        * _cos_probability(trap, "Trap CoS")
        * _cos_probability(seal, "Seal CoS")
    )
    values["presence_cos"] = str(int(round(probability * 100)))
    return values

def _seal_number(value, label):
    """Return one numeric Seal CoS input or raise a field-specific validation error."""
    if value is None or str(value).strip() == "":
        raise ValueError(f"{label} is required to calculate Seal CoS.")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} must be numeric.") from exc


def calculate_seal_cos(fields):
    """Calculate Seal CoS and return it as a whole-number percentage string.

    Rule:
    - activity > 0.9: activity × fracture permeability
    - activity <= 0.9: average(dip, azimuth vs. SHmax, fault confidence)
      × fracture permeability

    Input values are used exactly as entered; the displayed result is converted from
    a decimal probability to a percentage (e.g., 0.44 becomes ``44``).
    """
    values = fields or {}
    activity_raw = values.get("seal_recent_activity_age")
    fracture_raw = values.get("seal_fracture_permeability")
    # Allow a completely blank new form to be saved without creating a spurious error.
    inputs = [
        activity_raw,
        values.get("seal_dip"),
        values.get("seal_azimuth_vs_shmax"),
        values.get("seal_fault_level_confidence"),
        fracture_raw,
    ]
    if not any(str(value or "").strip() for value in inputs):
        return ""

    activity = _seal_number(activity_raw, "Most recent age of activity")
    fracture_permeability = _seal_number(fracture_raw, "Fracture Permeability")
    if activity > 0.9:
        seal_cos = activity * fracture_permeability
    else:
        dip = _seal_number(values.get("seal_dip"), "Dip")
        azimuth = _seal_number(values.get("seal_azimuth_vs_shmax"), "Azimuth vs. SHmax")
        fault_confidence = _seal_number(values.get("seal_fault_level_confidence"), "Fault Level of Confidence")
        seal_cos = ((dip + azimuth + fault_confidence) / 3.0) * fracture_permeability

    return str(int(round(seal_cos * 100)))

try:
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception:
    Font = PatternFill = Border = Side = Alignment = None
    get_column_letter = None
    Table = TableStyleInfo = None

APP_TITLE = "Segment Maturation and Execution System"
SCHEMA_VERSION = 15
DEFAULT_DB_PATH = Path(__file__).resolve().parent / "pipeline_tracker.db"
RF_MODEL_PATH = Path(os.environ.get("SEGMENT_TRACKER_RF_MODEL_PATH", str(Path(__file__).resolve().parent / "RF_model.joblib"))).expanduser().resolve()

# Shared/root directories only. Do NOT put a field name or well name here.
# The app automatically builds: ROOT / [Field Name] / [Well Name] / [Section Folder]
#
# Main well directory buttons:
#   Open Lead Folder, Open Well Folder, Segmentation, PDA, MTR
# Server/FastX path is optional and used only if the share is mounted on Linux.
# Windows client path is what folder buttons return to users' browsers.
# Example client PDA path: \\YOUR_WINDOWS_SERVER\YOUR_SHARE\Wells\MDFT\MDFT-3\PDA
WELL_OVERVIEW_DIRECTORY_ROOT = Path("/mnt/wells")
WINDOWS_WELL_SHARE_ROOT = r"\\YOUR_WINDOWS_SERVER\YOUR_SHARE\Wells"

# Separate lead-workflow directory used only by the Task Update stage buttons:
#   Open Identification Folder, Open Risking Folder, Open Segmentation Folder
# Example client path: \\YOUR_WINDOWS_SERVER\YOUR_SHARE\Lead_Workflow\MDFT\MDFT-3\Leads\Identification
LEAD_WORKFLOW_DIRECTORY_ROOT = Path("/mnt/lead_workflow")
WINDOWS_LEAD_WORKFLOW_SHARE_ROOT = r"\\YOUR_WINDOWS_SERVER\YOUR_SHARE\Lead_Workflow"

# Easy-to-change folder placeholders. Keys are used by frontend buttons; values are
# subfolders under the resolved ROOT / Field / Well path.
WELL_OVERVIEW_DIRECTORY_MAP = {
    # Main well directory buttons.
    "lead": "Leads",
    "well": "",
    "segmentation": "Segmentation",
    "pda": "PDA",
    "mtr": "MTR",

    # Task Update stage buttons in the separate lead-workflow directory.
    "identification_workflow": "Leads/Identification",
    "risking_workflow": "Leads/Risking",
    "segmentation_workflow": "Leads/Segmentation",
}

LEAD_WORKFLOW_SECTION_KEYS = {
    "identification_workflow",
    "risking_workflow",
    "segmentation_workflow",
}

# Components where users typically need a physical/share location for supporting files.
# The app automatically generates: ROOT / Field / Well / Component Files / Component Name
COMPONENT_FILE_SECTIONS = {
    # First nine Prospect Maturation components require generated file locations.
    "Reservoir Area Definition", "Thickness Estimation", "Lead Resource Assessment",
    "Seismic Signature Validation", "Reservoir CoS", "Trap CoS", "Seal CoS",
    "Presence CoS Evaluation", "Prospect Evaluation Presentation",
    # Additional components with supporting-file requirements.
    "Approval to Stake", "Well Proposal", "GHEER", "Quicklook Logs Interpretation", "Quicklook Logs", "SAD Model",
    "Executive Summary", "URED Update", "Aramco Picks", "Aramco Approved Picks", "Flowback Results",
    "SAD Update", "Executive Summary Final", "Final Log Analysis", "PVAD Structural MTR",
    "Resource Assessment Update", "PDA", "Pre-Drilling Resource Assessment",
    "Post-Drilling Resource Assessment",
}

def _safe_folder_name(name: str) -> str:
    text = str(name or "").strip() or "Component"
    # Windows also disallows backslash as part of a folder name.
    text = text.replace('\\', '-')
    for ch in '<>:"/|?*':
        text = text.replace(ch, "-")
    return " ".join(text.split())[:120]

STATUSES = [
    "Not Assigned",
    "Assigned",
    "In Progress",
    "Ready for Review",
    "Under Review",
    "Ready for Approval",
    "Returned for Update",
    "Approved",
    "Not Applicable",
]

DONE_STATUSES = {"Approved", "Not Applicable", "Complete"}
ACTIVE_STATUSES = {"Assigned", "In Progress", "Ready for Review", "Under Review", "Ready for Approval", "Returned for Update"}

STAGE_ORDER = [
    "Lead Identification",
    "Risking",
    "Segmentation",
    "Pre-Well Delivery",
    "Well Delivery",
    "Post-Drilling",
    "Post-Testing",
]

PROSPECT_STAGES = ["Lead Identification", "Risking", "Segmentation", "Pre-Well Delivery"]
BP_EXECUTION_STAGES = ["Well Delivery", "Post-Drilling", "Post-Testing"]
BOARD_STAGE_ORDER = STAGE_ORDER[:]

# Template tuple: id, component, stage, role, duration, depends_on, branch_type, output
PIPELINE_TEMPLATES = [
    (1, "Reservoir Area Definition", "Lead Identification", "Lead Owner", 3, None, "normal", "Reservoir area defined"),
    (2, "Thickness Estimation", "Lead Identification", "Lead Owner", 3, None, "normal", "Thickness estimated"),
    (3, "Lead Resource Assessment", "Lead Identification", "Reservoir Engineer", 3, None, "normal", "Lead resources assessed"),
    (4, "Seismic Signature Validation", "Risking", "Geologist", 2, None, "normal", "Seismic signature validated"),
    (5, "Reservoir CoS", "Risking", "Reservoir Engineer", 3, None, "normal", "Reservoir CoS entered"),
    (6, "Trap CoS", "Risking", "Geologist", 3, None, "normal", "Trap CoS entered"),
    (7, "Seal CoS", "Risking", "Geologist", 3, None, "normal", "Seal CoS entered"),
    (8, "Presence CoS Evaluation", "Risking", "Geologist", 2, None, "normal", "Presence CoS entered"),
    (9, "Prospect Evaluation Presentation", "Segmentation", "Lead Owner", 2, None, "normal", "Presentation prepared"),
    (10, "Well Creation", "Pre-Well Delivery", "Well Planner", 3, None, "normal", "Well created"),
    (11, "Pre-Drilling Resource Assessment", "Pre-Well Delivery", "Reservoir Engineer", 3, None, "normal", "Pre-drilling resources assessed"),
    (12, "Staking Moving Tolerance", "Pre-Well Delivery", "Geologist", 2, None, "normal", "Moving tolerance recorded"),
    (13, "Approval to Stake", "Pre-Well Delivery", "Stakeholder", 2, None, "normal", "Approval to stake complete"),
    (14, "BP Execution Gate", "Well Delivery", "Portfolio Team", 1, None, "normal", "BP execution gate complete"),
    (15, "Well Proposal", "Well Delivery", "Drilling Engineer", 3, None, "normal", "Well proposal complete"),
    (16, "Site Preparation", "Well Delivery", "Field Team", 4, None, "normal", "Site preparation complete"),
    (17, "Approval To Drill", "Well Delivery", "Approver", 2, None, "normal", "Approval to drill complete"),
    (18, "GHEER", "Well Delivery", "HSE / Review Team", 2, None, "normal", "GHEER complete"),
    (19, "Quicklook Logs Interpretation", "Post-Drilling", "Petrophysicist", 2, None, "normal", "Quicklook logs interpreted"),
    (20, "Aramco Picks", "Post-Drilling", "Geologist", 2, None, "normal", "Aramco picks complete"),
    (21, "Post-Drilling Resource Assessment", "Post-Drilling", "Reservoir Engineer", 3, None, "normal", "Post-drilling resources assessed"),
    (22, "SAD Model", "Post-Drilling", "PDA Owner", 3, None, "normal", "SAD model complete"),
    (23, "Executive Summary", "Post-Drilling", "Manager", 2, None, "normal", "Executive summary complete"),
    (24, "URED Update", "Post-Drilling", "Reservoir Engineer", 2, None, "normal", "URED update complete"),
    (25, "Post-Well Outcome & Decision Gate", "Post-Drilling", "Portfolio Team", 3, None, "normal", "Outcome decision complete"),
    (26, "Flowback Results", "Post-Testing", "Analyst", 3, None, "normal", "Flowback results captured"),
    (27, "SAD Update", "Post-Testing", "PDA Owner", 3, None, "normal", "SAD update complete"),
    (28, "Executive Summary Final", "Post-Testing", "Manager", 2, None, "normal", "Final executive summary complete"),
    (29, "Final Log Analysis", "Post-Testing", "Petrophysicist", 2, None, "normal", "Final log analysis complete"),
    (30, "PVAD Structural MTR", "Post-Testing", "Reporting Owner", 2, None, "normal", "PVAD structural MTR complete"),
    (31, "Resource Assessment Update", "Post-Testing", "Reservoir Engineer", 2, None, "normal", "Resource assessment updated"),
    (32, "PDA", "Post-Testing", "PDA Owner", 2, None, "normal", "PDA complete"),
]

WORKFLOW_TASK_RENAMES = {
    "Quicklook Logs": "Quicklook Logs Interpretation",
    "Aramco Approved Picks": "Aramco Picks",
    "Flowback": "Flowback Results",
    "Flow Back": "Flowback Results",
    "Post Test": "Flowback Results",
}

DYNAMIC_FIELD_OVERVIEW_MAP = {
    "lead_piip_gas_mean": "lead_ogip",
    "pre_drill_piip_gas_mean": "pre_drill_estimation",
    "post_drill_piip_gas_mean": "post_drill_estimation",
    "resource_update_gas_mean": "post_drill_estimation",
    "presence_cos": "derisking",
    "quicklook_pay_thickness_ft": "quick_look_pay",
    "quicklook_average_porosity_pct": "quick_look_porosity",
    "quicklook_average_swt_pct": "quick_look_swt",
    "flowback_gas_rate_mmscfd": "flowback_results",
}



def utc_now_str() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


def today_str() -> str:
    return date.today().isoformat()


def parse_iso_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(str(value))
    except Exception:
        return None


def days_between(start_val, end_val=None):
    start_dt = parse_iso_date(start_val)
    if not start_dt:
        return None
    end_dt = parse_iso_date(end_val) if end_val else date.today()
    if not end_dt:
        end_dt = date.today()
    return (end_dt - start_dt).days


def _to_float_or_none(value):
    if value is None:
        return None
    text = str(value).replace(',', '').strip()
    if not text or text == '-':
        return None
    try:
        return float(text)
    except Exception:
        return None

def compact_text(value, max_len=20):
    value = str(value or "-").strip()
    if len(value) <= max_len:
        return value
    return value[: max_len - 1].rstrip() + "…"



def health_from_target(target_date_str, overall_status):
    target_dt = parse_iso_date(target_date_str)
    if overall_status == "Completed":
        return "Completed"
    if not target_dt:
        return "On Track"
    delta = (target_dt - date.today()).days
    if delta < 0:
        return "Overdue"
    if delta <= 14:
        return "Due Soon"
    return "On Track"


def task_due_health(planned_finish, status):
    finish = parse_iso_date(planned_finish)
    if status == "Complete":
        return "Complete"
    if not finish:
        return "No Date"
    delta = (finish - date.today()).days
    if delta < 0:
        return "Overdue"
    if delta <= 3:
        return "Due Soon"
    return "OK"


def date_variance_days(planned_finish, actual_finish=None):
    planned = parse_iso_date(planned_finish)
    if not planned:
        return None
    actual = parse_iso_date(actual_finish) if actual_finish else date.today()
    return (actual - planned).days


def parse_optional_date(value, field_name):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return date.fromisoformat(value).isoformat()
    except Exception:
        raise ValueError(f"{field_name} must be blank or in YYYY-MM-DD format.")


def progress_bar_text(percent, width=18):
    percent = max(0, min(100, int(round(percent or 0))))
    filled = int(round((percent / 100) * width))
    return "█" * filled + "░" * (width - filled) + f" {percent}%"




def parse_field_and_well(project_name: str):
    name = (project_name or '').strip()
    if not name:
        return '', ''
    # Common convention: MDFT-3 -> field MDFT, well MDFT-3.
    if '-' in name:
        return name.split('-', 1)[0].strip(), name
    parts = name.split()
    if len(parts) > 1:
        return parts[0], name
    return name, name
class Database:
    def __init__(self, db_path: Path, bootstrap: bool = True):
        self.db_path = Path(db_path).resolve()
        self.conn = sqlite3.connect(str(self.db_path), timeout=30, check_same_thread=True)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON;")
        self.conn.execute("PRAGMA busy_timeout = 30000;")
        if bootstrap:
            self.conn.execute("PRAGMA journal_mode = WAL;")
            self.conn.execute("PRAGMA synchronous = NORMAL;")
            self.create_schema()
            self.seed_templates()
            self.apply_workflow_updates()

    @contextmanager
    def write_transaction(self, attempts: int = 5):
        """Acquire an immediate SQLite write transaction with bounded retries."""
        for attempt in range(attempts):
            try:
                self.conn.execute("BEGIN IMMEDIATE")
                break
            except sqlite3.OperationalError as exc:
                if "locked" not in str(exc).lower() or attempt == attempts - 1:
                    raise
                time.sleep(0.08 * (2 ** attempt))
        try:
            yield
        except Exception:
            self.conn.rollback()
            raise
        else:
            self.conn.commit()

    def close(self):
        try:
            self.conn.close()
        except Exception:
            pass

    def create_schema(self):
        cur = self.conn.cursor()
        cur.executescript("""
            CREATE TABLE IF NOT EXISTS task_templates (
                template_id INTEGER PRIMARY KEY,
                sequence_no INTEGER NOT NULL,
                task_name TEXT NOT NULL UNIQUE,
                stage_group TEXT NOT NULL,
                default_role TEXT,
                default_duration_days INTEGER DEFAULT 3,
                depends_on_template_id INTEGER,
                branch_type TEXT DEFAULT 'normal',
                mandatory_output TEXT
            );

            CREATE TABLE IF NOT EXISTS projects (
                project_id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_name TEXT NOT NULL UNIQUE,
                overall_status TEXT NOT NULL DEFAULT 'In Progress',
                current_stage TEXT,
                current_task TEXT,
                current_owner TEXT,
                drill_result TEXT,
                start_date TEXT,
                target_date TEXT,
                location TEXT,
                business_plan_enabled INTEGER NOT NULL DEFAULT 0,
                business_plan_year INTEGER,
                active_well_enabled INTEGER NOT NULL DEFAULT 0,
                pipeline_type TEXT NOT NULL DEFAULT 'prospect',
                current_stage_started_at TEXT,
                last_updated TEXT,
                archived INTEGER NOT NULL DEFAULT 0,
                lead_folder_path TEXT
            );

            CREATE TABLE IF NOT EXISTS project_tasks (
                task_id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                template_id INTEGER NOT NULL,
                sequence_no INTEGER NOT NULL,
                task_name TEXT NOT NULL,
                stage_group TEXT NOT NULL,
                assigned_to TEXT,
                backup_owner TEXT,
                approver TEXT,
                status TEXT NOT NULL DEFAULT 'Not Started',
                planned_start TEXT,
                actual_start TEXT,
                planned_finish TEXT,
                actual_finish TEXT,
                output_notes TEXT,
                comments TEXT,
                priority TEXT NOT NULL DEFAULT 'Normal',
                business_plan_enabled INTEGER NOT NULL DEFAULT 0,
                business_plan_year INTEGER,
                is_active INTEGER NOT NULL DEFAULT 1,
                last_updated TEXT,
                FOREIGN KEY(project_id) REFERENCES projects(project_id) ON DELETE CASCADE,
                FOREIGN KEY(template_id) REFERENCES task_templates(template_id)
            );

            CREATE TABLE IF NOT EXISTS task_history (
                history_id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER NOT NULL,
                project_id INTEGER NOT NULL,
                task_name TEXT NOT NULL,
                action_type TEXT,
                old_status TEXT,
                new_status TEXT,
                changed_at TEXT NOT NULL,
                changed_by TEXT,
                comment TEXT,
                FOREIGN KEY(task_id) REFERENCES project_tasks(task_id) ON DELETE CASCADE,
                FOREIGN KEY(project_id) REFERENCES projects(project_id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS project_overview (
                project_id INTEGER PRIMARY KEY,
                derisking TEXT,
                ogip TEXT,
                lead_ogip TEXT,
                preliminary_resource_estimation TEXT,
                pre_drill_estimation TEXT,
                post_drill_estimation TEXT,
                reservoir_pressure TEXT,
                reservoir_gradient TEXT,
                flowback_results TEXT,
                pay TEXT,
                porosity TEXT,
                swt TEXT,
                quick_look_pay TEXT,
                quick_look_porosity TEXT,
                quick_look_swt TEXT,
                last_updated TEXT,
                FOREIGN KEY(project_id) REFERENCES projects(project_id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS task_dynamic_fields (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER NOT NULL,
                field_key TEXT NOT NULL,
                field_value TEXT,
                updated_at TEXT,
                UNIQUE(task_id, field_key),
                FOREIGN KEY(task_id) REFERENCES project_tasks(task_id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS lead_summary_snapshots (
                project_id INTEGER PRIMARY KEY,
                snapshot_json TEXT NOT NULL,
                captured_at TEXT NOT NULL,
                captured_by TEXT,
                FOREIGN KEY(project_id) REFERENCES projects(project_id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS business_plan_commitment (
                commitment_id INTEGER PRIMARY KEY CHECK (commitment_id = 1),
                produced REAL NOT NULL DEFAULT 0,
                pending_tie_in REAL NOT NULL DEFAULT 0,
                base REAL NOT NULL DEFAULT 0,
                core_extension_wells REAL NOT NULL DEFAULT 0,
                planned_yet_to_find REAL NOT NULL DEFAULT 0,
                last_updated TEXT
            );
        """)
        self._ensure_column("project_tasks", "backup_owner", "TEXT")
        self._ensure_column("project_tasks", "approver", "TEXT")
        self._ensure_column("project_tasks", "output_notes", "TEXT")
        self._ensure_column("project_tasks", "comments", "TEXT")
        self._ensure_column("project_tasks", "priority", "TEXT NOT NULL DEFAULT 'Normal'")
        self._ensure_column("project_tasks", "business_plan_enabled", "INTEGER NOT NULL DEFAULT 0")
        self._ensure_column("project_tasks", "business_plan_year", "INTEGER")
        self._ensure_column("project_tasks", "last_updated", "TEXT")
        self._ensure_column("projects", "location", "TEXT")
        self._ensure_column("projects", "business_plan_enabled", "INTEGER NOT NULL DEFAULT 0")
        self._ensure_column("projects", "business_plan_year", "INTEGER")
        self._ensure_column("projects", "active_well_enabled", "INTEGER NOT NULL DEFAULT 0")
        self._ensure_column("projects", "pipeline_type", "TEXT NOT NULL DEFAULT 'prospect'")
        self._ensure_column("projects", "current_stage_started_at", "TEXT")
        self._ensure_column("projects", "last_updated", "TEXT")
        self._ensure_column("projects", "archived", "INTEGER NOT NULL DEFAULT 0")
        self._ensure_column("projects", "lead_folder_path", "TEXT")
        self._ensure_column("projects", "lead_x", "INTEGER")
        self._ensure_column("projects", "lead_y", "INTEGER")
        self._ensure_column("project_overview", "derisking", "TEXT")
        self._ensure_column("project_overview", "ogip", "TEXT")
        self._ensure_column("project_overview", "lead_ogip", "TEXT")
        self._ensure_column("project_overview", "preliminary_resource_estimation", "TEXT")
        self._ensure_column("project_overview", "pre_drill_estimation", "TEXT")
        self._ensure_column("project_overview", "post_drill_estimation", "TEXT")
        self._ensure_column("project_overview", "reservoir_pressure", "TEXT")
        self._ensure_column("project_overview", "reservoir_gradient", "TEXT")
        self._ensure_column("project_overview", "flowback_results", "TEXT")
        self._ensure_column("project_overview", "pay", "TEXT")
        self._ensure_column("project_overview", "porosity", "TEXT")
        self._ensure_column("project_overview", "swt", "TEXT")
        self._ensure_column("project_overview", "quick_look_pay", "TEXT")
        self._ensure_column("project_overview", "quick_look_porosity", "TEXT")
        self._ensure_column("project_overview", "quick_look_swt", "TEXT")
        self._ensure_column("project_overview", "last_updated", "TEXT")
        self.conn.execute("""
            UPDATE project_overview
            SET lead_ogip = COALESCE(NULLIF(lead_ogip, ''), NULLIF(preliminary_resource_estimation, ''), NULLIF(ogip, ''))
            WHERE COALESCE(lead_ogip, '') = ''
              AND (COALESCE(preliminary_resource_estimation, '') != '' OR COALESCE(ogip, '') != '')
        """)
        self.conn.execute("""
            INSERT OR IGNORE INTO business_plan_commitment (commitment_id, last_updated)
            VALUES (1, ?)
        """, (utc_now_str(),))
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)
        self._ensure_column("projects", "revision", "INTEGER NOT NULL DEFAULT 0")
        self._ensure_column("project_tasks", "revision", "INTEGER NOT NULL DEFAULT 0")
        self.conn.executescript("""
            CREATE INDEX IF NOT EXISTS idx_project_tasks_project_active_sequence
                ON project_tasks(project_id, is_active, sequence_no);
            CREATE INDEX IF NOT EXISTS idx_project_tasks_project_status
                ON project_tasks(project_id, status);
            CREATE INDEX IF NOT EXISTS idx_project_tasks_project_name
                ON project_tasks(project_id, task_name);
            CREATE INDEX IF NOT EXISTS idx_task_dynamic_fields_task_key
                ON task_dynamic_fields(task_id, field_key);
            CREATE INDEX IF NOT EXISTS idx_task_history_project_changed
                ON task_history(project_id, changed_at DESC);
            CREATE INDEX IF NOT EXISTS idx_projects_archived_stage
                ON projects(archived, pipeline_type, current_stage);
            CREATE INDEX IF NOT EXISTS idx_projects_portfolio
                ON projects(archived, business_plan_enabled, business_plan_year, active_well_enabled);
        """)
        self.conn.commit()

    def _ensure_column(self, table_name, column_name, column_def):
        cols = [r["name"] for r in self.conn.execute(f"PRAGMA table_info({table_name})").fetchall()]
        if column_name not in cols:
            self.conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_def}")


    def normalize_pipeline_types(self, commit=True):
        """Separate workflow placement from the Business Plan reporting flag.

        `pipeline_type` determines which operational board a record belongs to.
        `business_plan_enabled` only controls inclusion in Portfolio.  The small
        migration below preserves legacy BP-created wells by identifying records
        whose entire Prospect Maturation task set was initialized as Not Applicable.
        """
        projects = self.conn.execute("SELECT project_id, pipeline_type, business_plan_enabled FROM projects").fetchall()
        for project in projects:
            pipeline = str(project["pipeline_type"] or "").strip().lower()
            # Legacy BP-created wells have all Prospect tasks marked Not Applicable.
            if pipeline not in {"prospect", "bp"}:
                pipeline = "prospect"
            if pipeline == "prospect" and int(project["business_plan_enabled"] or 0) == 1:
                counts = self.conn.execute("""
                    SELECT
                        COUNT(*) AS total,
                        SUM(CASE WHEN status = 'Not Applicable' THEN 1 ELSE 0 END) AS not_applicable
                    FROM project_tasks
                    WHERE project_id = ? AND stage_group IN (?,?,?,?)
                """, (project["project_id"], *PROSPECT_STAGES)).fetchone()
                if counts and int(counts["total"] or 0) > 0 and int(counts["total"] or 0) == int(counts["not_applicable"] or 0):
                    pipeline = "bp"
            self.conn.execute("UPDATE projects SET pipeline_type = ? WHERE project_id = ?", (pipeline, project["project_id"]))
        if commit:
            self.conn.commit()

    def normalize_legacy_statuses(self, commit=True):
        legacy_map = {
            'Not Started': 'Not Assigned',
            'Ready': 'In Progress',
            'Waiting': 'Under Review',
            'Complete': 'Approved',
            'Done': 'Approved',
            'Completed': 'Approved',
        }
        for old, new in legacy_map.items():
            self.conn.execute("UPDATE project_tasks SET status = ? WHERE status = ?", (new, old))
            self.conn.execute("UPDATE task_history SET old_status = ? WHERE old_status = ?", (new, old))
            self.conn.execute("UPDATE task_history SET new_status = ? WHERE new_status = ?", (new, old))
        self.conn.execute("UPDATE projects SET overall_status = 'In Progress' WHERE overall_status IN ('Ready', 'Waiting')")
        self.conn.execute("UPDATE projects SET overall_status = 'Completed' WHERE overall_status IN ('Approved')")
        if commit:
            self.conn.commit()

    def seed_templates(self):
        cur = self.conn.cursor()
        count = cur.execute("SELECT COUNT(*) AS c FROM task_templates").fetchone()["c"]
        if count:
            return
        cur.executemany("""
            INSERT INTO task_templates (
                template_id, sequence_no, task_name, stage_group, default_role,
                default_duration_days, depends_on_template_id, branch_type, mandatory_output
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, [(t[0], idx + 1, t[1], t[2], t[3], t[4], t[5], t[6], t[7]) for idx, t in enumerate(PIPELINE_TEMPLATES)])
        self.conn.commit()

    def _schema_version(self) -> int:
        row = self.conn.execute("SELECT value FROM app_settings WHERE key = 'schema_version'").fetchone()
        try:
            return int(row["value"]) if row else 0
        except (TypeError, ValueError):
            return 0

    def _set_schema_version(self, version: int) -> None:
        self.conn.execute("""
            INSERT INTO app_settings (key, value) VALUES ('schema_version', ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """, (str(version),))

    def apply_workflow_updates(self):
        """Run each workflow migration once without deleting historic task records.

        Current tasks are brought to their canonical names, sequence and stage. Retired
        tasks are retained as inactive records so their inputs and audit trail survive.
        """
        if self._schema_version() >= SCHEMA_VERSION:
            return

        cur = self.conn.cursor()
        now = utc_now_str()
        # Rename live records in place, preserving task_id, dynamic values and history.
        for legacy_name, canonical_name in WORKFLOW_TASK_RENAMES.items():
            template = cur.execute("SELECT template_id FROM task_templates WHERE task_name = ?", (legacy_name,)).fetchone()
            current = cur.execute("SELECT template_id FROM task_templates WHERE task_name = ?", (canonical_name,)).fetchone()
            if template and not current:
                cur.execute("UPDATE task_templates SET task_name = ? WHERE template_id = ?", (canonical_name, template["template_id"]))
            elif template and current:
                # Keep the legacy template as an archived definition rather than deleting it.
                cur.execute("UPDATE task_templates SET task_name = ? WHERE template_id = ?", (legacy_name + " (Legacy)", template["template_id"]))

            conflicts = cur.execute("""
                SELECT project_id FROM project_tasks
                WHERE task_name = ?
                  AND project_id IN (SELECT project_id FROM project_tasks WHERE task_name = ?)
            """, (legacy_name, canonical_name)).fetchall()
            conflict_ids = {row["project_id"] for row in conflicts}
            if conflict_ids:
                placeholders = ",".join("?" for _ in conflict_ids)
                cur.execute(
                    f"UPDATE project_tasks SET is_active = 0, last_updated = ? WHERE task_name = ? AND project_id IN ({placeholders})",
                    [now, legacy_name, *conflict_ids],
                )
            cur.execute("""
                UPDATE project_tasks
                SET task_name = ?, last_updated = COALESCE(last_updated, ?)
                WHERE task_name = ? AND is_active = 1
            """, (canonical_name, now, legacy_name))

        desired_names = {tpl[1] for tpl in PIPELINE_TEMPLATES}
        template_map = {}
        for sequence_no, tpl in enumerate(PIPELINE_TEMPLATES, start=1):
            preferred_id, task_name, stage_group, default_role, duration, _depends_on, _branch_type, output = tpl
            existing = cur.execute("SELECT template_id FROM task_templates WHERE task_name = ?", (task_name,)).fetchone()
            if existing:
                template_id = existing["template_id"]
                cur.execute("""
                    UPDATE task_templates
                    SET sequence_no = ?, stage_group = ?, default_role = ?, default_duration_days = ?,
                        depends_on_template_id = NULL, branch_type = 'normal', mandatory_output = ?
                    WHERE template_id = ?
                """, (sequence_no, stage_group, default_role, duration, output, template_id))
            else:
                occupied = cur.execute("SELECT 1 FROM task_templates WHERE template_id = ?", (preferred_id,)).fetchone()
                template_id = preferred_id if not occupied else cur.execute(
                    "SELECT COALESCE(MAX(template_id), 0) + 1 AS next_id FROM task_templates"
                ).fetchone()["next_id"]
                cur.execute("""
                    INSERT INTO task_templates (template_id, sequence_no, task_name, stage_group, default_role,
                                                default_duration_days, depends_on_template_id, branch_type, mandatory_output)
                    VALUES (?, ?, ?, ?, ?, ?, NULL, 'normal', ?)
                """, (template_id, sequence_no, task_name, stage_group, default_role, duration, output))
            template_map[task_name] = (template_id, sequence_no, stage_group, default_role, duration, output)
            cur.execute("""
                UPDATE project_tasks
                SET template_id = ?, sequence_no = ?, stage_group = ?, is_active = 1,
                    last_updated = COALESCE(last_updated, ?)
                WHERE task_name = ?
            """, (template_id, sequence_no, stage_group, now, task_name))

        # Templates not in the current workflow remain discoverable but are no longer active.
        if desired_names:
            placeholders = ",".join("?" for _ in desired_names)
            cur.execute(f"UPDATE project_tasks SET is_active = 0 WHERE task_name NOT IN ({placeholders})", list(desired_names))

        # Backfill only genuinely missing active tasks. Historic inputs are never overwritten.
        projects = cur.execute("SELECT project_id, pipeline_type FROM projects").fetchall()
        for project in projects:
            project_id = project["project_id"]
            existing_names = {r["task_name"] for r in cur.execute(
                "SELECT task_name FROM project_tasks WHERE project_id = ?", (project_id,)
            ).fetchall()}
            pipeline = str(project["pipeline_type"] or "prospect").lower()
            for task_name, values in template_map.items():
                if task_name in existing_names:
                    continue
                template_id, sequence_no, stage_group, _role, duration, output = values
                initial_status = "Not Applicable" if pipeline == "bp" and stage_group in PROSPECT_STAGES else "Not Assigned"
                cur.execute("""
                    INSERT INTO project_tasks (
                        project_id, template_id, sequence_no, task_name, stage_group, status,
                        planned_start, planned_finish, output_notes, priority, is_active, last_updated
                    ) VALUES (?, ?, ?, ?, ?, ?, NULL, NULL, ?, 'Medium', 1, ?)
                """, (project_id, template_id, sequence_no, task_name, stage_group, initial_status, output, now))

        self.normalize_legacy_statuses(commit=False)
        self.normalize_pipeline_types(commit=False)
        # v15: migrate existing Presence CoS values to the automatic Reservoir × Trap × Seal calculation.
        for project in projects:
            self.recalculate_presence_cos(project["project_id"], "System Migration")
        self._set_schema_version(SCHEMA_VERSION)
        self.conn.commit()

    def cleanup_duplicate_project_tasks(self):
        """Remove duplicate per-well tasks created by older migration attempts, keeping the lowest task_id."""
        duplicate_groups = self.conn.execute("""
            SELECT project_id, task_name, COUNT(*) AS c, MIN(task_id) AS keep_id
            FROM project_tasks
            GROUP BY project_id, task_name
            HAVING COUNT(*) > 1
        """).fetchall()
        for group in duplicate_groups:
            self.conn.execute("""
                DELETE FROM project_tasks
                WHERE project_id = ? AND task_name = ? AND task_id != ?
            """, (group["project_id"], group["task_name"], group["keep_id"]))
        self.conn.commit()

    def add_project(self, project_name, start_date=None, target_date=None, changed_by="System", lead_x=None, lead_y=None,
                    business_plan_year=None, business_plan_enabled=False, active_well_enabled=False, pipeline_type="prospect"):
        project_name = (project_name or '').strip()
        if not project_name:
            raise ValueError("Lead / well name is required.")
        if len(project_name) > 120:
            raise ValueError("Lead / well name must be 120 characters or less.")
        pipeline_type = str(pipeline_type or "prospect").strip().lower()
        if pipeline_type not in {"prospect", "bp"}:
            pipeline_type = "prospect"
        cur = self.conn.cursor()
        now = utc_now_str()
        start_date = start_date or today_str()
        target_date = target_date or ""
        year_val = int(business_plan_year) if business_plan_year else None
        bp_enabled = 1 if business_plan_enabled or year_val else 0
        if bp_enabled and (year_val is None or year_val < 2026 or year_val > 2040):
            raise ValueError("Select a business plan year from 2026 to 2040.")

        templates = list(self.get_templates())
        if not templates:
            raise RuntimeError("Workflow templates are not available.")
        first_template = next((t for t in templates if t["stage_group"] in BP_EXECUTION_STAGES), templates[0]) if pipeline_type == "bp" else templates[0]
        cur.execute("""
            INSERT INTO projects (
                project_name, overall_status, current_stage, current_task, current_owner,
                drill_result, start_date, target_date, current_stage_started_at, last_updated,
                lead_folder_path, lead_x, lead_y, business_plan_enabled, business_plan_year,
                active_well_enabled, pipeline_type
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            project_name, "In Progress", first_template["stage_group"], first_template["task_name"], None,
            None, start_date, target_date, start_date, now,
            self.default_lead_folder_path(project_name), lead_x or None, lead_y or None,
            bp_enabled, year_val, 1 if active_well_enabled else 0, pipeline_type
        ))
        project_id = cur.lastrowid
        try:
            start_dt = date.fromisoformat(start_date)
        except Exception:
            start_dt = date.today()
        first_task_id = None
        first_sequence = first_template["sequence_no"]
        for row in templates:
            is_bp_stage = row["stage_group"] in BP_EXECUTION_STAGES
            if pipeline_type == "bp" and not is_bp_stage:
                initial_status = "Not Applicable"
            elif pipeline_type == "prospect" and is_bp_stage:
                initial_status = "Not Assigned"
            else:
                initial_status = "Assigned" if row["sequence_no"] == first_sequence else "Not Assigned"
            planned_start = start_dt.isoformat() if row["sequence_no"] == first_sequence else None
            planned_finish = ((start_dt + timedelta(days=row["default_duration_days"])).isoformat()
                              if row["sequence_no"] == first_sequence else None)
            cur.execute("""
                INSERT INTO project_tasks (
                    project_id, template_id, sequence_no, task_name, stage_group, assigned_to,
                    backup_owner, approver, status, planned_start, actual_start, planned_finish,
                    actual_finish, output_notes, comments, priority, business_plan_enabled,
                    business_plan_year, is_active, last_updated
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
            """, (
                project_id, row["template_id"], row["sequence_no"], row["task_name"], row["stage_group"],
                None, None, None, initial_status, planned_start, None, planned_finish, None,
                row["mandatory_output"], None, "Medium", bp_enabled, year_val, now
            ))
            if row["sequence_no"] == first_sequence:
                first_task_id = cur.lastrowid

        cur.execute("INSERT OR IGNORE INTO project_overview (project_id, last_updated) VALUES (?, ?)", (project_id, now))
        if first_task_id is not None:
            action = "Well Added to BP" if pipeline_type == "bp" else "Lead Created"
            comment = f"{'Well added to Business Plan Execution' if pipeline_type == 'bp' else 'Lead created'}: {project_name}"
            self.log_task_event(
                task_id=first_task_id,
                project_id=project_id,
                task_name=first_template["task_name"],
                action_type=action,
                old_status=None,
                new_status="Created",
                changed_by=changed_by,
                comment=comment,
            )
        self.refresh_project_state(project_id)
        self.conn.commit()
        return project_id

    def get_templates(self):
        return self.conn.execute("SELECT * FROM task_templates ORDER BY sequence_no").fetchall()

    def get_projects(self, search_text="", stage_filter="All", status_filter="All",
                     owner_filter="All", health_filter="All", sort_key="Well Name", pipeline_filter="All"):
        conditions = ["COALESCE(p.archived, 0) = 0"]
        params = []
        needle = (search_text or "").strip().lower()
        if needle:
            conditions.append("LOWER(COALESCE(p.project_name, '')) LIKE ?")
            params.append(f"%{needle}%")
        if stage_filter != "All":
            conditions.append("p.current_stage = ?")
            params.append(stage_filter)
        if status_filter != "All":
            conditions.append("p.overall_status = ?")
            params.append(status_filter)
        if owner_filter != "All":
            conditions.append("p.current_owner = ?")
            params.append(owner_filter)
        if pipeline_filter in {"prospect", "bp"}:
            conditions.append("LOWER(COALESCE(p.pipeline_type, 'prospect')) = ?")
            params.append(pipeline_filter)
        where_clause = " AND ".join(conditions)
        rows = self.conn.execute(f"""
            SELECT p.*,
                   COALESCE(pt_current.priority, 'Medium') AS current_task_priority,
                   COALESCE(priority_flags.has_high_priority_tasks, 0) AS has_high_priority_tasks,
                   CASE WHEN p.current_stage = 'Post-Drilling'
                             AND LOWER(COALESCE(active_drilling.field_value, '')) IN ('1', 'true', 'yes', 'on')
                        THEN 1 ELSE 0 END AS active_drilling
            FROM projects p
            LEFT JOIN project_tasks pt_current
              ON pt_current.project_id = p.project_id
             AND pt_current.task_name = p.current_task
             AND pt_current.is_active = 1
            LEFT JOIN (
                SELECT project_id,
                       MAX(CASE WHEN priority = 'High' THEN 1 ELSE 0 END) AS has_high_priority_tasks
                FROM project_tasks
                WHERE is_active = 1
                GROUP BY project_id
            ) priority_flags ON priority_flags.project_id = p.project_id
            LEFT JOIN (
                SELECT pt.project_id, tdf.field_value
                FROM project_tasks pt
                JOIN task_dynamic_fields tdf ON tdf.task_id = pt.task_id
                WHERE pt.task_name IN ('Quicklook Logs Interpretation', 'Quicklook Logs')
                  AND tdf.field_key = 'active_drilling'
            ) active_drilling ON active_drilling.project_id = p.project_id
            WHERE {where_clause}
            ORDER BY p.project_id DESC
        """, params).fetchall()
        filtered = []
        for row in rows:
            item = dict(row)
            item["active_well_enabled"] = int(item.get("active_well_enabled") or 0)
            item["health"] = health_from_target(item.get("target_date"), item.get("overall_status"))
            if health_filter != "All" and item["health"] != health_filter:
                continue
            filtered.append(item)

        def sort_fn(item):
            if sort_key == "Well Name":
                return (item.get("project_name") or "").lower()
            if sort_key == "Date Created":
                return -(item.get("project_id") or 0)
            if sort_key == "Stage":
                return STAGE_ORDER.index(item["current_stage"]) if item.get("current_stage") in STAGE_ORDER else 999
            if sort_key == "Assignee":
                return (item.get("current_owner") or "").lower()
            if sort_key == "Health":
                return {"Overdue": 0, "Due Soon": 1, "On Track": 2, "Completed": 3}.get(item["health"], 99)
            return parse_iso_date(item.get("target_date")) or date.max
        filtered.sort(key=sort_fn)
        return filtered

    def get_projects_by_stage(self, search_text=""):
        grouped = {stage: [] for stage in STAGE_ORDER}
        for row in self.get_projects(search_text=search_text):
            stage = row["current_stage"] or "Closure"
            grouped.setdefault(stage, []).append(row)
        return grouped

    def get_project(self, project_id):
        row = self.conn.execute("SELECT * FROM projects WHERE project_id = ?", (project_id,)).fetchone()
        if not row:
            return None
        project = dict(row)
        if not project.get("lead_folder_path"):
            project["lead_folder_path"] = self.default_lead_folder_path(project.get("project_name") or "")
        return project

    def default_lead_folder_path(self, project_name):
        field_name, well_name = parse_field_and_well(project_name or "")
        return _windows_join(WINDOWS_WELL_SHARE_ROOT, field_name, well_name, WELL_OVERVIEW_DIRECTORY_MAP.get("lead", "Leads"))

    def update_project_lead_folder(self, project_id, lead_folder_path):
        lead_folder_path = (lead_folder_path or "").strip()
        if not lead_folder_path:
            project = self.get_project(project_id)
            if not project:
                raise ValueError("Well not found.")
            lead_folder_path = self.default_lead_folder_path(project.get("project_name") or "")
        self.conn.execute(
            "UPDATE projects SET lead_folder_path = ?, last_updated = ? WHERE project_id = ?",
            (lead_folder_path, utc_now_str(), project_id)
        )
        self.conn.commit()
        return lead_folder_path

    def get_project_tasks(self, project_id):
        rows = self.conn.execute("""
            SELECT * FROM project_tasks
            WHERE project_id = ? AND is_active = 1
            ORDER BY sequence_no
        """, (project_id,)).fetchall()
        return [dict(r) for r in rows]

    def get_task(self, task_id):
        row = self.conn.execute("SELECT * FROM project_tasks WHERE task_id = ?", (task_id,)).fetchone()
        return dict(row) if row else None

    def project_completion_percent(self, project_id):
        row = self.conn.execute("""
            SELECT
                SUM(CASE WHEN status != 'Not Applicable' THEN 1 ELSE 0 END) AS applicable_total,
                SUM(CASE WHEN status IN ('Approved', 'Complete') THEN 1 ELSE 0 END) AS done
            FROM project_tasks
            WHERE project_id = ? AND is_active = 1
        """, (project_id,)).fetchone()
        total = int(row["applicable_total"] or 0)
        done = int(row["done"] or 0)
        return round((done / total) * 100, 1) if total else 0.0

    def update_project_name(self, project_id, new_name, changed_by="Admin", lead_x=None, lead_y=None, business_plan_year=None, business_plan_enabled=None, active_well_enabled=None):
        new_name = (new_name or "").strip()
        if not new_name:
            raise ValueError("Lead / well name is required.")
        if len(new_name) > 120:
            raise ValueError("Lead / well name must be 120 characters or less.")
        old = self.get_project(project_id)
        if not old:
            raise ValueError("Lead / well not found.")
        updates = {"project_name": new_name, "last_updated": utc_now_str()}
        old_default_folder = self.default_lead_folder_path(old.get("project_name") or "")
        if not old.get("lead_folder_path") or old.get("lead_folder_path") == old_default_folder:
            updates["lead_folder_path"] = self.default_lead_folder_path(new_name)
        if lead_x is not None:
            updates["lead_x"] = lead_x or None
        if lead_y is not None:
            updates["lead_y"] = lead_y or None
        if business_plan_enabled is not None:
            updates["business_plan_enabled"] = 1 if business_plan_enabled else 0
        if active_well_enabled is not None:
            updates["active_well_enabled"] = 1 if active_well_enabled else 0
        if business_plan_year is not None and str(business_plan_year).strip():
            y = int(business_plan_year)
            if y < 2026 or y > 2040:
                raise ValueError("Select a business plan year from 2026 to 2040.")
            updates["business_plan_year"] = y
        assignments = ", ".join([f"{k} = ?" for k in updates])
        self.conn.execute(f"UPDATE projects SET {assignments} WHERE project_id = ?", [*updates.values(), project_id])
        # Keep the optional mounted server folder aligned when it is available.
        # UNC links are always regenerated from the current record name.
        try:
            old_field, old_well = parse_field_and_well(old.get("project_name") or "")
            new_field, new_well = parse_field_and_well(new_name)
            for root in (WELL_OVERVIEW_DIRECTORY_ROOT, LEAD_WORKFLOW_DIRECTORY_ROOT):
                old_path = root / old_field / old_well
                new_path = root / new_field / new_well
                if old_path.exists() and not new_path.exists():
                    new_path.parent.mkdir(parents=True, exist_ok=True)
                    old_path.rename(new_path)
        except Exception:
            # Folder links must not prevent a record rename when the share is not mounted.
            pass
        first_task = self.conn.execute("SELECT task_id, task_name FROM project_tasks WHERE project_id = ? ORDER BY sequence_no LIMIT 1", (project_id,)).fetchone()
        if first_task:
            record_type = "Well" if str((old or {}).get("pipeline_type") or "prospect").lower() == "bp" else "Lead"
            self.log_task_event(first_task["task_id"], project_id, first_task["task_name"], f"{record_type} Renamed", old.get("project_name") if old else None, new_name, changed_by, f"Renamed {record_type.lower()} to {new_name}")
        self.conn.commit()

    def archive_project(self, project_id, changed_by="Admin", *args, **kwargs):
        project = self.get_project(project_id)
        if not project:
            raise ValueError("Lead / well not found.")
        if int(project.get("archived") or 0):
            return
        with self.write_transaction():
            self.conn.execute(
                "UPDATE projects SET archived = 1, last_updated = ?, revision = COALESCE(revision, 0) + 1 WHERE project_id = ?",
                (utc_now_str(), project_id),
            )
            first_task = self.conn.execute(
                "SELECT task_id, task_name FROM project_tasks WHERE project_id = ? ORDER BY sequence_no LIMIT 1",
                (project_id,),
            ).fetchone()
            if first_task:
                self.log_task_event(
                    first_task["task_id"], project_id, first_task["task_name"], "Well Archived", None, "Archived",
                    changed_by, f"Archived well: {project.get('project_name') or project_id}"
                )

    def restore_project(self, project_id, changed_by="Admin"):
        project = self.get_project(project_id)
        if not project:
            raise ValueError("Lead / well not found.")
        if not int(project.get("archived") or 0):
            return
        with self.write_transaction():
            self.conn.execute(
                "UPDATE projects SET archived = 0, last_updated = ?, revision = COALESCE(revision, 0) + 1 WHERE project_id = ?",
                (utc_now_str(), project_id),
            )
            first_task = self.conn.execute(
                "SELECT task_id, task_name FROM project_tasks WHERE project_id = ? ORDER BY sequence_no LIMIT 1",
                (project_id,),
            ).fetchone()
            if first_task:
                self.log_task_event(
                    first_task["task_id"], project_id, first_task["task_name"], "Well Restored", "Archived", "Active",
                    changed_by, f"Restored well: {project.get('project_name') or project_id}"
                )

    def delete_project(self, project_id, changed_by="Admin"):
        """Permanent deletion is reserved for controlled maintenance only; web routes archive records."""
        with self.write_transaction():
            self.conn.execute("DELETE FROM projects WHERE project_id = ?", (project_id,))

    def get_first_open_task(self, project_id):
        # Ensure the workflow is reconciled before asking the UI to move to the next task.
        self.reconcile_project_flow(project_id)
        row = self.conn.execute("""
            SELECT * FROM project_tasks
            WHERE project_id = ? AND is_active = 1 AND status IN ('Assigned','In Progress','Ready for Review','Under Review','Ready for Approval','Returned for Update')
            ORDER BY sequence_no
            LIMIT 1
        """, (project_id,)).fetchone()
        return dict(row) if row else None

    def reconcile_project_flow(self, project_id):
        """Return the current open task without enforcing dependencies or locks.

        If no task is explicitly In Progress/Waiting, continue from the current
        task position instead of jumping back to the first incomplete item.
        """
        project = self.get_project(project_id) or {}
        pipeline_type = str(project.get("pipeline_type") or "prospect").lower()
        applicable_stages = BP_EXECUTION_STAGES if pipeline_type == "bp" else PROSPECT_STAGES
        stage_filter_sql = " AND stage_group IN ({})".format(",".join("?" for _ in applicable_stages))
        stage_params = applicable_stages

        row = self.conn.execute(f"""
            SELECT * FROM project_tasks
            WHERE project_id = ? AND is_active = 1 AND status IN ('Assigned','In Progress','Ready for Review','Under Review','Ready for Approval','Returned for Update')
            {stage_filter_sql}
            ORDER BY sequence_no
            LIMIT 1
        """, [project_id, *stage_params]).fetchone()
        if row:
            return dict(row)

        current_task = project.get("current_task")
        current_seq = 0
        if current_task:
            seq_row = self.conn.execute("""
                SELECT sequence_no FROM project_tasks
                WHERE project_id = ? AND task_name = ? AND is_active = 1
                LIMIT 1
            """, (project_id, current_task)).fetchone()
            current_seq = seq_row["sequence_no"] if seq_row else 0

        row = self.conn.execute(f"""
            SELECT * FROM project_tasks
            WHERE project_id = ? AND is_active = 1 AND status NOT IN ('Approved','Not Applicable','Complete') AND sequence_no > ?
            {stage_filter_sql}
            ORDER BY sequence_no
            LIMIT 1
        """, [project_id, current_seq, *stage_params]).fetchone()
        if row:
            return dict(row)

        row = self.conn.execute(f"""
            SELECT * FROM project_tasks
            WHERE project_id = ? AND is_active = 1 AND status NOT IN ('Approved','Not Applicable','Complete')
            {stage_filter_sql}
            ORDER BY sequence_no
            LIMIT 1
        """, [project_id, *stage_params]).fetchone()
        return dict(row) if row else None

    def get_distinct_owners(self):
        rows = self.conn.execute("""
            SELECT DISTINCT assigned_to AS owner_name FROM project_tasks WHERE assigned_to IS NOT NULL AND TRIM(assigned_to) != ''
            UNION
            SELECT DISTINCT current_owner AS owner_name FROM projects WHERE current_owner IS NOT NULL AND TRIM(current_owner) != ''
            ORDER BY owner_name
        """).fetchall()
        return [r["owner_name"] for r in rows]

    def get_project_overview(self, project_id):
        self.conn.execute("INSERT OR IGNORE INTO project_overview (project_id, last_updated) VALUES (?, ?)", (project_id, utc_now_str()))
        self.conn.commit()
        row = self.conn.execute("SELECT * FROM project_overview WHERE project_id = ?", (project_id,)).fetchone()
        return dict(row) if row else {}

    def update_project_overview_fields(self, project_id, fields):
        allowed = {
            "derisking", "ogip", "lead_ogip", "preliminary_resource_estimation", "pre_drill_estimation",
            "post_drill_estimation", "reservoir_pressure", "reservoir_gradient",
            "flowback_results", "pay", "porosity", "swt",
            "quick_look_pay", "quick_look_porosity", "quick_look_swt"
        }
        clean = {k: (v.strip() if isinstance(v, str) else v) for k, v in (fields or {}).items() if k in allowed}
        if not clean:
            return
        self.conn.execute("INSERT OR IGNORE INTO project_overview (project_id, last_updated) VALUES (?, ?)", (project_id, utc_now_str()))
        assignments = ", ".join([f"{k} = ?" for k in clean]) + ", last_updated = ?"
        self.conn.execute(f"UPDATE project_overview SET {assignments} WHERE project_id = ?", [*clean.values(), utc_now_str(), project_id])

    def get_well_overview_rows(self):
        rows = self.conn.execute("""
            SELECT p.project_id, p.project_name, p.current_stage, p.current_task, p.overall_status,
                   COALESCE(o.derisking, '') AS derisking,
                   COALESCE(NULLIF(o.lead_ogip, ''), NULLIF(o.preliminary_resource_estimation, ''), NULLIF(o.ogip, ''), '') AS lead_ogip,
                   COALESCE(NULLIF(o.preliminary_resource_estimation, ''), '') AS preliminary_resource_estimation,
                   COALESCE(o.pre_drill_estimation, '') AS pre_drill_estimation,
                   COALESCE(o.post_drill_estimation, '') AS post_drill_estimation,
                   COALESCE(NULLIF(o.post_drill_estimation, ''), NULLIF(o.pre_drill_estimation, ''), NULLIF(o.lead_ogip, ''), NULLIF(o.preliminary_resource_estimation, ''), NULLIF(o.ogip, ''), '') AS business_plan_ogip,
                   COALESCE(o.reservoir_pressure, '') AS reservoir_pressure,
                   COALESCE(o.reservoir_gradient, '') AS reservoir_gradient,
                   COALESCE(o.flowback_results, '') AS flowback_results, COALESCE(o.pay, '') AS pay,
                   COALESCE(o.porosity, '') AS porosity, COALESCE(o.swt, '') AS swt,
                   COALESCE(o.quick_look_pay, '') AS quick_look_pay,
                   COALESCE(o.quick_look_porosity, '') AS quick_look_porosity,
                   COALESCE(o.quick_look_swt, '') AS quick_look_swt,
                   MAX(CASE WHEN pt.task_name = 'Prospect Evaluation Presentation' AND pt.status IN ('Approved','Complete') THEN 1 ELSE 0 END) AS segment_done,
                   MAX(CASE WHEN pt.task_name = 'PDA' AND pt.status IN ('Approved','Complete') THEN 1 ELSE 0 END) AS pda_done,
                   MAX(CASE WHEN pt.task_name = 'PVAD Structural MTR' AND pt.status IN ('Approved','Complete') THEN 1 ELSE 0 END) AS mtr_done
            FROM projects p
            LEFT JOIN project_overview o ON o.project_id = p.project_id
            LEFT JOIN project_tasks pt ON pt.project_id = p.project_id AND pt.is_active = 1
            WHERE COALESCE(p.archived, 0) = 0
            GROUP BY p.project_id
            ORDER BY p.project_name COLLATE NOCASE
        """).fetchall()
        result = []
        for r in rows:
            item = dict(r)
            item["segment_class"] = self.segment_class(item.get("business_plan_ogip"), item.get("derisking"))
            result.append(item)
        return result

    def get_lead_summary_snapshot(self, project_id: int):
        row = self.conn.execute("SELECT snapshot_json, captured_at, captured_by FROM lead_summary_snapshots WHERE project_id = ?", (project_id,)).fetchone()
        if not row:
            return None
        try:
            fields = json.loads(row["snapshot_json"] or "{}")
        except json.JSONDecodeError:
            fields = {}
        return {"fields": fields, "captured_at": row["captured_at"], "captured_by": row["captured_by"]}

    def _capture_lead_summary_snapshot(self, project_id: int, changed_by: str):
        """Capture the Lead Summary immediately before promotion to BP Execution.

        Re-promotion refreshes the snapshot so the BP Well always carries the
        current Lead Summary that was moved with it.
        """
        rows = self.conn.execute("""
            SELECT pt.task_name, tdf.field_key, tdf.field_value
            FROM project_tasks pt
            LEFT JOIN task_dynamic_fields tdf ON tdf.task_id = pt.task_id
            WHERE pt.project_id = ? AND pt.stage_group IN ({})
            ORDER BY pt.sequence_no, tdf.field_key
        """.format(",".join("?" for _ in PROSPECT_STAGES)), [project_id, *PROSPECT_STAGES]).fetchall()
        fields = {}
        for row in rows:
            fields.setdefault(row["task_name"], {})
            if row["field_key"]:
                fields[row["task_name"]][row["field_key"]] = row["field_value"] or ""
        self.conn.execute("""
            INSERT INTO lead_summary_snapshots(project_id, snapshot_json, captured_at, captured_by)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(project_id) DO UPDATE SET
                snapshot_json = excluded.snapshot_json,
                captured_at = excluded.captured_at,
                captured_by = excluded.captured_by
        """, (project_id, json.dumps(fields, separators=(",", ":")), utc_now_str(), changed_by))

    def _move_lead_to_bp_execution(self, project_id: int, year_val: int, changed_by: str):
        """Promote a matured lead into the BP Execution pipeline without losing its lead record.

        Prospect tasks and inputs remain intact for the captured Lead Summary. BP tasks become
        operational; the first BP component is assigned when no BP component is already active.
        """
        project = self.get_project(project_id)
        if not project:
            raise ValueError("Lead / well not found.")
        if str(project.get("pipeline_type") or "prospect").lower() != "bp":
            self._capture_lead_summary_snapshot(project_id, changed_by)

        bp_tasks = self.conn.execute("""
            SELECT * FROM project_tasks
            WHERE project_id = ? AND stage_group IN ({}) AND is_active = 1
            ORDER BY sequence_no
        """.format(",".join("?" for _ in BP_EXECUTION_STAGES)), [project_id, *BP_EXECUTION_STAGES]).fetchall()
        if not bp_tasks:
            raise RuntimeError("Business Plan workflow is not available for this lead.")
        now = utc_now_str()
        has_existing_bp_work = any((task["status"] or "") not in {"Not Applicable", "Not Assigned"} for task in bp_tasks)
        for index, task in enumerate(bp_tasks):
            old_status = task["status"] or "Not Assigned"
            # A promoted prospect normally has BP tasks in Not Assigned state. Activate
            # the first BP task exactly once; leave any existing BP progress untouched.
            should_activate_first = index == 0 and not has_existing_bp_work and old_status in {"Not Applicable", "Not Assigned"}
            if old_status == "Not Applicable" or should_activate_first:
                next_status = "Assigned" if should_activate_first else "Not Assigned"
                self.conn.execute("""
                    UPDATE project_tasks
                    SET status = ?, business_plan_enabled = 1, business_plan_year = ?, last_updated = ?, revision = COALESCE(revision, 0) + 1
                    WHERE task_id = ?
                """, (next_status, year_val, now, task["task_id"]))
            else:
                self.conn.execute("""
                    UPDATE project_tasks
                    SET business_plan_enabled = 1, business_plan_year = ?, last_updated = ?
                    WHERE task_id = ?
                """, (year_val, now, task["task_id"]))

        first_open = self.conn.execute("""
            SELECT * FROM project_tasks
            WHERE project_id = ? AND stage_group IN ({}) AND is_active = 1
              AND status NOT IN ('Approved','Not Applicable','Complete')
            ORDER BY sequence_no LIMIT 1
        """.format(",".join("?" for _ in BP_EXECUTION_STAGES)), [project_id, *BP_EXECUTION_STAGES]).fetchone()
        if not first_open:
            first_open = bp_tasks[0]
        self.conn.execute("""
            UPDATE projects
            SET pipeline_type = 'bp', business_plan_enabled = 1, business_plan_year = ?,
                current_stage = ?, current_task = ?, current_owner = ?, current_stage_started_at = ?,
                overall_status = 'In Progress', last_updated = ?, revision = COALESCE(revision, 0) + 1
            WHERE project_id = ?
        """, (year_val, first_open["stage_group"], first_open["task_name"], first_open["assigned_to"], today_str(), now, project_id))

    def _move_bp_to_lead_phase(self, project_id: int, changed_by: str):
        """Return a promoted BP Well to Prospect Maturation without data loss.

        Business Plan tasks, inputs, the Lead Summary snapshot and history stay
        intact for a future promotion. Only the operating pipeline changes.
        """
        project = self.get_project(project_id)
        if not project:
            raise ValueError("Lead / well not found.")
        now = utc_now_str()
        self.conn.execute("""
            UPDATE projects
            SET pipeline_type = 'prospect', business_plan_enabled = 0,
                business_plan_year = NULL, last_updated = ?,
                revision = COALESCE(revision, 0) + 1
            WHERE project_id = ?
        """, (now, project_id))
        self.conn.execute("""
            UPDATE project_tasks
            SET business_plan_enabled = 0, business_plan_year = NULL,
                last_updated = ?
            WHERE project_id = ?
        """, (now, project_id))
        lead_open = self.conn.execute("""
            SELECT * FROM project_tasks
            WHERE project_id = ? AND stage_group IN ({}) AND is_active = 1
              AND status NOT IN ('Approved','Not Applicable','Complete')
            ORDER BY sequence_no LIMIT 1
        """.format(",".join("?" for _ in PROSPECT_STAGES)), [project_id, *PROSPECT_STAGES]).fetchone()
        if not lead_open:
            lead_open = self.conn.execute("""
                SELECT * FROM project_tasks
                WHERE project_id = ? AND stage_group IN ({}) AND is_active = 1
                ORDER BY sequence_no LIMIT 1
            """.format(",".join("?" for _ in PROSPECT_STAGES)), [project_id, *PROSPECT_STAGES]).fetchone()
        if lead_open:
            self.conn.execute("""
                UPDATE projects
                SET current_stage = ?, current_task = ?, current_owner = ?,
                    overall_status = 'In Progress', current_stage_started_at = ?,
                    last_updated = ?
                WHERE project_id = ?
            """, (lead_open['stage_group'], lead_open['task_name'], lead_open['assigned_to'], today_str(), now, project_id))

    def set_business_plan(self, project_id, enabled, year=None, changed_by="Admin", *args, **kwargs):
        old = self.get_project(project_id)
        if not old:
            raise ValueError("Lead / well not found.")
        enabled_int = 1 if enabled else 0
        year_val = None
        if enabled_int:
            year_val = int(year or old.get("business_plan_year") or 0)
            if year_val < 2026 or year_val > 2040:
                raise ValueError("Select a business plan year from 2026 to 2040.")
        with self.write_transaction():
            if enabled_int:
                self._move_lead_to_bp_execution(project_id, year_val, changed_by)
            else:
                # Removing Business Plan returns the record to the Lead pipeline.
                # No BP values, task data, Lead Summary snapshot, or history is deleted.
                self._move_bp_to_lead_phase(project_id, changed_by)
            first_task = self.conn.execute("SELECT task_id, task_name FROM project_tasks WHERE project_id = ? ORDER BY sequence_no LIMIT 1", (project_id,)).fetchone()
            if first_task:
                old_state = f"{old.get('business_plan_enabled') or 0}/{old.get('business_plan_year') or '-'}"
                new_state = f"{enabled_int}/{year_val or '-'}"
                action = "Lead Promoted to BP Execution" if enabled_int and str(old.get("pipeline_type") or "prospect").lower() != "bp" else ("Well Added to BP" if enabled_int else "BP Well Returned to Lead Phase")
                self.log_task_event(first_task["task_id"], project_id, first_task["task_name"], action, old_state, new_state, changed_by, "Business plan assignment updated.")
            self.refresh_project_state(project_id)

    def get_business_plan_rows(self):
        rows = self.conn.execute("""
            SELECT p.project_id,
                   p.business_plan_year AS year,
                   p.project_name AS well_name,
                   COALESCE(o.pre_drill_estimation, '') AS pre_drill_ogip,
                   COALESCE(o.post_drill_estimation, '') AS post_drill_ogip,
                   COALESCE(o.derisking, '') AS chance_of_success,
                   COALESCE(p.active_well_enabled, 0) AS active_well_enabled
            FROM projects p
            LEFT JOIN project_overview o ON o.project_id = p.project_id
            WHERE COALESCE(p.archived, 0) = 0 AND COALESCE(p.business_plan_enabled, 0) = 1
            ORDER BY p.business_plan_year, p.project_name COLLATE NOCASE
        """).fetchall()
        result = []
        for r in rows:
            item = dict(r)
            # Classification should be blank if required inputs are incomplete.
            class_ogip = item.get("post_drill_ogip") or item.get("pre_drill_ogip")
            item["segment_class"] = self.segment_class(class_ogip, item.get("chance_of_success"))
            result.append(item)
        return result

    def get_portfolio_rows(self, year="All", activity="All"):
        """Return the Portfolio view for BP-enabled wells only.

        This reporting method uses the same canonical values and class logic as the
        business-plan scorecard.  Filters are applied here so the cumulative OGIP
        always represents the currently displayed portfolio rows.
        """
        selected_year = str(year or "All").strip()
        selected_activity = str(activity or "All").strip()
        if selected_year != "All":
            try:
                selected_year_int = int(selected_year)
            except (TypeError, ValueError):
                raise ValueError("Select a business plan year from 2026 to 2040.")
            if selected_year_int < 2026 or selected_year_int > 2040:
                raise ValueError("Select a business plan year from 2026 to 2040.")
        else:
            selected_year_int = None

        if selected_activity not in {"All", "Active", "Non-Active"}:
            selected_activity = "All"

        rows = self.get_business_plan_rows()
        filtered = []
        cumulative_ogip = 0.0
        for item in rows:
            if selected_year_int is not None and int(item.get("year") or 0) != selected_year_int:
                continue
            is_active = int(item.get("active_well_enabled") or 0) == 1
            if selected_activity == "Active" and not is_active:
                continue
            if selected_activity == "Non-Active" and is_active:
                continue

            # Latest available OGIP is post-drill; otherwise use pre-drill.
            current_ogip = _to_float_or_none(item.get("post_drill_ogip"))
            if current_ogip is None:
                current_ogip = _to_float_or_none(item.get("pre_drill_ogip"))
            if current_ogip is not None:
                cumulative_ogip += current_ogip
            filtered.append(item)

        return {
            "rows": filtered,
            "summary": {
                "business_plan_wells": len(filtered),
                "cumulative_ogip": round(cumulative_ogip, 1),
            },
        }

    def segment_class(self, ogip_value, chance_value):
        ogip = _to_float_or_none(ogip_value)
        chance = _to_float_or_none(chance_value)
        if ogip is None or chance is None:
            return ""
        high_resource = ogip >= 10.0
        high_chance = chance >= 50.0
        if high_resource and high_chance:
            return "Super Star"
        if high_resource and not high_chance:
            return "Risk Taker"
        if not high_resource and high_chance:
            return "Value Hunter"
        return "Dog"

    def update_project_flags(self, project_id, business_plan_enabled=None, active_well_enabled=None, business_plan_year=None, changed_by="Web User"):
        old = self.get_project(project_id)
        if not old:
            raise ValueError("Lead / well not found.")
        # Promotion is an atomic business operation: capture lead summary, switch pipeline, activate BP tasks.
        if business_plan_enabled is not None:
            requested_year = business_plan_year if business_plan_enabled else None
            self.set_business_plan(project_id, bool(business_plan_enabled), requested_year, changed_by)
        if active_well_enabled is not None:
            with self.write_transaction():
                new_active = 1 if active_well_enabled else 0
                self.conn.execute("UPDATE projects SET active_well_enabled = ?, last_updated = ?, revision = COALESCE(revision, 0) + 1 WHERE project_id = ?", (new_active, utc_now_str(), project_id))
                first_task = self.conn.execute("SELECT task_id, task_name FROM project_tasks WHERE project_id = ? ORDER BY sequence_no LIMIT 1", (project_id,)).fetchone()
                if first_task and new_active != int(old.get("active_well_enabled") or 0):
                    self.log_task_event(first_task["task_id"], project_id, first_task["task_name"], "Active Well Flag", old.get("active_well_enabled") or 0, new_active, changed_by, "Active well flag updated.")

    def get_business_plan_commitment(self):
        self.conn.execute("""
            INSERT OR IGNORE INTO business_plan_commitment (commitment_id, last_updated)
            VALUES (1, ?)
        """, (utc_now_str(),))
        self.conn.commit()
        row = self.conn.execute("SELECT * FROM business_plan_commitment WHERE commitment_id = 1").fetchone()
        return dict(row) if row else {
            "produced": 0,
            "pending_tie_in": 0,
            "base": 0,
            "core_extension_wells": 0,
            "planned_yet_to_find": 0,
        }

    def update_business_plan_commitment(self, fields):
        allowed = {"produced", "pending_tie_in", "base", "core_extension_wells", "planned_yet_to_find"}
        clean = {}
        for key, value in (fields or {}).items():
            if key not in allowed:
                continue
            try:
                text = str(value or "").replace(",", "").strip()
                clean[key] = max(0.0, float(text)) if text else 0.0
            except Exception:
                clean[key] = 0.0
        if not clean:
            return
        self.conn.execute("""
            INSERT OR IGNORE INTO business_plan_commitment (commitment_id, last_updated)
            VALUES (1, ?)
        """, (utc_now_str(),))
        assignments = ", ".join([f"{k} = ?" for k in clean]) + ", last_updated = ?"
        self.conn.execute(
            f"UPDATE business_plan_commitment SET {assignments} WHERE commitment_id = 1",
            [*clean.values(), utc_now_str()]
        )
        self.conn.commit()

    def set_task_priority(self, task_id, priority_value="Medium", changed_by="Admin"):
        task = self.get_task(task_id)
        if not task:
            raise ValueError("Component not found.")
        if isinstance(priority_value, bool):
            new_priority = "High" if priority_value else "Medium"
        else:
            new_priority = str(priority_value or "Medium").strip().title()
        if new_priority not in {"Low", "Medium", "High"}:
            new_priority = "Medium"
        old_priority = task.get("priority") or "Medium"
        if new_priority == old_priority:
            return
        self.conn.execute("""
            UPDATE project_tasks
            SET priority = ?, last_updated = ?
            WHERE task_id = ?
        """, (new_priority, utc_now_str(), task_id))
        self.log_task_event(
            task_id=task_id,
            project_id=task["project_id"],
            task_name=task["task_name"],
            action_type="Priority Update",
            old_status=old_priority,
            new_status=new_priority,
            changed_by=changed_by,
            comment=f"Priority set to {new_priority}.",
        )
        self.conn.commit()

    def update_task(self, task_id, status, assigned_to, comments,
                    changed_by="Web User", enforce_conflict_ts=None, overview_fields=None,
                    business_plan_enabled=None, business_plan_year=None):
        task = self.get_task(task_id)
        if not task:
            raise ValueError("Component not found.")
        # Conflict checks were intentionally removed for this directly-editable web app.
        # Dynamic field saves update last_updated before the main task save, so enforcing
        # a stale timestamp caused false "updated by someone else" errors.
        if status not in STATUSES:
            raise ValueError("Invalid component status.")
        old_status = task["status"]
        old_assigned_to = (task.get("assigned_to") or "").strip()
        old_comments = (task.get("comments") or "").strip()
        actual_start = task["actual_start"]
        actual_finish = task["actual_finish"]
        today = today_str()
        now = utc_now_str()
        if status == "In Progress" and not actual_start:
            actual_start = today
        if status in DONE_STATUSES:
            actual_finish = today
            if not actual_start:
                actual_start = today
        else:
            # Fully reversible status logic. Approved/Not Applicable -> open status clears finish date.
            actual_finish = None
            if status == "Not Assigned":
                actual_start = None

        project = self.get_project(task["project_id"]) or {}
        if business_plan_enabled is None:
            business_plan_enabled = task.get("business_plan_enabled") or project.get("business_plan_enabled") or 0
        enabled_int = 1 if bool(business_plan_enabled) else 0
        year_val = None
        if enabled_int:
            existing_year = business_plan_year or task.get("business_plan_year") or project.get("business_plan_year")
            year_val = int(existing_year) if existing_year else None
            if year_val is None or year_val < 2026 or year_val > 2040:
                raise ValueError("Select a business plan year from 2026 to 2040.")

        self.conn.execute("""
            UPDATE project_tasks
            SET status = ?, assigned_to = ?, comments = ?, actual_start = ?, actual_finish = ?,
                business_plan_enabled = ?, business_plan_year = ?, last_updated = ?
            WHERE task_id = ?
        """, (status, assigned_to.strip() or None, comments.strip() or None, actual_start, actual_finish,
              enabled_int, year_val, now, task_id))
        if enabled_int:
            self.conn.execute("""
                UPDATE projects
                SET business_plan_enabled = 1, business_plan_year = ?, last_updated = ?
                WHERE project_id = ?
            """, (year_val, now, task["project_id"]))

        current = self.get_task(task_id)
        self.update_project_overview_fields(current["project_id"], overview_fields or {})
        new_assigned_to = (assigned_to or "").strip()
        new_comments = (comments or "").strip()
        if status != old_status or new_assigned_to != old_assigned_to or new_comments != old_comments:
            change_note = new_comments or f"Status set to {status}."
            self.log_task_event(task_id, current["project_id"], current["task_name"], "Component Update",
                                old_status, status, changed_by, change_note)
        if status in DONE_STATUSES:
            # Anchor the project on the completed task before recalculating so the
            # next current task advances forward instead of jumping back to an
            # earlier independent Not Started task.
            self.conn.execute("""
                UPDATE projects
                SET current_stage = ?, current_task = ?, current_owner = ?, last_updated = ?
                WHERE project_id = ?
            """, (current["stage_group"], current["task_name"], current["assigned_to"], now, current["project_id"]))
        self.refresh_project_state(current["project_id"])
        self.conn.commit()

    def recalculate_presence_cos(self, project_id, changed_by="System"):
        """Persist the automatic Presence CoS reading for one project.

        The final Reservoir CoS is the last completed row in Reservoir CoS. Presence
        CoS is read-only in the UI and always equals Reservoir × Trap × Seal.
        """
        target = self.conn.execute("""
            SELECT task_id FROM project_tasks
            WHERE project_id = ? AND task_name = 'Presence CoS Evaluation'
            ORDER BY task_id DESC LIMIT 1
        """, (project_id,)).fetchone()
        if not target:
            return {}
        values = calculate_presence_cos(self.conn, project_id)
        now = utc_now_str()
        existing = self.get_task_dynamic_fields(target["task_id"])
        changed = any(str(existing.get(k, "")) != str(v) for k, v in values.items())
        for key, value in values.items():
            self.conn.execute("""
                INSERT INTO task_dynamic_fields (task_id, field_key, field_value, updated_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(task_id, field_key) DO UPDATE SET
                    field_value = excluded.field_value, updated_at = excluded.updated_at
            """, (target["task_id"], key, str(value), now))
        self.update_project_overview_fields(project_id, {"derisking": values.get("presence_cos", "")})
        if changed:
            note = "Automatically recalculated from final Reservoir CoS × Trap CoS × Seal CoS."
            self.log_task_event(target["task_id"], project_id, "Presence CoS Evaluation",
                                "Presence CoS Calculated", None, None, changed_by, note)
        return values

    def get_task_dynamic_fields(self, task_id):
        rows = self.conn.execute("SELECT field_key, field_value FROM task_dynamic_fields WHERE task_id = ?", (task_id,)).fetchall()
        return {r["field_key"]: r["field_value"] for r in rows}

    def save_task_dynamic_fields(self, task_id, fields, changed_by="Web User"):
        task = self.get_task(task_id)
        if not task:
            raise ValueError("Component not found.")
        fields = fields or {}
        if task.get("task_name") == "Seal CoS":
            fields = dict(fields)
            fields["seal_cos_pct"] = calculate_seal_cos(fields)
        now = utc_now_str()
        overview_updates = {}
        changed_keys = []
        for key, value in fields.items():
            val = "" if value is None else str(value).strip()
            existing = self.conn.execute(
                "SELECT field_value FROM task_dynamic_fields WHERE task_id = ? AND field_key = ?",
                (task_id, key),
            ).fetchone()
            old_val = "" if not existing or existing["field_value"] is None else str(existing["field_value"])
            if old_val != val:
                changed_keys.append(key)
            self.conn.execute("""
                INSERT INTO task_dynamic_fields (task_id, field_key, field_value, updated_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(task_id, field_key) DO UPDATE SET field_value = excluded.field_value, updated_at = excluded.updated_at
            """, (task_id, key, val, now))
            if key in DYNAMIC_FIELD_OVERVIEW_MAP:
                overview_updates[DYNAMIC_FIELD_OVERVIEW_MAP[key]] = val
        if overview_updates:
            self.update_project_overview_fields(task["project_id"], overview_updates)
        self.conn.execute("UPDATE project_tasks SET last_updated = ? WHERE task_id = ?", (now, task_id))
        if changed_keys:
            readable = [key.replace("_", " ") for key in changed_keys]
            listed = ", ".join(readable[:8])
            if len(readable) > 8:
                listed += ", and more"
            self.log_task_event(
                task_id, task["project_id"], task["task_name"], "Component Inputs Updated",
                None, None, changed_by, f"Updated inputs: {listed}."
            )
        if task.get("task_name") in {"Reservoir CoS", "Trap CoS", "Seal CoS"}:
            self.recalculate_presence_cos(task["project_id"], changed_by)
        self.conn.commit()

    def save_task(self, task_id, payload, changed_by="Web User"):
        """Save a component atomically, including fields, priority and workflow state.

        `revision` is optional for backward compatibility. When provided, stale edits
        are rejected rather than silently overwriting a newer change.
        """
        payload = payload or {}
        fields = payload.get("fields") if isinstance(payload.get("fields"), dict) else {}
        expected_revision = payload.get("revision")
        status = str(payload.get("status") or "Not Assigned")
        assigned_to = str(payload.get("assigned_to") or "").strip()
        comments = str(payload.get("comments") or "").strip()
        priority = str(payload.get("priority") or "Medium").strip().title()
        if priority not in {"Low", "Medium", "High"}:
            priority = "Medium"
        if status not in STATUSES:
            raise ValueError("Invalid component status.")

        result = {}
        with self.write_transaction():
            task = self.get_task(task_id)
            if not task:
                raise ValueError("Component not found.")
            current_revision = int(task.get("revision") or 0)
            if expected_revision is not None:
                try:
                    if int(expected_revision) != current_revision:
                        raise RuntimeError("This component was updated by someone else. Refresh and review the latest values.")
                except (TypeError, ValueError):
                    raise ValueError("Invalid component revision.")

            old_status = task.get("status") or "Not Assigned"
            old_assigned_to = (task.get("assigned_to") or "").strip()
            old_comments = (task.get("comments") or "").strip()
            old_priority = task.get("priority") or "Medium"
            actual_start = task.get("actual_start")
            actual_finish = task.get("actual_finish")
            today = today_str()
            now = utc_now_str()
            if status == "In Progress" and not actual_start:
                actual_start = today
            if status in DONE_STATUSES:
                actual_finish = today
                if not actual_start:
                    actual_start = today
            else:
                actual_finish = None
                if status == "Not Assigned":
                    actual_start = None

            project = self.get_project(task["project_id"]) or {}
            bp_enabled = payload.get("business_plan_enabled")
            if bp_enabled is None:
                bp_enabled = task.get("business_plan_enabled") or project.get("business_plan_enabled") or 0
            enabled_int = 1 if bool(bp_enabled) else 0
            year_val = None
            if enabled_int:
                selected_year = payload.get("business_plan_year") or task.get("business_plan_year") or project.get("business_plan_year")
                year_val = int(selected_year) if selected_year else None
                if year_val is None or year_val < 2026 or year_val > 2040:
                    raise ValueError("Select a business plan year from 2026 to 2040.")

            # Reservoir CoS is model-derived, not manually keyed. The saved result is a whole-number percent.
            if task.get("task_name") == "Reservoir CoS" and "reservoir_cos_rows" in fields:
                fields = dict(fields)
                fields["reservoir_cos_rows"] = calculate_reservoir_cos_rows(fields.get("reservoir_cos_rows"))

            # Seal CoS is formula-derived, not manually entered. The result is stored
            # as a whole-number percentage string, e.g., 44 for 44%.
            if task.get("task_name") == "Seal CoS":
                fields = dict(fields)
                fields["seal_cos_pct"] = calculate_seal_cos(fields)

            changed_keys = []
            overview_updates = {}
            for key, value in fields.items():
                val = "" if value is None else str(value).strip()
                existing = self.conn.execute(
                    "SELECT field_value FROM task_dynamic_fields WHERE task_id = ? AND field_key = ?",
                    (task_id, key),
                ).fetchone()
                old_val = "" if not existing or existing["field_value"] is None else str(existing["field_value"])
                if old_val != val:
                    changed_keys.append(key)
                self.conn.execute("""
                    INSERT INTO task_dynamic_fields (task_id, field_key, field_value, updated_at)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(task_id, field_key) DO UPDATE
                    SET field_value = excluded.field_value, updated_at = excluded.updated_at
                """, (task_id, key, val, now))
                if key in DYNAMIC_FIELD_OVERVIEW_MAP:
                    overview_updates[DYNAMIC_FIELD_OVERVIEW_MAP[key]] = val

            update_result = self.conn.execute("""
                UPDATE project_tasks
                SET status = ?, assigned_to = ?, comments = ?, priority = ?, actual_start = ?, actual_finish = ?,
                    business_plan_enabled = ?, business_plan_year = ?, last_updated = ?, revision = revision + 1
                WHERE task_id = ? AND revision = ?
            """, (status, assigned_to or None, comments or None, priority, actual_start, actual_finish,
                  enabled_int, year_val, now, task_id, current_revision))
            if update_result.rowcount != 1:
                raise RuntimeError("This component was updated by someone else. Refresh and review the latest values.")

            if enabled_int:
                self.conn.execute("""
                    UPDATE projects
                    SET business_plan_enabled = 1, business_plan_year = ?, last_updated = ?, revision = revision + 1
                    WHERE project_id = ?
                """, (year_val, now, task["project_id"]))

            if overview_updates:
                self.update_project_overview_fields(task["project_id"], overview_updates)

            if task.get("task_name") in {"Reservoir CoS", "Trap CoS", "Seal CoS"}:
                self.recalculate_presence_cos(task["project_id"], changed_by)

            if changed_keys:
                readable = [key.replace("_", " ") for key in changed_keys]
                listed = ", ".join(readable[:8]) + (", and more" if len(readable) > 8 else "")
                self.log_task_event(task_id, task["project_id"], task["task_name"], "Component Inputs Updated",
                                    None, None, changed_by, f"Updated inputs: {listed}.")
            if status != old_status or assigned_to != old_assigned_to or comments != old_comments or priority != old_priority:
                self.log_task_event(task_id, task["project_id"], task["task_name"], "Component Update",
                                    old_status, status, changed_by, comments or f"Status set to {status}.")

            current = self.get_task(task_id)
            if status in DONE_STATUSES:
                self.conn.execute("""
                    UPDATE projects
                    SET current_stage = ?, current_task = ?, current_owner = ?, last_updated = ?, revision = revision + 1
                    WHERE project_id = ?
                """, (current["stage_group"], current["task_name"], current["assigned_to"], now, current["project_id"]))
            self.refresh_project_state(current["project_id"])
            self.conn.execute("UPDATE projects SET revision = revision + 1 WHERE project_id = ?", (current["project_id"],))
            result = self.get_task(task_id) or {}
        return result

    def refresh_project_state(self, project_id):
        active = self.reconcile_project_flow(project_id)

        if active:
            new_stage = active["stage_group"]
            new_task = active["task_name"]
            new_owner = active["assigned_to"]
            overall_status = "In Progress"
            project = self.get_project(project_id)
            current_stage_started_at = project["current_stage_started_at"]
            if project["current_stage"] != new_stage:
                current_stage_started_at = today_str()
            self.conn.execute("""
                UPDATE projects
                SET current_stage = ?, current_task = ?, current_owner = ?, overall_status = ?,
                    current_stage_started_at = ?, last_updated = ?
                WHERE project_id = ?
            """, (
                new_stage, new_task, new_owner, overall_status,
                current_stage_started_at, utc_now_str(), project_id
            ))
        else:
            project = self.get_project(project_id) or {}
            applicable_stages = BP_EXECUTION_STAGES if str(project.get("pipeline_type") or "prospect").lower() == "bp" else PROSPECT_STAGES
            incomplete = self.conn.execute("""
                SELECT COUNT(*) AS c FROM project_tasks
                WHERE project_id = ? AND is_active = 1 AND stage_group IN ({})
                  AND status NOT IN ('Approved','Not Applicable','Complete')
            """.format(",".join("?" for _ in applicable_stages)), [project_id, *applicable_stages]).fetchone()["c"]
            overall_status = "Completed" if incomplete == 0 else "In Progress"
            if overall_status == "Completed":
                final_done = self.conn.execute("""
                    SELECT task_name, stage_group, assigned_to
                    FROM project_tasks
                    WHERE project_id = ? AND task_name = 'PDA'
                    LIMIT 1
                """, (project_id,)).fetchone()
                final_task_name = final_done["task_name"] if final_done else "PDA"
                final_stage = final_done["stage_group"] if final_done else "Post-Testing"
                self.conn.execute("""
                    UPDATE projects
                    SET current_stage = ?, current_task = ?, current_owner = NULL,
                        overall_status = ?, current_stage_started_at = ?, last_updated = ?
                    WHERE project_id = ?
                """, (final_stage, final_task_name, overall_status, today_str(), utc_now_str(), project_id))
            else:
                earliest_open = self.conn.execute("""
                    SELECT stage_group, task_name, assigned_to
                    FROM project_tasks
                    WHERE project_id = ? AND is_active = 1 AND stage_group IN ({})
                      AND status NOT IN ('Approved','Not Applicable','Complete')
                    ORDER BY sequence_no
                    LIMIT 1
                """.format(",".join("?" for _ in applicable_stages)), [project_id, *applicable_stages]).fetchone()
                if earliest_open:
                    self.conn.execute("""
                        UPDATE projects
                        SET current_stage = ?, current_task = ?, current_owner = ?, overall_status = ?, last_updated = ?
                        WHERE project_id = ?
                    """, (earliest_open['stage_group'], earliest_open['task_name'], earliest_open['assigned_to'], overall_status, utc_now_str(), project_id))
                else:
                    self.conn.execute("""
                        UPDATE projects SET overall_status = ?, last_updated = ? WHERE project_id = ?
                    """, (overall_status, utc_now_str(), project_id))

    def update_project_location(self, project_id, location_text):
        self.conn.execute(
            "UPDATE projects SET location = ?, last_updated = ? WHERE project_id = ?",
            ((location_text or '').strip() or None, utc_now_str(), project_id)
        )
        self.conn.commit()

    def log_task_event(self, task_id, project_id, task_name, action_type, old_status, new_status, changed_by, comment):
        self.conn.execute("""
            INSERT INTO task_history (
                task_id, project_id, task_name, action_type, old_status, new_status, changed_at, changed_by, comment
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            task_id, project_id, task_name, action_type, old_status, new_status, utc_now_str(), changed_by, comment
        ))

    def get_activity_log(self, project_id=None, limit=500):
        base_query = """
            SELECT th.*, p.project_name
            FROM task_history th
            LEFT JOIN projects p ON p.project_id = th.project_id
        """
        if project_id:
            rows = self.conn.execute(
                base_query + " WHERE th.project_id = ? ORDER BY th.changed_at DESC LIMIT ?",
                (project_id, limit),
            ).fetchall()
        else:
            rows = self.conn.execute(
                base_query + " ORDER BY th.changed_at DESC LIMIT ?",
                (limit,),
            ).fetchall()
        return [dict(r) for r in rows]

    def dashboard_metrics(self):
        rows = self.conn.execute("""
            SELECT overall_status, current_stage, current_owner, target_date, current_stage_started_at
            FROM projects
            WHERE COALESCE(archived, 0) = 0
        """).fetchall()
        metrics = {
            "Completed Wells": sum(1 for row in rows if row["overall_status"] == "Completed"),
            "Components Under Review": self.conn.execute(
                "SELECT COUNT(*) AS c FROM project_tasks WHERE is_active = 1 AND status = 'Under Review'"
            ).fetchone()["c"],
        }
        stage_counts = {stage: 0 for stage in STAGE_ORDER}
        owner_workload = {}
        for row in rows:
            stage = row["current_stage"]
            owner = row["current_owner"]
            if stage in stage_counts:
                stage_counts[stage] += 1
            if owner:
                owner_workload[owner] = owner_workload.get(owner, 0) + 1
        return metrics, stage_counts, owner_workload

    def monthly_progress_metrics(self, limit=12):
        rows = self.conn.execute("""
            WITH activity AS (
                SELECT
                    substr(changed_at, 1, 7) AS month,
                    SUM(CASE WHEN action_type = 'Lead Created' THEN 1 ELSE 0 END) AS leads_created,
                    SUM(CASE WHEN action_type = 'Well Added to BP' THEN 1 ELSE 0 END) AS wells_added_to_bp,
                    SUM(CASE WHEN action_type IN ('Task Update', 'Component Update')
                                  AND new_status IN ('Approved', 'Complete') THEN 1 ELSE 0 END) AS components_completed
                FROM task_history
                WHERE changed_at IS NOT NULL
                GROUP BY substr(changed_at, 1, 7)
            ), completed AS (
                SELECT substr(COALESCE(last_updated, ''), 1, 7) AS month, COUNT(*) AS completed_wells
                FROM projects
                WHERE overall_status = 'Completed'
                GROUP BY substr(COALESCE(last_updated, ''), 1, 7)
            )
            SELECT activity.month, activity.leads_created, activity.wells_added_to_bp,
                   activity.components_completed, COALESCE(completed.completed_wells, 0) AS completed_wells
            FROM activity
            LEFT JOIN completed ON completed.month = activity.month
            ORDER BY activity.month DESC
            LIMIT ?
        """, (limit,)).fetchall()
        monthly = []
        for row in rows:
            leads = int(row["leads_created"] or 0)
            completed_components = int(row["components_completed"] or 0)
            added_to_bp = int(row["wells_added_to_bp"] or 0)
            monthly.append({
                "month": row["month"] or "Unknown",
                "leads_created": leads,
                "wells_created": leads,
                "wells_completed": int(row["completed_wells"] or 0),
                "components_completed": completed_components,
                "wells_added_to_bp": added_to_bp,
                "progress_index": leads + completed_components + added_to_bp,
            })
        return list(reversed(monthly))

    def export_to_excel(self, filepath):
        if pd is None:
            raise RuntimeError("pandas is not available in this environment.")
        if Font is None or get_column_letter is None:
            raise RuntimeError("openpyxl styling tools are not available in this environment.")

        projects_df = pd.read_sql_query("SELECT * FROM projects", self.conn)
        tasks_df = pd.read_sql_query("SELECT * FROM project_tasks", self.conn)
        history_df = pd.read_sql_query("SELECT * FROM task_history", self.conn)

        if not projects_df.empty:
            for col in ["project_id", "project_name", "overall_status", "current_stage", "current_task",
                        "current_owner", "drill_result", "start_date", "target_date", "location",
                        "current_stage_started_at", "last_updated"]:
                if col not in projects_df.columns:
                    projects_df[col] = None
            projects_df["health"] = projects_df.apply(
                lambda r: health_from_target(r.get("target_date"), r.get("overall_status")), axis=1
            )
            projects_df["days_to_target"] = pd.to_datetime(projects_df["target_date"], errors="coerce")
            projects_df["days_to_target"] = (projects_df["days_to_target"] - pd.Timestamp.today().normalize()).dt.days
            overview_df = projects_df.reindex(columns=[
                "project_id", "project_name", "overall_status", "health", "current_stage",
                "current_task", "current_owner", "drill_result", "start_date", "target_date",
                "location", "current_stage_started_at", "last_updated"
            ]).copy()
            overview_df.columns = [
                "Well ID", "Well Name", "Overall Status", "Health", "Current Stage",
                "Current Task", "Assignee", "Drill Result", "Start Date", "Target Date",
                "Location", "Stage Started", "Last Updated"
            ]
            health_order = {"Overdue": 0, "Due Soon": 1, "On Track": 2, "Completed": 3}
            overview_df["_sort"] = overview_df["Health"].map(health_order).fillna(9)
            overview_df = overview_df.sort_values(["_sort", "Target Date", "Well Name"]).drop(columns=["_sort"])
        else:
            overview_df = pd.DataFrame(columns=[
                "Well ID", "Well Name", "Overall Status", "Health", "Current Stage",
                "Current Task", "Assignee", "Drill Result", "Start Date", "Target Date",
                "Location", "Stage Started", "Last Updated"
            ])

        task_export_df = tasks_df.copy()
        if not task_export_df.empty:
            for col in ["project_id", "sequence_no", "task_name", "stage_group", "assigned_to",
                        "status", "planned_start", "planned_finish", "actual_start", "actual_finish",
                        "comments", "is_active"]:
                if col not in task_export_df.columns:
                    task_export_df[col] = None
            task_export_df = task_export_df.reindex(columns=[
                "project_id", "sequence_no", "task_name", "stage_group", "assigned_to",
                "status", "planned_start", "planned_finish", "actual_start", "actual_finish",
                "comments", "is_active"
            ]).copy()
            task_export_df.columns = [
                "Well ID", "Seq", "Component", "Stage", "Assignee",
                "Status", "Planned Start", "Planned Finish", "Actual Start", "Actual Finish",
                "Comments", "Active"
            ]
            task_export_df = task_export_df.sort_values(["Well ID", "Seq"])
        else:
            task_export_df = pd.DataFrame(columns=[
                "Well ID", "Seq", "Component", "Stage", "Assignee",
                "Status", "Planned Start", "Planned Finish", "Actual Start", "Actual Finish",
                "Comments", "Active"
            ])

        monthly_df = pd.DataFrame(self.monthly_progress_metrics(limit=12))
        if not monthly_df.empty:
            monthly_df = monthly_df.rename(columns={
                "month": "Month",
                "leads_created": "Leads Created",
                "wells_created": "Wells Created",
                "wells_completed": "Wells Completed",
                "components_completed": "Components Completed",
                "wells_added_to_bp": "Wells Added to BP",
                "progress_index": "Progress Index",
            })

        projects = self.get_projects()
        _, stage_counts, owner_workload = self.dashboard_metrics()
        monthly_all = self.monthly_progress_metrics(limit=120)
        summary_rows = [
            ["Metric", "Value"],
            ["Leads Created", sum(int(r.get("leads_created") or 0) for r in monthly_all)],
            ["Wells Completed", len([p for p in projects if p.get("overall_status") == "Completed"])],
            ["Components Completed", sum(int(r.get("components_completed") or 0) for r in monthly_all)],
            ["Wells Added to BP", len([p for p in projects if int(p.get("business_plan_enabled") or 0) == 1])],
        ]

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            pd.DataFrame(summary_rows[1:], columns=summary_rows[0]).to_excel(writer, sheet_name="Executive Summary", index=False, startrow=3)
            overview_df.to_excel(writer, sheet_name="Wells Overview", index=False, startrow=3)
            task_export_df.to_excel(writer, sheet_name="Task Register", index=False, startrow=3)
            if not monthly_df.empty:
                monthly_df.to_excel(writer, sheet_name="Monthly Progress", index=False, startrow=3)
            else:
                pd.DataFrame(columns=["Month", "Wells Created", "Components Completed", "Waiting Events", "Drill Updates", "Progress Index", "Cumulative Completed"]).to_excel(writer, sheet_name="Monthly Progress", index=False, startrow=3)

            book = writer.book
            ws_summary = writer.sheets["Executive Summary"]
            ws_overview = writer.sheets["Wells Overview"]
            ws_tasks = writer.sheets["Task Register"]
            ws_monthly = writer.sheets["Monthly Progress"]

            title_fill = PatternFill("solid", fgColor="0F2747")
            header_fill = PatternFill("solid", fgColor="163A6B")
            soft_fill = PatternFill("solid", fgColor="F4F7FB")
            white_font = Font(color="FFFFFF", bold=True)
            title_font = Font(color="FFFFFF", bold=True, size=16)
            header_font = Font(color="FFFFFF", bold=True, size=11)
            label_font = Font(color="334155", bold=True)
            body_font = Font(color="0F172A", size=10)
            thin = Side(style="thin", color="D8E1EB")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center")
            left = Alignment(horizontal="left", vertical="center")

            def style_sheet(ws, title, subtitle):
                max_col = max(ws.max_column, 1)
                end_col = get_column_letter(max_col)
                ws.merge_cells(f"A1:{end_col}1")
                ws["A1"] = title
                ws["A1"].fill = title_fill
                ws["A1"].font = title_font
                ws["A1"].alignment = left
                ws.row_dimensions[1].height = 24

                ws.merge_cells(f"A2:{end_col}2")
                ws["A2"] = subtitle
                ws["A2"].fill = soft_fill
                ws["A2"].font = Font(color="475569", italic=True, size=10)
                ws["A2"].alignment = left
                ws.row_dimensions[2].height = 20

                for cell in ws[4]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                    cell.border = border

                ws.freeze_panes = "A5"
                ws.sheet_view.showGridLines = False

                for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border
                        cell.alignment = left
                        cell.font = body_font

                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = 0
                    for row_idx in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        val = "" if cell.value is None else str(cell.value)
                        max_len = max(max_len, len(val))
                    ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 32)

                if Table is not None and ws.max_row >= 4 and ws.max_column >= 1:
                    ref = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"

                    # Excel requires workbook-level table names to be unique.
                    # The previous implementation built table names from the report title and
                    # truncated them to 20 characters, so every sheet became
                    # "TblSegmentFactoryTracke" and export failed after the first sheet.
                    # Build the table name from the worksheet title instead and guard against
                    # collisions for older/future export sections.
                    base_name = ''.join(ch for ch in ws.title.title() if ch.isalnum()) or "Sheet"
                    table_name = f"Tbl{base_name[:24]}"
                    existing_names = set()
                    try:
                        for existing_ws in book.worksheets:
                            existing_names.update(str(name) for name in existing_ws.tables.keys())
                    except Exception:
                        existing_names = set()
                    candidate = table_name
                    suffix = 1
                    while candidate in existing_names:
                        suffix += 1
                        candidate = f"{table_name[:24]}{suffix}"

                    table = Table(displayName=candidate, ref=ref)
                    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
                    ws.add_table(table)

            exported_at = datetime.now().strftime("%Y-%m-%d %H:%M")
            style_sheet(ws_summary, "UR Segment Factory Tracker — Executive Summary", f"Exported on {exported_at}")
            style_sheet(ws_overview, "UR Segment Factory Tracker — Wells Overview", f"Executive export generated on {exported_at}")
            style_sheet(ws_tasks, "UR Segment Factory Tracker — Task Register", f"Detailed task register exported on {exported_at}")
            style_sheet(ws_monthly, "UR Segment Factory Tracker — Monthly Progress", f"Progress trend exported on {exported_at}")

            # Executive Summary enhancements
            ws_summary["D4"] = "Stage"
            ws_summary["E4"] = "Count"
            ws_summary["G4"] = "Assignee"
            ws_summary["H4"] = "Open Wells"
            for cell in [ws_summary["D4"], ws_summary["E4"], ws_summary["G4"], ws_summary["H4"]]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            row = 5
            for stage, count in stage_counts.items():
                ws_summary[f"D{row}"] = stage
                ws_summary[f"E{row}"] = count
                ws_summary[f"D{row}"].font = body_font
                ws_summary[f"E{row}"].font = body_font
                ws_summary[f"D{row}"].border = border
                ws_summary[f"E{row}"].border = border
                row += 1

            row = 5
            for owner, count in sorted(owner_workload.items(), key=lambda x: (-x[1], x[0]))[:10]:
                ws_summary[f"G{row}"] = owner
                ws_summary[f"H{row}"] = count
                ws_summary[f"G{row}"].font = body_font
                ws_summary[f"H{row}"].font = body_font
                ws_summary[f"G{row}"].border = border
                ws_summary[f"H{row}"].border = border
                row += 1

            # color the metric values
            metric_colors = {
                "Leads Created": "1D4ED8",
                "Wells Completed": "059669",
                "Components Completed": "7C3AED",
                "Wells Added to BP": "0F766E",
            }
            for r in range(5, 9):
                metric_name = ws_summary[f"A{r}"].value
                ws_summary[f"B{r}"].font = Font(color=metric_colors.get(metric_name, "0F172A"), bold=True, size=12)
                ws_summary[f"A{r}"].font = label_font

            # Conditional colors in overview
            health_col = None
            status_col = None
            for c in range(1, ws_overview.max_column + 1):
                v = ws_overview.cell(4, c).value
                if v == "Health":
                    health_col = c
                if v == "Overall Status":
                    status_col = c
            fill_map = {
                "Overdue": PatternFill("solid", fgColor="FEE2E2"),
                "Due Soon": PatternFill("solid", fgColor="FEF3C7"),
                "On Track": PatternFill("solid", fgColor="DCFCE7"),
                "Completed": PatternFill("solid", fgColor="DBEAFE"),
                "Waiting": PatternFill("solid", fgColor="F3E8FF"),
                "In Progress": PatternFill("solid", fgColor="DBEAFE"),
            }
            for r in range(5, ws_overview.max_row + 1):
                if health_col:
                    cell = ws_overview.cell(r, health_col)
                    if cell.value in fill_map:
                        cell.fill = fill_map[cell.value]
                        cell.font = Font(color="0F172A", bold=True)
                if status_col:
                    cell = ws_overview.cell(r, status_col)
                    if cell.value in fill_map:
                        cell.fill = fill_map[cell.value]
                        cell.font = Font(color="0F172A", bold=True)

            # make summary a bit tighter
            ws_summary.column_dimensions["A"].width = 22
            ws_summary.column_dimensions["B"].width = 14
            ws_summary.column_dimensions["D"].width = 20
            ws_summary.column_dimensions["E"].width = 10
            ws_summary.column_dimensions["G"].width = 22
            ws_summary.column_dimensions["H"].width = 12





    def get_project_dynamic_field_map(self, project_id: int):
        rows = self.conn.execute("""
            SELECT pt.task_name, pt.sequence_no, tdf.field_key, tdf.field_value
            FROM project_tasks pt
            LEFT JOIN task_dynamic_fields tdf ON tdf.task_id = pt.task_id
            WHERE pt.project_id = ? AND pt.is_active = 1
            ORDER BY pt.sequence_no, tdf.field_key
        """, (project_id,)).fetchall()
        data = {}
        for row in rows:
            name = row["task_name"]
            data.setdefault(name, {})
            if row["field_key"]:
                data[name][row["field_key"]] = row["field_value"] or ""
        return data

    def get_component_folder_link(self, project_id: int, task_id: int):
        project = self.get_project(project_id)
        task = self.get_task(task_id)
        if not project or not task or int(task.get("project_id") or 0) != int(project_id):
            raise ValueError("Component folder could not be resolved.")
        task_name = task.get("task_name") or "Component"
        requires_folder = task_name in COMPONENT_FILE_SECTIONS
        field_name, well_name = parse_field_and_well(project.get("project_name") or "")
        section = _safe_folder_name(task_name)
        server_path = WELL_OVERVIEW_DIRECTORY_ROOT / field_name / well_name / "Component Files" / section
        unc_path = _windows_join(WINDOWS_WELL_SHARE_ROOT, field_name, well_name, "Component Files", section)
        return {
            "requires_folder": 1 if requires_folder else 0,
            "path": unc_path,
            "unc_path": unc_path,
            "file_url": _windows_path_to_file_url(unc_path),
            "section": task_name,
            "server_path": str(server_path),
        }

    def get_open_folder_path(self, project_id: int, section_key: str = "well"):
        project = self.get_project(project_id)
        if not project:
            raise ValueError("Well not found.")
        if section_key not in WELL_OVERVIEW_DIRECTORY_MAP:
            raise ValueError(f"Unknown folder section: {section_key}")
        root = LEAD_WORKFLOW_DIRECTORY_ROOT if section_key in LEAD_WORKFLOW_SECTION_KEYS else WELL_OVERVIEW_DIRECTORY_ROOT
        section = WELL_OVERVIEW_DIRECTORY_MAP.get(section_key, "")
        field_name, well_name = parse_field_and_well(project.get("project_name") or "")
        path = root / field_name / well_name
        return path / section if section else path

    def get_client_folder_link(self, project_id: int, section_key: str = "well"):
        project = self.get_project(project_id)
        if not project:
            raise ValueError("Well not found.")
        if section_key not in WELL_OVERVIEW_DIRECTORY_MAP:
            raise ValueError(f"Unknown folder section: {section_key}")
        root = WINDOWS_LEAD_WORKFLOW_SHARE_ROOT if section_key in LEAD_WORKFLOW_SECTION_KEYS else WINDOWS_WELL_SHARE_ROOT
        section = WELL_OVERVIEW_DIRECTORY_MAP.get(section_key, "")
        field_name, well_name = parse_field_and_well(project.get("project_name") or "")
        unc_path = _windows_join(root, field_name, well_name, section)
        return {"path": unc_path, "unc_path": unc_path, "file_url": _windows_path_to_file_url(unc_path), "section": section_key}

    def ensure_well_folders(self, project_id: int):
        created = []
        for section_key in WELL_OVERVIEW_DIRECTORY_MAP:
            path = self.get_open_folder_path(project_id, section_key)
            try:
                path.mkdir(parents=True, exist_ok=True)
                created.append(str(path))
            except Exception:
                pass
        for task in self.get_project_tasks(project_id):
            if task.get("task_name") not in COMPONENT_FILE_SECTIONS:
                continue
            info = self.get_component_folder_link(project_id, task["task_id"])
            try:
                Path(info["server_path"]).mkdir(parents=True, exist_ok=True)
                created.append(info["server_path"])
            except Exception:
                pass
        return str(self.get_open_folder_path(project_id, "well"))

    def open_folder(self, project_id: int, section_key: str = "well"):
        return self.get_client_folder_link(project_id, section_key)

    def bottleneck_rows(self, threshold_days: int = 14):
        rows = self.conn.execute("""
            SELECT p.project_id, p.project_name, p.current_stage, p.current_task, p.current_owner,
                   p.current_stage_started_at, p.target_date
            FROM projects p
            WHERE COALESCE(p.archived, 0) = 0 AND p.overall_status != 'Completed'
            ORDER BY p.project_name COLLATE NOCASE
        """).fetchall()
        result = []
        for row in rows:
            item = dict(row)
            age = days_between(item.get("current_stage_started_at")) or 0
            item["stage_age_days"] = age
            item["is_bottleneck"] = 1 if age >= threshold_days else 0
            result.append(item)
        return result

    def attention_rows(self):
        rows = []
        for project in self.get_projects():
            health = health_from_target(project.get("target_date"), project.get("overall_status"))
            if health in ("Overdue", "Due Soon") or project.get("has_high_priority_tasks"):
                item = dict(project)
                item["health"] = health
                rows.append(item)
        return rows




def _windows_join(*parts):
    """Join Windows/UNC path parts without depending on the server OS."""
    clean = []
    for idx, part in enumerate(parts):
        text = str(part or "").strip()
        if not text:
            continue
        if idx == 0:
            clean.append(text.rstrip("\\/"))
        else:
            clean.append(text.strip("\\/"))
    return "\\".join(clean)


def _windows_path_to_file_url(path_text: str):
    """Convert a Windows UNC or drive path to a browser file:// URL."""
    path_text = str(path_text or "").strip()
    if not path_text:
        return ""
    normalized = path_text.replace("\\", "/")
    if normalized.startswith("//"):
        # UNC path: //server/share/folder -> file://///server/share/folder
        return "file:" + normalized
    if len(normalized) >= 2 and normalized[1] == ":":
        # Drive path: C:/folder -> file:///C:/folder
        return "file:///" + normalized
    return "file:///" + normalized.lstrip("/")








