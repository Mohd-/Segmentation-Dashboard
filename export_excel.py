"""Excel workbook export and its styling.

What belongs here:
- ``export_to_excel(session, filepath)`` and the openpyxl styling helpers that
  build the branded multi-sheet workbook.

What does NOT belong here:
- The metric calculations themselves (reporting.py / workflow.py) -- this module
  only lays them out.

pandas reads the raw tables straight off the SQLAlchemy connection; the derived
metrics come from the reporting/workflow layers.
"""
from __future__ import annotations

from datetime import datetime

import reporting
import workflow
from helpers import health_from_target

try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

try:
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception:  # pragma: no cover
    Font = PatternFill = Border = Side = Alignment = None
    get_column_letter = None
    Table = TableStyleInfo = None


def export_to_excel(session, filepath):
    """Write the branded multi-sheet Excel workbook to ``filepath``."""
    if pd is None:
        raise RuntimeError("pandas is not available in this environment.")
    if Font is None or get_column_letter is None:
        raise RuntimeError("openpyxl styling tools are not available in this environment.")

    connection = session.connection()
    projects_df = pd.read_sql_query("SELECT * FROM projects", connection)
    tasks_df = pd.read_sql_query("SELECT * FROM project_tasks", connection)

    if not projects_df.empty:
        for col in ["project_id", "project_name", "overall_status", "current_stage", "current_task",
                    "current_owner", "start_date", "target_date",
                    "current_stage_started_at", "last_updated"]:
            if col not in projects_df.columns:
                projects_df[col] = None
        projects_df["health"] = projects_df.apply(
            lambda r: health_from_target(r.get("target_date"), r.get("overall_status")), axis=1
        )
        projects_df["days_to_target"] = pd.to_datetime(projects_df["target_date"], errors="coerce")
        projects_df["days_to_target"] = (projects_df["days_to_target"] - pd.Timestamp.today().normalize()).dt.days
        overview_df = projects_df.reindex(columns=[
            "project_id", "project_name", "overall_status", "health", "current_stage",
            "current_task", "current_owner", "start_date", "target_date",
            "current_stage_started_at", "last_updated"
        ]).copy()
        overview_df.columns = [
            "Well ID", "Well Name", "Overall Status", "Health", "Current Stage",
            "Current Task", "Assignee", "Start Date", "Target Date",
            "Stage Started", "Last Updated"
        ]
        health_order = {"Overdue": 0, "Due Soon": 1, "On Track": 2, "Completed": 3}
        overview_df["_sort"] = overview_df["Health"].map(health_order).fillna(9)
        overview_df = overview_df.sort_values(["_sort", "Target Date", "Well Name"]).drop(columns=["_sort"])
    else:
        overview_df = pd.DataFrame(columns=[
            "Well ID", "Well Name", "Overall Status", "Health", "Current Stage",
            "Current Task", "Assignee", "Start Date", "Target Date",
            "Stage Started", "Last Updated"
        ])

    task_export_df = tasks_df.copy()
    if not task_export_df.empty:
        for col in ["project_id", "sequence_no", "task_name", "stage_group", "assigned_to",
                    "status", "actual_start", "actual_finish",
                    "comments", "is_active"]:
            if col not in task_export_df.columns:
                task_export_df[col] = None
        task_export_df = task_export_df.reindex(columns=[
            "project_id", "sequence_no", "task_name", "stage_group", "assigned_to",
            "status", "actual_start", "actual_finish",
            "comments", "is_active"
        ]).copy()
        task_export_df.columns = [
            "Well ID", "Seq", "Component", "Stage", "Assignee",
            "Status", "Actual Start", "Actual Finish",
            "Comments", "Active"
        ]
        task_export_df = task_export_df.sort_values(["Well ID", "Seq"])
    else:
        task_export_df = pd.DataFrame(columns=[
            "Well ID", "Seq", "Component", "Stage", "Assignee",
            "Status", "Actual Start", "Actual Finish",
            "Comments", "Active"
        ])

    monthly_df = pd.DataFrame(reporting.monthly_progress_metrics(session, limit=12))
    if not monthly_df.empty:
        monthly_df = monthly_df.rename(columns={
            "month": "Month",
            "leads_created": "Leads Created",
            "wells_created": "Wells Created",
            "wells_completed": "Wells Completed",
            "components_completed": "Components Completed",
            "wells_added_to_bp": "Wells Added to BP",
            "progress_index": "Progress Index",
        })

    projects = workflow.get_projects(session)
    _, stage_counts, owner_workload = reporting.dashboard_metrics(session)
    monthly_all = reporting.monthly_progress_metrics(session, limit=120)
    summary_rows = [
        ["Metric", "Value"],
        ["Leads Created", sum(int(r.get("leads_created") or 0) for r in monthly_all)],
        ["Wells Completed", len([p for p in projects if p.get("overall_status") == "Completed"])],
        ["Components Completed", sum(int(r.get("components_completed") or 0) for r in monthly_all)],
        ["Wells Added to BP", len([p for p in projects if int(p.get("business_plan_enabled") or 0) == 1])],
    ]

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows[1:], columns=summary_rows[0]).to_excel(writer, sheet_name="Executive Summary", index=False, startrow=3)
        overview_df.to_excel(writer, sheet_name="Wells Overview", index=False, startrow=3)
        task_export_df.to_excel(writer, sheet_name="Task Register", index=False, startrow=3)
        if not monthly_df.empty:
            monthly_df.to_excel(writer, sheet_name="Monthly Progress", index=False, startrow=3)
        else:
            pd.DataFrame(columns=["Month", "Wells Created", "Components Completed", "Waiting Events", "Drill Updates", "Progress Index", "Cumulative Completed"]).to_excel(writer, sheet_name="Monthly Progress", index=False, startrow=3)

        book = writer.book
        ws_summary = writer.sheets["Executive Summary"]
        ws_overview = writer.sheets["Wells Overview"]
        ws_tasks = writer.sheets["Task Register"]
        ws_monthly = writer.sheets["Monthly Progress"]

        title_fill = PatternFill("solid", fgColor="0F2747")
        header_fill = PatternFill("solid", fgColor="163A6B")
        soft_fill = PatternFill("solid", fgColor="F4F7FB")
        white_font = Font(color="FFFFFF", bold=True)
        title_font = Font(color="FFFFFF", bold=True, size=16)
        header_font = Font(color="FFFFFF", bold=True, size=11)
        label_font = Font(color="334155", bold=True)
        body_font = Font(color="0F172A", size=10)
        thin = Side(style="thin", color="D8E1EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        def style_sheet(ws, title, subtitle):
            max_col = max(ws.max_column, 1)
            end_col = get_column_letter(max_col)
            ws.merge_cells(f"A1:{end_col}1")
            ws["A1"] = title
            ws["A1"].fill = title_fill
            ws["A1"].font = title_font
            ws["A1"].alignment = left
            ws.row_dimensions[1].height = 24

            ws.merge_cells(f"A2:{end_col}2")
            ws["A2"] = subtitle
            ws["A2"].fill = soft_fill
            ws["A2"].font = Font(color="475569", italic=True, size=10)
            ws["A2"].alignment = left
            ws.row_dimensions[2].height = 20

            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            ws.freeze_panes = "A5"
            ws.sheet_view.showGridLines = False

            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = left
                    cell.font = body_font

            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 32)

            if Table is not None and ws.max_row >= 4 and ws.max_column >= 1:
                ref = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"

                # Excel requires workbook-level table names to be unique.
                # The previous implementation built table names from the report title and
                # truncated them to 20 characters, so every sheet became
                # "TblSegmentFactoryTracke" and export failed after the first sheet.
                # Build the table name from the worksheet title instead and guard against
                # collisions for older/future export sections.
                base_name = ''.join(ch for ch in ws.title.title() if ch.isalnum()) or "Sheet"
                table_name = f"Tbl{base_name[:24]}"
                existing_names = set()
                try:
                    for existing_ws in book.worksheets:
                        existing_names.update(str(name) for name in existing_ws.tables.keys())
                except Exception:
                    existing_names = set()
                candidate = table_name
                suffix = 1
                while candidate in existing_names:
                    suffix += 1
                    candidate = f"{table_name[:24]}{suffix}"

                table = Table(displayName=candidate, ref=ref)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
                ws.add_table(table)

        exported_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        style_sheet(ws_summary, "UR Segment Factory Tracker — Executive Summary", f"Exported on {exported_at}")
        style_sheet(ws_overview, "UR Segment Factory Tracker — Wells Overview", f"Executive export generated on {exported_at}")
        style_sheet(ws_tasks, "UR Segment Factory Tracker — Task Register", f"Detailed task register exported on {exported_at}")
        style_sheet(ws_monthly, "UR Segment Factory Tracker — Monthly Progress", f"Progress trend exported on {exported_at}")

        # Executive Summary enhancements
        ws_summary["D4"] = "Stage"
        ws_summary["E4"] = "Count"
        ws_summary["G4"] = "Assignee"
        ws_summary["H4"] = "Open Wells"
        for cell in [ws_summary["D4"], ws_summary["E4"], ws_summary["G4"], ws_summary["H4"]]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        row = 5
        for stage, count in stage_counts.items():
            ws_summary[f"D{row}"] = stage
            ws_summary[f"E{row}"] = count
            ws_summary[f"D{row}"].font = body_font
            ws_summary[f"E{row}"].font = body_font
            ws_summary[f"D{row}"].border = border
            ws_summary[f"E{row}"].border = border
            row += 1

        row = 5
        for owner, count in sorted(owner_workload.items(), key=lambda x: (-x[1], x[0]))[:10]:
            ws_summary[f"G{row}"] = owner
            ws_summary[f"H{row}"] = count
            ws_summary[f"G{row}"].font = body_font
            ws_summary[f"H{row}"].font = body_font
            ws_summary[f"G{row}"].border = border
            ws_summary[f"H{row}"].border = border
            row += 1

        # color the metric values
        metric_colors = {
            "Leads Created": "1D4ED8",
            "Wells Completed": "059669",
            "Components Completed": "7C3AED",
            "Wells Added to BP": "0F766E",
        }
        for r in range(5, 9):
            metric_name = ws_summary[f"A{r}"].value
            ws_summary[f"B{r}"].font = Font(color=metric_colors.get(metric_name, "0F172A"), bold=True, size=12)
            ws_summary[f"A{r}"].font = label_font

        # Conditional colors in overview
        health_col = None
        status_col = None
        for c in range(1, ws_overview.max_column + 1):
            v = ws_overview.cell(4, c).value
            if v == "Health":
                health_col = c
            if v == "Overall Status":
                status_col = c
        fill_map = {
            "Overdue": PatternFill("solid", fgColor="FEE2E2"),
            "Due Soon": PatternFill("solid", fgColor="FEF3C7"),
            "On Track": PatternFill("solid", fgColor="DCFCE7"),
            "Completed": PatternFill("solid", fgColor="DBEAFE"),
            "Waiting": PatternFill("solid", fgColor="F3E8FF"),
            "In Progress": PatternFill("solid", fgColor="DBEAFE"),
        }
        for r in range(5, ws_overview.max_row + 1):
            if health_col:
                cell = ws_overview.cell(r, health_col)
                if cell.value in fill_map:
                    cell.fill = fill_map[cell.value]
                    cell.font = Font(color="0F172A", bold=True)
            if status_col:
                cell = ws_overview.cell(r, status_col)
                if cell.value in fill_map:
                    cell.fill = fill_map[cell.value]
                    cell.font = Font(color="0F172A", bold=True)

        # make summary a bit tighter
        ws_summary.column_dimensions["A"].width = 22
        ws_summary.column_dimensions["B"].width = 14
        ws_summary.column_dimensions["D"].width = 20
        ws_summary.column_dimensions["E"].width = 10
        ws_summary.column_dimensions["G"].width = 22
        ws_summary.column_dimensions["H"].width = 12
