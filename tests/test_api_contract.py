"""Characterization tests for HTTP status codes and JSON response shapes.

These pin the *current* API contract (main.py routes backed by database.py) so a
SQLAlchemy refactor can be checked against them. Where actual behavior surprised
us relative to a naive reading of the spec, a comment marks it.
"""
from __future__ import annotations

import io
import json

import openpyxl
import pytest

import config
import portfolio_export
from conftest import create_project, get_task_by_name, get_tasks, raw_sqlite_connect


# ---------------------------------------------------------------------------
# /api/health
# ---------------------------------------------------------------------------

def test_health_ok(client):
    resp = client.get("/api/health")
    assert resp.status_code == 200
    body = resp.get_json()
    for key in ("ok", "app", "version", "backend", "db"):
        assert key in body
    assert body["ok"] is True
    # Pin the release label from config.APP_VERSION (v18 release; a product
    # axis distinct from the database schema version -- see config.py).
    assert body["version"] == "v18"


# ---------------------------------------------------------------------------
# /api/meta
# ---------------------------------------------------------------------------

def test_meta_serves_the_user_maintained_pick_lists(client, tmp_path, monkeypatch):
    """Formations and wellbore sizes come from config/lists.yaml, not code.

    The client's schema.js copy is a boot fallback, so this endpoint has to be
    the one that reflects an edit to that file. Both lists also fall back
    cleanly when the file is absent -- a deployment that never writes one must
    behave exactly as it did before the file existed.
    """
    body = client.get("/api/meta").get_json()
    assert body["formations"] == list(config.formations())
    assert body["hole_sections"] == list(config.hole_sections())
    # Hole sections used to be empty (env-var only, and nobody set it), which
    # is why the BP Gate's interval dropdowns offered formations alone.
    assert body["hole_sections"], "the shipped list is no longer empty"

    lists_file = tmp_path / "lists.yaml"
    lists_file.write_text("formations:\n  - ZETA\n  - ETA\nhole_sections:\n  - 9.5in Section\n",
                          encoding="utf-8")
    monkeypatch.setenv("SEGMENT_TRACKER_LISTS_PATH", str(lists_file))
    body = client.get("/api/meta").get_json()
    assert body["formations"] == ["ZETA", "ETA"]
    assert body["hole_sections"] == ["9.5in Section"]

    # An absent file is not an error; it is "use the built-in defaults".
    monkeypatch.setenv("SEGMENT_TRACKER_LISTS_PATH", str(tmp_path / "nope.yaml"))
    body = client.get("/api/meta").get_json()
    assert body["formations"] == list(config.DEFAULT_FORMATIONS)
    assert body["hole_sections"] == []


def test_user_list_survives_a_malformed_file(tmp_path, monkeypatch):
    """A hand-edited YAML that does not parse must not take the app down."""
    bad = tmp_path / "lists.yaml"
    bad.write_text("formations: [unclosed\n", encoding="utf-8")
    monkeypatch.setenv("SEGMENT_TRACKER_LISTS_PATH", str(bad))
    assert config.formations() == config.DEFAULT_FORMATIONS
    # A key of the wrong shape falls back the same way.
    bad.write_text("formations: SARH\n", encoding="utf-8")
    assert config.formations() == config.DEFAULT_FORMATIONS
    # Blanks and duplicates are cleaned rather than reaching a dropdown.
    bad.write_text("formations:\n  - SARH\n  - '  '\n  - SARH\n  - QASM\n", encoding="utf-8")
    assert config.formations() == ("SARH", "QASM")


def test_meta_shape_matches_workflow_constants(client):
    import workflow
    resp = client.get("/api/meta")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["prospect_stages"] == workflow.PROSPECT_STAGES
    assert body["bp_stages"] == workflow.BP_EXECUTION_STAGES
    assert body["stage_order"] == workflow.STAGE_ORDER
    # The 4 lifecycle statuses, straight from the domain layer.
    assert body["statuses"] == ["Not Assigned", "In Progress", "Ready", "Approved"]
    assert workflow.STATUSES == ["Not Assigned", "In Progress", "Ready", "Approved"]
    assert body["roles"] == ["supervisor", "staff", "employee"]
    # Block name -> [AR number, ...], feeding the Portfolio's dependent
    # Block/AR dropdowns (config.SEISMIC_BLOCK_AR_MAP, from seismic_blocks.json).
    assert "seismic_blocks" in body
    assert isinstance(body["seismic_blocks"], dict)
    for ars in body["seismic_blocks"].values():
        assert isinstance(ars, list)
    # The 4 configured resource-assessment scenarios feed the Lead Resource
    # Assessment pop-up calculator's scenario dropdown.
    scenarios = body["resource_scenarios"]
    assert isinstance(scenarios, list) and len(scenarios) == 4
    for entry in scenarios:
        assert set(entry) == {"id", "label", "resource_type"}


# ---------------------------------------------------------------------------
# POST /api/projects
# ---------------------------------------------------------------------------

def test_create_project_valid(client):
    resp = client.post("/api/projects", json={"project_name": "ALPHA-1"})
    assert resp.status_code == 201
    body = resp.get_json()
    assert "project_id" in body
    assert "folder_path" in body


def test_component_folder_uses_leads_for_prospect_steps_and_wells_for_bp_steps(client):
    """Folder roots follow the component's stage, not just the record's current
    pipeline, so historical prospect steps remain under Leads after promotion."""
    pid = create_project(client, "PATH-1", pipeline_type="bp",
                         business_plan_enabled=True, business_plan_year=2030)
    prospect_task = get_task_by_name(client, pid, "Lead Assessment")
    bp_task = get_task_by_name(client, pid, "Well Proposal")

    prospect = client.get(
        f"/api/projects/{pid}/component-folder/{prospect_task['task_id']}"
    ).get_json()
    bp = client.get(
        f"/api/projects/{pid}/component-folder/{bp_task['task_id']}"
    ).get_json()

    assert prospect["unc_path"].startswith("\\\\aramco.com\\ecc\\data\\NAUGAD\\Leads\\")
    assert prospect["server_path"].startswith("/mnt/leads/")
    assert prospect["unc_path"].endswith(
        r"PATH\PATH-1\Component Files\Lead Assessment"
    )
    assert bp["unc_path"].startswith("\\\\aramco.com\\ecc\\data\\NAUGAD\\Wells\\")
    assert bp["server_path"].startswith("/mnt/wells/")
    assert bp["unc_path"].endswith(r"PATH\PATH-1\Component Files\Well Proposal")


# ---------------------------------------------------------------------------
# GET /api/projects/<id>/folders/<section_key>
# ---------------------------------------------------------------------------

def test_section_folder_returns_resolved_link_for_a_known_section(client):
    pid = create_project(client, "SECFLD-1")
    resp = client.get(f"/api/projects/{pid}/folders/well")
    assert resp.status_code == 200
    body = resp.get_json()
    assert set(body) == {"path", "unc_path", "file_url", "section", "server_path"}
    assert body["path"] == body["unc_path"]
    assert body["unc_path"] == r"\\aramco.com\ecc\data\NAUGAD\Wells\SECFLD\SECFLD-1"
    assert body["file_url"].startswith("file:")
    assert body["section"] == "Well Folder"
    assert body["server_path"] == "/mnt/wells/SECFLD/SECFLD-1"


def test_section_folder_unknown_section_is_400(client):
    pid = create_project(client, "SECFLD-2")
    resp = client.get(f"/api/projects/{pid}/folders/not-a-real-section")
    assert resp.status_code == 400
    assert "Unknown folder section" in resp.get_json()["detail"]


def test_section_folder_unknown_project_is_404(client):
    resp = client.get("/api/projects/999999/folders/well")
    assert resp.status_code == 404


def test_create_project_empty_name(client):
    resp = client.post("/api/projects", json={"project_name": ""})
    assert resp.status_code == 400
    assert "detail" in resp.get_json()


def test_create_project_duplicate_name(client):
    create_project(client, "DUP-1")
    resp = client.post("/api/projects", json={"project_name": "DUP-1"})
    assert resp.status_code == 400
    assert "detail" in resp.get_json()


def test_create_project_name_too_long(client):
    resp = client.post("/api/projects", json={"project_name": "A" * 121})
    assert resp.status_code == 400
    assert "detail" in resp.get_json()


# --- Card 1D: name uniqueness is case- and whitespace-insensitive ------------

@pytest.mark.parametrize("collider", ["wwww-44", "WWWW-44", " WWWW-44 ", "  wWwW-44"])
def test_create_project_duplicate_name_ignores_case_and_surrounding_space(client, collider):
    """'WWWW-44', 'wwww-44' and ' WWWW-44 ' are ONE lead.

    The DB's UNIQUE(project_name) index is case-sensitive, so the rule lives in
    workflow.add_project's pre-check -- pinned here per collision shape.
    """
    create_project(client, "WWWW-44")
    resp = client.post("/api/projects", json={"project_name": collider})
    assert resp.status_code == 400
    assert resp.get_json()["detail"] == "A lead with this name already exists."


def test_create_project_stores_the_original_casing(client):
    """Case-insensitive COMPARISON, case-preserving STORAGE."""
    pid = create_project(client, "  MiXeD-Case-7  ")
    assert client.get(f"/api/projects/{pid}").get_json()["project_name"] == "MiXeD-Case-7"


def test_create_bp_well_duplicate_keeps_the_lead_or_well_wording(client):
    create_project(client, "BPDUP-1", pipeline_type="bp",
                   business_plan_enabled=True, business_plan_year=2030)
    resp = client.post("/api/projects", json={
        "project_name": "bpdup-1", "pipeline_type": "bp",
        "business_plan_enabled": True, "business_plan_year": 2030,
    })
    assert resp.status_code == 400
    assert resp.get_json()["detail"] == "A lead / well with this name already exists."


def test_rename_cannot_create_a_case_variant_duplicate(client):
    create_project(client, "RENCASE-1")
    other = create_project(client, "RENCASE-2")
    resp = client.patch(f"/api/projects/{other}/rename", json={"new_name": "rencase-1"})
    assert resp.status_code == 400
    assert "already exists" in resp.get_json()["detail"]


def test_rename_can_recase_its_own_name(client):
    """The project_id exclusion keeps a pure re-casing of the record itself legal."""
    pid = create_project(client, "RECASE-9")
    resp = client.patch(f"/api/projects/{pid}/rename", json={"new_name": "recase-9"})
    assert resp.status_code == 200
    assert client.get(f"/api/projects/{pid}").get_json()["project_name"] == "recase-9"


# --- Card 1D: coordinates are optional, but a supplied value must be numeric --

@pytest.mark.parametrize("key,label", [("lead_x", "X"), ("lead_y", "Y")])
@pytest.mark.parametrize("value", ["abc", "12.3.4", "1,5", "NaN", "inf", "-inf", "12 34"])
def test_create_project_rejects_non_numeric_coordinates(client, key, label, value):
    resp = client.post("/api/projects", json={"project_name": f"COORD-{label}-BAD", key: value})
    assert resp.status_code == 400
    assert resp.get_json()["detail"] == f"Enter a valid Lead {label} Coordinate."


@pytest.mark.parametrize("lead_x,lead_y", [
    ("100", "200"),
    ("-3.5", "0"),                      # NO positive-only rule: signed is valid
    ("612345.678", "2734567.891"),      # UTM-scale precision survives
    ("1e3", "2E-2"),
])
def test_create_project_accepts_signed_and_precise_coordinates(client, lead_x, lead_y):
    pid = create_project(client, f"COORD-OK-{lead_x}-{lead_y}", lead_x=lead_x, lead_y=lead_y)
    row = client.get(f"/api/projects/{pid}").get_json()
    # projects.lead_x/lead_y are REAL columns, so the value round-trips as a
    # number: what is pinned is that the entered magnitude/precision survives
    # (no truncation, no sign rule), not the literal string.
    assert float(row["lead_x"]) == float(lead_x)
    assert float(row["lead_y"]) == float(lead_y)


def test_create_project_still_allows_missing_coordinates(client):
    """The API contract stays coordinate-OPTIONAL (Excel importer, older callers);
    Card 1D's requirement is a client-side rule on the Add New Lead control."""
    pid = create_project(client, "COORD-NONE-1")
    row = client.get(f"/api/projects/{pid}").get_json()
    assert row["lead_x"] in (None, "")
    assert row["lead_y"] in (None, "")


def test_create_project_is_not_role_gated_unless_born_bp(client):
    """Pinning the CURRENT permission model: plain creation carries no
    require_role check (unlike approve / delete / priority, which are
    supervisor-only) -- Card 1D preserves it exactly. A born-BP creation
    (business_plan_enabled truthy) is the one exception: it is a promotion
    done at creation time, so it gates on supervisor exactly like
    PATCH /api/projects/<id>/flags does. Change this test only with a
    deliberate decision.
    """
    resp = client.post("/api/projects", json={"project_name": "ROLE-UNGATED-1"})
    assert resp.status_code == 201

    resp = client.post("/api/login", json={"name": "Employee"})
    assert resp.status_code == 200, resp.get_json()
    resp = client.post("/api/projects", json={
        "project_name": "ROLE-GATED-BP-1", "business_plan_enabled": True,
        "business_plan_year": 2026,
    })
    assert resp.status_code == 403


@pytest.mark.parametrize("payload", [
    {"business_plan_enabled": True},
    # Floor is 1990 (admits imported historical wells); 1989 stays invalid.
    {"business_plan_enabled": True, "business_plan_year": 1989},
    {"business_plan_enabled": True, "business_plan_year": 2041},
])
def test_create_project_bp_enabled_needs_valid_year(client, payload):
    body = {"project_name": "BADYEAR"}
    body.update(payload)
    resp = client.post("/api/projects", json=body)
    assert resp.status_code == 400
    assert "detail" in resp.get_json()


def test_create_project_bp_enabled_valid_year_ok(client):
    resp = client.post("/api/projects", json={
        "project_name": "GOODYEAR", "business_plan_enabled": True, "business_plan_year": 2026,
    })
    assert resp.status_code == 201


# ---------------------------------------------------------------------------
# GET /api/projects
# ---------------------------------------------------------------------------

def test_list_projects_row_shape(client):
    """The list endpoint returns EXACTLY the board projection -- no more, no
    less. The full row (dates, folder path, coordinates, revision, ...) is a
    single-project concern served by GET /api/projects/<id>; the projection
    exists to keep the 300+-row board payload small, so a key added here must
    be a deliberate main._PROJECT_LIST_FIELDS change, not a raw-row leak."""
    import main

    create_project(client, "ROWSHAPE-1")
    resp = client.get("/api/projects")
    assert resp.status_code == 200
    rows = resp.get_json()
    assert len(rows) == 1
    assert set(rows[0].keys()) == set(main._PROJECT_LIST_FIELDS)
    # Card 1B widened the projection deliberately: four DERIVED lead-card
    # fields (no stored column, no extra query -- see
    # workflow.projects._annotate_card_state). Pinned by name and by shape so a
    # later card can't quietly drop one the board renders from.
    row = rows[0]
    assert row["display_stage"] == "Lead Assessment"      # the stored stage group itself
    # Creation auto-assignment: the configured rule assignees (Tahira on
    # Seismic Signature Validation, then the Saad/Salem Pre-Well picks) appear
    # on a brand-new lead; the anonymous creator does not.
    assert row["assignees"][0] == "Tahira"
    assert set(row["assignees"][1:]) <= {"Saad", "Salem"}
    assert row["lead_priority"] in ("High", "Medium", "Low")
    assert len(row["tracked_items"]) == 12
    # Card 2A widened each ITEM by one key: `steps`, the item's source step
    # names, so the lead detail page's three-stage sidebar can open the real
    # step behind an item without re-implementing the _TRACKED_ITEMS mapping.
    assert all(set(item) == {"stage", "label", "status", "steps"} for item in row["tracked_items"])
    assert all(isinstance(item["steps"], list) for item in row["tracked_items"])
    # Card 1C widened it by one more DERIVED key: the record's field, feeding
    # the lead board's Field filter (there is no stored field column).
    assert row["field"] == "ROWSHAPE"
    # Card 1E widened it by one more: the latest saved Mean Gas (BCF), derived
    # from the assessment steps' dynamic fields. A brand-new lead has recorded
    # none, so the payload carries an explicit null (the KPI tile reads 0) --
    # the key is always PRESENT, which is what the tile can rely on.
    assert "mean_gas_bcf" in row
    assert row["mean_gas_bcf"] is None
    # The single-project route stays full-row (the editor/detail surfaces
    # read promotion flags, dates and folder path from it).
    full = client.get(f"/api/projects/{rows[0]['project_id']}").get_json()
    for key in ("start_date", "target_date", "lead_folder_path", "revision",
                "business_plan_enabled"):
        assert key in full, key


def test_list_projects_search_filter(client):
    create_project(client, "ALPHA-1")
    create_project(client, "BETA-2")
    resp = client.get("/api/projects?search=alpha")
    names = [p["project_name"] for p in resp.get_json()]
    assert names == ["ALPHA-1"]


def test_list_projects_pipeline_filter(client):
    create_project(client, "PROSPECT-A")
    create_project(client, "BP-A", pipeline_type="bp", business_plan_enabled=True, business_plan_year=2030)
    resp_bp = client.get("/api/projects?pipeline_filter=bp")
    assert [p["project_name"] for p in resp_bp.get_json()] == ["BP-A"]
    resp_prospect = client.get("/api/projects?pipeline_filter=prospect")
    assert [p["project_name"] for p in resp_prospect.get_json()] == ["PROSPECT-A"]


def test_list_projects_field_is_the_folder_field_derivation(client):
    """`field` must be the SAME split folders.py builds share paths from.

    There is no stored field column: the field is the first segment of the
    record name. Pinning it against folders.parse_field_and_well keeps the
    board's Field filter and the folder links from drifting into two
    conventions.
    """
    import folders

    for name in ("GALV-2", "LUNA-10", "SOLO"):
        create_project(client, name)
    rows = {row["project_name"]: row["field"] for row in client.get("/api/projects").get_json()}
    for name, field in rows.items():
        assert field == folders.parse_field_and_well(name)[0]
    assert rows == {"GALV-2": "GALV", "LUNA-10": "LUNA", "SOLO": "SOLO"}


def test_list_projects_include_completed_is_opt_in(client):
    """A matured lead leaves the board by default and returns only on request.

    Card 1C's lead board offers an explicit Completed status filter and asks
    for those leads with include_completed=1; every other caller (the BP board,
    the portfolio, the tests) keeps the historical behaviour.
    """
    from test_portfolio import _approve_all_prospect_tasks

    done_pid = create_project(client, "MATURED-1")
    open_pid = create_project(client, "OPEN-1")
    _approve_all_prospect_tasks(client, done_pid)

    default_ids = [r["project_id"] for r in client.get("/api/projects?pipeline_filter=prospect").get_json()]
    assert default_ids == [open_pid]

    rows = client.get("/api/projects?pipeline_filter=prospect&include_completed=1").get_json()
    by_id = {row["project_id"]: row for row in rows}
    assert set(by_id) == {open_pid, done_pid}
    assert by_id[done_pid]["overall_status"] == "Completed"
    # It lands in the board's last column, with every tracked item done -- the
    # board can render it without a special case.
    assert by_id[done_pid]["display_stage"] == "Pre-Well Delivery"
    # Since v5 every tracked item has a real step, so a fully approved lead
    # reads ALL twelve done (pre-v5 two items were pinned at In Progress).
    assert {item["status"] for item in by_id[done_pid]["tracked_items"]} == {"Completed"}


# ---------------------------------------------------------------------------
# GET /api/projects/<id>
# ---------------------------------------------------------------------------

def test_get_project_ok(client):
    pid = create_project(client, "GETME-1")
    resp = client.get(f"/api/projects/{pid}")
    assert resp.status_code == 200
    assert resp.get_json()["project_id"] == pid


def test_get_project_not_found(client):
    resp = client.get("/api/projects/999999")
    assert resp.status_code == 404
    assert resp.get_json()["detail"] == "Lead / well not found"


# ---------------------------------------------------------------------------
# GET /api/projects/<id>/detail
# ---------------------------------------------------------------------------

def test_project_detail_shape(client):
    pid = create_project(client, "DETAIL-1")
    resp = client.get(f"/api/projects/{pid}/detail")
    assert resp.status_code == 200
    body = resp.get_json()
    for key in ("project", "tasks", "completion", "fields", "lead_summary", "overview"):
        assert key in body
    assert "percent" in body["completion"]
    assert body["lead_summary"] is None  # never promoted
    # overview is composed from the task inputs at read time (no stored
    # project_overview table); derisking carries the computed Total CoS.
    assert isinstance(body["overview"], dict)
    assert "derisking" in body["overview"]
    # Card 2A: the detail payload's project row is the FULL project dict, so it
    # already carries the same derived card fields the board rows do
    # (get_project -> _annotate_derived_state -> _annotate_card_state). The
    # lead detail page's three-stage sidebar and its Lead Summary progress bar
    # read tracked_items straight from here -- same derivation as the board, so
    # the two surfaces can never disagree. Pinned so the detail payload is not
    # narrowed to the board's projection without noticing.
    assert len(body["project"]["tracked_items"]) == 12
    assert all(set(item) == {"stage", "label", "status", "steps"}
               for item in body["project"]["tracked_items"])
    assert body["project"]["display_stage"] == "Lead Assessment"


# ---------------------------------------------------------------------------
# PATCH /api/projects/<id>/rename
# ---------------------------------------------------------------------------

def test_rename_project_ok(client):
    pid = create_project(client, "RENAME-1")
    resp = client.patch(f"/api/projects/{pid}/rename", json={"new_name": "RENAME-1-NEW"})
    assert resp.status_code == 200
    assert resp.get_json()["ok"] is True


def test_rename_project_empty(client):
    pid = create_project(client, "RENAME-2")
    resp = client.patch(f"/api/projects/{pid}/rename", json={"new_name": ""})
    assert resp.status_code == 400


def test_rename_project_too_long(client):
    pid = create_project(client, "RENAME-3")
    resp = client.patch(f"/api/projects/{pid}/rename", json={"new_name": "B" * 121})
    assert resp.status_code == 400


def test_rename_project_duplicate_name_rejected(client):
    pid = create_project(client, "RENAME-DUP-1")
    create_project(client, "RENAME-DUP-2")
    resp = client.patch(f"/api/projects/{pid}/rename", json={"new_name": "RENAME-DUP-2"})
    assert resp.status_code == 400
    assert "already exists" in resp.get_json()["detail"]
    assert client.get(f"/api/projects/{pid}").get_json()["project_name"] == "RENAME-DUP-1"


# ---------------------------------------------------------------------------
# DELETE / PATCH restore
# ---------------------------------------------------------------------------

def test_delete_archives_and_restore_brings_back(client):
    pid = create_project(client, "ARCHIVE-1")
    resp = client.delete(f"/api/projects/{pid}")
    assert resp.status_code == 200
    assert resp.get_json() == {"ok": True, "archived": True}

    ids = [p["project_id"] for p in client.get("/api/projects").get_json()]
    assert pid not in ids

    resp = client.patch(f"/api/projects/{pid}/restore")
    assert resp.status_code == 200
    assert resp.get_json()["ok"] is True

    ids = [p["project_id"] for p in client.get("/api/projects").get_json()]
    assert pid in ids


# ---------------------------------------------------------------------------
# PATCH /api/tasks/<id> (save)
# ---------------------------------------------------------------------------

def test_save_task_ok(client):
    pid = create_project(client, "TASKSAVE-1")
    task = get_tasks(client, pid)[0]
    resp = client.patch(f"/api/tasks/{task['task_id']}", json={
        "status": "In Progress", "revision": task["revision"],
    })
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert "task" in body


def test_save_task_invalid_status(client):
    pid = create_project(client, "TASKSAVE-2")
    task = get_tasks(client, pid)[0]
    resp = client.patch(f"/api/tasks/{task['task_id']}", json={
        "status": "Bogus Status", "revision": task["revision"],
    })
    assert resp.status_code == 400


def test_save_task_stale_revision_conflict(client):
    pid = create_project(client, "TASKSAVE-3")
    task = get_tasks(client, pid)[0]
    stale_revision = task["revision"]
    resp1 = client.patch(f"/api/tasks/{task['task_id']}", json={
        "status": "In Progress", "revision": stale_revision,
    })
    assert resp1.status_code == 200
    resp2 = client.patch(f"/api/tasks/{task['task_id']}", json={
        "status": "Approved", "revision": stale_revision,
    })
    assert resp2.status_code == 409
    assert "detail" in resp2.get_json()


# ---------------------------------------------------------------------------
# POST /api/tasks/<id>/assign
# ---------------------------------------------------------------------------

def test_assign_task_shape_and_canonical_casing(client):
    pid = create_project(client, "ASSIGN-CONTRACT-1")
    task = get_tasks(client, pid)[0]
    # Lowercase on purpose: the response must carry the users-table casing.
    resp = client.post(f"/api/tasks/{task['task_id']}/assign", json={
        "assignee": "supervisor", "cascade": False, "revision": task["revision"],
    })
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["task"]["assigned_to"] == "Supervisor"
    assert body["task"]["status"] == "In Progress"
    assert body["task"]["revision"] == task["revision"] + 1


def test_assign_task_stale_revision_conflict(client):
    pid = create_project(client, "ASSIGN-CONTRACT-2")
    task = get_tasks(client, pid)[0]
    resp = client.post(f"/api/tasks/{task['task_id']}/assign", json={
        "assignee": "Supervisor", "revision": task["revision"] + 5,
    })
    assert resp.status_code == 409
    assert "detail" in resp.get_json()


def test_assign_task_unknown_assignee(client):
    pid = create_project(client, "ASSIGN-CONTRACT-3")
    task = get_tasks(client, pid)[0]
    resp = client.post(f"/api/tasks/{task['task_id']}/assign", json={
        "assignee": "Nobody In Particular", "revision": task["revision"],
    })
    assert resp.status_code == 400
    assert resp.get_json()["detail"] == "Unknown or inactive user."


# ---------------------------------------------------------------------------
# POST /api/tasks/<id>/transition
# ---------------------------------------------------------------------------

def test_transition_task_shape(client):
    pid = create_project(client, "TRANSITION-CONTRACT-1")
    task = get_tasks(client, pid)[0]
    assigned = client.post(f"/api/tasks/{task['task_id']}/assign", json={
        "assignee": "Supervisor", "cascade": False, "revision": task["revision"],
    }).get_json()["task"]
    resp = client.post(f"/api/tasks/{task['task_id']}/transition", json={
        "action": "submit", "revision": assigned["revision"],
    })
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["task"]["status"] == "Ready"
    assert body["task"]["revision"] == assigned["revision"] + 1


def test_transition_task_stale_revision_conflict(client):
    pid = create_project(client, "TRANSITION-CONTRACT-2")
    task = get_tasks(client, pid)[0]
    assigned = client.post(f"/api/tasks/{task['task_id']}/assign", json={
        "assignee": "Supervisor", "cascade": False, "revision": task["revision"],
    }).get_json()["task"]
    resp = client.post(f"/api/tasks/{task['task_id']}/transition", json={
        "action": "submit", "revision": assigned["revision"] + 5,
    })
    assert resp.status_code == 409
    assert "detail" in resp.get_json()


# ---------------------------------------------------------------------------
# GET /api/tasks/<id>
# ---------------------------------------------------------------------------

def test_get_task_ok(client):
    pid = create_project(client, "TASKGET-1")
    task = get_tasks(client, pid)[0]
    resp = client.get(f"/api/tasks/{task['task_id']}")
    assert resp.status_code == 200


def test_get_task_missing(client):
    resp = client.get("/api/tasks/999999")
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Dynamic fields round trip
# ---------------------------------------------------------------------------

def test_dynamic_fields_round_trip_wrapped_and_bare(client):
    pid = create_project(client, "DYNFIELDS-1")
    task = get_tasks(client, pid)[0]
    tid = task["task_id"]

    resp = client.patch(f"/api/tasks/{tid}/dynamic-fields", json={"fields": {"foo": "bar"}})
    assert resp.status_code == 200
    got = client.get(f"/api/tasks/{tid}/dynamic-fields").get_json()
    assert got == {"foo": "bar"}

    resp = client.patch(f"/api/tasks/{tid}/dynamic-fields", json={"baz": "qux"})
    assert resp.status_code == 200
    got = client.get(f"/api/tasks/{tid}/dynamic-fields").get_json()
    assert got == {"foo": "bar", "baz": "qux"}


# ---------------------------------------------------------------------------
# PATCH /api/tasks/<id>/priority
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("value,expected", [("Low", "Low"), ("Medium", "Medium"), ("High", "High")])
def test_set_task_priority_valid(client, value, expected):
    pid = create_project(client, f"PRIORITY-{value}")
    task = get_tasks(client, pid)[0]
    resp = client.patch(f"/api/tasks/{task['task_id']}/priority", json={"priority": value})
    assert resp.status_code == 200
    got = client.get(f"/api/tasks/{task['task_id']}").get_json()
    assert got["priority"] == expected


def test_set_task_priority_unknown_falls_back_to_medium(client):
    pid = create_project(client, "PRIORITY-BOGUS")
    task = get_tasks(client, pid)[0]
    # Actual behavior: the endpoint still returns 200 (it never validates and
    # rejects); invalid values are silently normalized to Medium.
    resp = client.patch(f"/api/tasks/{task['task_id']}/priority", json={"priority": "bogus"})
    assert resp.status_code == 200
    got = client.get(f"/api/tasks/{task['task_id']}").get_json()
    assert got["priority"] == "Medium"


# ---------------------------------------------------------------------------
# PATCH /api/projects/<id>/priority -- the LEAD-LEVEL priority (v9)
# ---------------------------------------------------------------------------

def _priority_history(client, project_id):
    conn = raw_sqlite_connect(client.db_path)
    try:
        return [tuple(row) for row in conn.execute("""
            SELECT task_name, old_status, new_status, changed_by, comment
            FROM task_history
            WHERE project_id = ? AND action_type = 'Priority Changed'
            ORDER BY history_id
        """, (project_id,))]
    finally:
        conn.close()


def test_set_project_priority_updates_stored_value_and_logs_one_event(client):
    pid = create_project(client, "LEADPRIO-1")
    resp = client.patch(f"/api/projects/{pid}/priority",
                        json={"priority": "high", "changed_by": "Supervisor"})
    assert resp.status_code == 200
    # Input is normalized (title-cased) and echoed back.
    assert resp.get_json() == {"ok": True, "priority": "High"}
    # The detail payload's project dict carries the stored priority, and both
    # legacy derived keys now read the SAME stored value.
    project = client.get(f"/api/projects/{pid}/detail").get_json()["project"]
    assert project["priority"] == "High"
    assert project["lead_priority"] == "High"
    assert project["current_task_priority"] == "High"
    # Exactly ONE history event, anchored on the first active task (the same
    # anchor "Lead Created" uses), with the old and new values.
    assert _priority_history(client, pid) == [
        ("Lead Assessment", "Low", "High", "Supervisor", "Priority set to High.")]


def test_set_project_priority_rejects_invalid_values(client):
    import db as dbmod
    import workflow

    pid = create_project(client, "LEADPRIO-2")
    # Route: unlike the legacy per-task endpoint, an unknown value is REJECTED
    # (400), never silently defaulted.
    resp = client.patch(f"/api/projects/{pid}/priority", json={"priority": "Urgent"})
    assert resp.status_code == 400
    assert client.get(f"/api/projects/{pid}").get_json()["priority"] == "Low"
    assert _priority_history(client, pid) == []
    # Domain: the same rejection is a ValueError at the function boundary.
    session = dbmod.get_session()
    with pytest.raises(ValueError, match="Priority must be Low, Medium or High."):
        workflow.set_project_priority(session, pid, "bogus")
    with pytest.raises(ValueError, match="Priority must be Low, Medium or High."):
        workflow.set_project_priority(session, pid, None)


def test_set_project_priority_unchanged_value_writes_nothing(client):
    pid = create_project(client, "LEADPRIO-3")
    before = client.get(f"/api/projects/{pid}").get_json()["last_updated"]
    resp = client.patch(f"/api/projects/{pid}/priority", json={"priority": "Low"})
    assert resp.status_code == 200
    assert resp.get_json() == {"ok": True, "priority": "Low"}
    after = client.get(f"/api/projects/{pid}").get_json()
    assert after["priority"] == "Low"
    assert after["last_updated"] == before   # the no-op touched nothing
    assert _priority_history(client, pid) == []


def test_set_project_priority_missing_project_is_404(client):
    resp = client.patch("/api/projects/424242/priority", json={"priority": "High"})
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Completion
# ---------------------------------------------------------------------------

def test_completion_percent_zero_for_new_project(client):
    pid = create_project(client, "COMPLETION-1")
    resp = client.get(f"/api/projects/{pid}/completion")
    assert resp.status_code == 200
    assert resp.get_json() == {"percent": 0.0}


# ---------------------------------------------------------------------------
# Business plan rows / portfolio
# ---------------------------------------------------------------------------

def test_business_plan_rows_and_portfolio_rows(client):
    create_project(client, "BPROW-1", pipeline_type="bp", business_plan_enabled=True, business_plan_year=2028)
    resp = client.get("/api/business-plan/rows")
    assert resp.status_code == 200
    assert isinstance(resp.get_json(), list)

    resp = client.get("/api/portfolio/rows")
    assert resp.status_code == 200
    body = resp.get_json()
    assert "rows" in body
    assert "summary" in body
    assert "business_plan_wells" in body["summary"]
    assert "cumulative_ogip" in body["summary"]
    # WS7: every row carries exactly the 8 analysis columns (plus internals).
    assert len(body["rows"]) == 1
    row = body["rows"][0]
    for key in ("well_name", "gas_field", "seismic_block", "classification",
                "year", "fluid", "mean_ogip", "total_cos"):
        assert key in row, key


def test_portfolio_rows_invalid_year(client):
    # Floor is 1990 (admits imported historical wells); 1989 stays invalid.
    resp = client.get("/api/portfolio/rows?year=1989")
    assert resp.status_code == 400
    assert "detail" in resp.get_json()


# ---------------------------------------------------------------------------
# Activity log
# ---------------------------------------------------------------------------

def test_activity_contains_lead_created(client):
    pid = create_project(client, "ACTIVITY-1")
    resp = client.get("/api/activity")
    assert resp.status_code == 200
    events = resp.get_json()
    assert any(e["action_type"] == "Lead Created" and e["project_id"] == pid for e in events)


def test_activity_filters_by_project_id(client):
    pid1 = create_project(client, "ACTIVITY-A")
    pid2 = create_project(client, "ACTIVITY-B")
    resp = client.get(f"/api/activity?project_id={pid1}")
    events = resp.get_json()
    assert len(events) >= 1
    assert all(e["project_id"] == pid1 for e in events)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def test_export_excel(client):
    create_project(client, "EXPORT-1")
    # A BP well with a saved flowback field exercises the Portfolio Export
    # composer's task-input read path (WS4), not just an empty sheet.
    bp_pid = create_project(client, "EXPORT-BP-1", pipeline_type="bp",
                             business_plan_enabled=True, business_plan_year=2030)
    flowback_task = get_task_by_name(client, bp_pid, "Flowback Results")
    resp = client.patch(f"/api/tasks/{flowback_task['task_id']}/dynamic-fields",
                         json={"fields": {"flowback_dynamic_ogip_bcf": "12.5"}})
    assert resp.status_code == 200

    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    assert resp.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = openpyxl.load_workbook(io.BytesIO(resp.data))
    assert workbook.sheetnames == [
        "Executive Summary", "Wells Overview", "Monthly Progress",
        "Portfolio Export", "Staking Options",
    ]

    ws_export = workbook["Portfolio Export"]
    header = [cell.value for cell in ws_export[4]]
    assert header == portfolio_export.PORTFOLIO_EXPORT_COLUMNS

    well_col = header.index("Well Name")
    dyn_mean_col = header.index("Dynamic Mean (BCF)")
    matching_rows = [row for row in ws_export.iter_rows(min_row=5, max_row=ws_export.max_row, values_only=True)
                     if row[well_col] == "EXPORT-BP-1"]
    assert len(matching_rows) == 1
    assert matching_rows[0][dyn_mean_col] == "12.5"

    ws_staking = workbook["Staking Options"]
    staking_header = [cell.value for cell in ws_staking[4]]
    assert staking_header == portfolio_export.STAKING_EXPORT_COLUMNS

    # The choke column header matches the field's inch scale (flowback_choke_size_in
    # and the UI's 'Choke Size (in)'), not the legacy 1/64" convention.
    assert "Choke Size (in)" in header


def test_export_does_not_execute_raw_sql_through_pandas(client, monkeypatch):
    """Production can combine pandas with SQLAlchemy 2.x, where a plain SQL
    string is not executable. The export must use the dialect-safe db helper."""
    import export_excel

    create_project(client, "EXPORT-SQL-1")

    def reject_raw_pandas_sql(*_args, **_kwargs):
        raise AssertionError("Export bypassed db.fetch_all")

    monkeypatch.setattr(export_excel.pd, "read_sql_query", reject_raw_pandas_sql)
    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    assert resp.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.close()


def test_export_reservoir_cos_columns_read_one_primary_row(client):
    """A blank leading reservoir_cos_rows row must not split the export between
    vintages: Pull-up/Amplitude/BTS/CoS AND the Seismic Block AR all come from
    the ONE primary row (first row with a CoS pct or AR number filled), so the
    Excel row matches what the Portfolio UI derives."""
    bp_pid = create_project(client, "EXPORT-BP-2", pipeline_type="bp",
                             business_plan_enabled=True, business_plan_year=2030)
    reservoir_task = get_task_by_name(client, bp_pid, "Reservoir CoS")
    resp = client.patch(f"/api/tasks/{reservoir_task['task_id']}/dynamic-fields",
                         json={"fields": {"reservoir_cos_rows": json.dumps([
                             {"seismic_volume_ar_number": "", "reservoir_cos_pct": "",
                              "pull_up": "", "amplitude_ratio": "", "base_tight_sarah": ""},
                             {"seismic_volume_ar_number": "AR-777", "reservoir_cos_pct": "55",
                              "pull_up": "Yes", "amplitude_ratio": "0.8", "base_tight_sarah": "0.6"},
                         ])}})
    assert resp.status_code == 200

    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws_export = workbook["Portfolio Export"]
    header = [cell.value for cell in ws_export[4]]
    matching_rows = [row for row in ws_export.iter_rows(min_row=5, max_row=ws_export.max_row, values_only=True)
                     if row[header.index("Well Name")] == "EXPORT-BP-2"]
    assert len(matching_rows) == 1
    row = matching_rows[0]
    assert row[header.index("Pull-up")] == "Yes"
    assert row[header.index("Amplitude Ratio")] == "0.8"
    assert row[header.index("BTS")] == "0.6"
    assert row[header.index("Reservoir CoS (%)")] == "55"
    # AR-777 is unmapped in the shipped block map, so the raw AR shows -- and it
    # comes from the SAME row as the four values above, never a different one.
    assert row[header.index("Seismic Block")] == "AR-777"
    assert row[header.index("AR Number")] == "AR-777"


def test_export_flowback_columns_read_primary_stage(client):
    """The flowback columns come from the FIRST non-empty stage of the
    flowback_stages_rows mini-sheet, as one unit: a well that also carries the
    retired flat keys must not mix them in (single-vintage rule, mirroring the
    Reservoir CoS primary row) -- even for a measurement the stage left blank."""
    bp_pid = create_project(client, "EXPORT-BP-FB1", pipeline_type="bp",
                             business_plan_enabled=True, business_plan_year=2030)
    flowback_task = get_task_by_name(client, bp_pid, "Flowback Results")
    resp = client.patch(f"/api/tasks/{flowback_task['task_id']}/dynamic-fields",
                         json={"fields": {
                             "flowback_stages_rows": json.dumps([
                                 {"flowback_gas_rate_mmscfd": "", "flowback_water_rate_bwpd": "",
                                  "flowback_choke_size_in": "", "flowback_fwhp_psi": ""},
                                 {"flowback_gas_rate_mmscfd": "9.5", "flowback_water_rate_bwpd": "120",
                                  "flowback_liquid_rate_bpd": "75",
                                  "flowback_choke_size_in": "0.5", "flowback_fwhp_psi": ""},
                                 {"flowback_gas_rate_mmscfd": "4.2"},
                             ]),
                             # Retired flat keys: must lose to the stage row wholesale.
                             "flowback_gas_rate_mmscfd": "1.1",
                             "flowback_fwhp_psi": "888",
                         }})
    assert resp.status_code == 200

    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws_export = workbook["Portfolio Export"]
    header = [cell.value for cell in ws_export[4]]
    matching_rows = [row for row in ws_export.iter_rows(min_row=5, max_row=ws_export.max_row, values_only=True)
                     if row[header.index("Well Name")] == "EXPORT-BP-FB1"]
    assert len(matching_rows) == 1
    row = matching_rows[0]
    assert row[header.index("Gas Rate (MMSCFD)")] == "9.5"
    assert row[header.index("Water Rate (BWPD)")] == "120"
    assert row[header.index("Condensate Rate (BPD)")] == "75"
    assert row[header.index("Choke Size (in)")] == "0.5"
    # The primary stage left WHP blank; the flat key's 888 must NOT leak in.
    assert row[header.index("WHP (psi)")] in (None, "")


def test_export_flowback_depth_only_stage_is_not_primary(client):
    """A stage row carrying only Top/Base depths (the stage sheet's first two
    columns) is not data-bearing: the export's primary-stage pick still lands
    on the first row with a rate/choke/WHP measurement, and the depth keys
    round-trip untouched inside the stored flowback_stages_rows blob."""
    bp_pid = create_project(client, "EXPORT-BP-FB3", pipeline_type="bp",
                             business_plan_enabled=True, business_plan_year=2030)
    flowback_task = get_task_by_name(client, bp_pid, "Flowback Results")
    resp = client.patch(f"/api/tasks/{flowback_task['task_id']}/dynamic-fields",
                         json={"fields": {
                             "flowback_stages_rows": json.dumps([
                                 {"flowback_top_md": "11200", "flowback_base_md": "11450"},
                                 {"flowback_top_md": "11500", "flowback_base_md": "11720",
                                  "flowback_gas_rate_mmscfd": "6.1", "flowback_choke_size_in": "0.5"},
                             ]),
                         }})
    assert resp.status_code == 200

    stored = client.get(f"/api/tasks/{flowback_task['task_id']}/dynamic-fields").get_json()
    rows = json.loads(stored["flowback_stages_rows"])
    assert rows[0] == {"flowback_top_md": "11200", "flowback_base_md": "11450"}

    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws_export = workbook["Portfolio Export"]
    header = [cell.value for cell in ws_export[4]]
    matching_rows = [row for row in ws_export.iter_rows(min_row=5, max_row=ws_export.max_row, values_only=True)
                     if row[header.index("Well Name")] == "EXPORT-BP-FB3"]
    assert len(matching_rows) == 1
    row = matching_rows[0]
    assert row[header.index("Gas Rate (MMSCFD)")] == "6.1"
    assert row[header.index("Choke Size (in)")] == "0.5"


def test_export_flowback_columns_fall_back_to_flat_keys(client):
    """A well written before the stages mini-sheet existed (retired flat keys
    only, no flowback_stages_rows) still fills the flowback columns."""
    bp_pid = create_project(client, "EXPORT-BP-FB2", pipeline_type="bp",
                             business_plan_enabled=True, business_plan_year=2030)
    flowback_task = get_task_by_name(client, bp_pid, "Flowback Results")
    resp = client.patch(f"/api/tasks/{flowback_task['task_id']}/dynamic-fields",
                         json={"fields": {
                             "flowback_gas_rate_mmscfd": "3.3",
                             "flowback_water_rate_bwpd": "450",
                             "flowback_liquid_rate_bpd": "60",
                             "flowback_choke_size_in": "0.75",
                             "flowback_fwhp_psi": "2100",
                         }})
    assert resp.status_code == 200

    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws_export = workbook["Portfolio Export"]
    header = [cell.value for cell in ws_export[4]]
    matching_rows = [row for row in ws_export.iter_rows(min_row=5, max_row=ws_export.max_row, values_only=True)
                     if row[header.index("Well Name")] == "EXPORT-BP-FB2"]
    assert len(matching_rows) == 1
    row = matching_rows[0]
    assert row[header.index("Gas Rate (MMSCFD)")] == "3.3"
    assert row[header.index("Water Rate (BWPD)")] == "450"
    assert row[header.index("Condensate Rate (BPD)")] == "60"
    assert row[header.index("Choke Size (in)")] == "0.75"
    assert row[header.index("WHP (psi)")] == "2100"


def test_export_xy_columns_and_fixed_layout(client):
    """X/Y open the sheet with the project's lead coordinates, AR Number sits
    8th (right after Seismic Block), and Condensate Rate follows Water Rate --
    the fixed column layout external consumers of the sheet rely on."""
    cols = portfolio_export.PORTFOLIO_EXPORT_COLUMNS
    assert cols[:2] == ["X", "Y"]
    assert cols[6] == "Seismic Block"
    assert cols[7] == "AR Number"
    assert cols[cols.index("Water Rate (BWPD)") + 1] == "Condensate Rate (BPD)"

    create_project(client, "EXPORT-XY-1", lead_x=512345.5, lead_y=2871000)
    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws_export = workbook["Portfolio Export"]
    header = [cell.value for cell in ws_export[4]]
    matching_rows = [row for row in ws_export.iter_rows(min_row=5, max_row=ws_export.max_row, values_only=True)
                     if row[header.index("Well Name")] == "EXPORT-XY-1"]
    assert len(matching_rows) == 1
    row = matching_rows[0]
    assert float(row[header.index("X")]) == 512345.5
    assert float(row[header.index("Y")]) == 2871000


def test_export_includes_proposed_leads_with_latest_estimates(client):
    """The Portfolio Export sheet is one row per NON-ARCHIVED project: a
    still-maturing (Proposed) lead must appear, its estimate columns filled
    from the latest available (lead-phase) inputs; a just-created bare lead
    must appear too. The Staking Options sheet now shares that same
    non-archived membership (it filters reporting._portfolio_projects to
    business_plan_enabled == 0, which is every lead once that reader was
    widened), so both leads must show there too."""
    bare_pid = create_project(client, "EXPORT-LEAD-BARE")
    lead_pid = create_project(client, "EXPORT-LEAD-1")
    lead_ra_task = get_task_by_name(client, lead_pid, "Lead Assessment")
    resp = client.patch(f"/api/tasks/{lead_ra_task['task_id']}/dynamic-fields",
                         json={"fields": {"lead_piip_gas_p90": "3.1",
                                          "lead_piip_gas_mean": "7.5",
                                          "lead_piip_gas_p10": "15.2"}})
    assert resp.status_code == 200
    area_task = lead_ra_task
    resp = client.patch(f"/api/tasks/{area_task['task_id']}/dynamic-fields",
                         json={"fields": {"p90_area_km2": "2.4", "p10_area_km2": "9.8"}})
    assert resp.status_code == 200
    thickness_task = lead_ra_task
    resp = client.patch(f"/api/tasks/{thickness_task['task_id']}/dynamic-fields",
                         json={"fields": {"reservoir_thickness_ft": "88"}})
    assert resp.status_code == 200

    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws_export = workbook["Portfolio Export"]
    header = [cell.value for cell in ws_export[4]]
    by_name = {row[header.index("Well Name")]: row
               for row in ws_export.iter_rows(min_row=5, max_row=ws_export.max_row, values_only=True)}

    lead_row = by_name["EXPORT-LEAD-1"]
    assert lead_row[header.index("Status")] == "Proposed"
    assert lead_row[header.index("BP Year")] in ("", None)
    assert lead_row[header.index("OGIP P90 (BCF)")] == "3.1"
    assert lead_row[header.index("OGIP Mean (BCF)")] == "7.5"
    assert lead_row[header.index("OGIP P10 (BCF)")] == "15.2"
    assert lead_row[header.index("P90 Area (km2)")] == "2.4"
    assert lead_row[header.index("P10 Area (km2)")] == "9.8"
    # An undrilled lead has no SARH formation row: P50 Pay falls back to the
    # Thickness Estimation step's reservoir thickness estimate.
    assert lead_row[header.index("P50 Pay Thickness (ft)")] == "88"

    bare_row = by_name["EXPORT-LEAD-BARE"]
    assert bare_row[header.index("Status")] == "Proposed"
    assert bare_row[header.index("OGIP Mean (BCF)")] in ("", None)

    ws_staking = workbook["Staking Options"]
    staking_header = [cell.value for cell in ws_staking[4]]
    staking_by_name = {row[staking_header.index("Lead Name")]: row
                       for row in ws_staking.iter_rows(min_row=5, max_row=ws_staking.max_row, values_only=True)}
    assert "EXPORT-LEAD-1" in staking_by_name
    assert "EXPORT-LEAD-BARE" in staking_by_name
    # A membership assert alone would pass even if the row's content were
    # wrong: the bare lead was created with no lead_x/lead_y and has never
    # touched the Moving Tolerance step, so its X/Y must come back blank, not
    # some stray/junk coordinate.
    bare_staking_row = staking_by_name["EXPORT-LEAD-BARE"]
    assert bare_staking_row[staking_header.index("X")] in ("", None)
    assert bare_staking_row[staking_header.index("Y")] in ("", None)


def test_export_status_reads_sarh_formation_fluid(client):
    """The export Status column resolves fluid down the SARH-aware well-fluid
    ladder (reporting.resolve_well_fluid): a SARH 'final'-phase formation fluid
    shows as the record's Status, since the step-level fluid selects are gone."""
    bp_pid = create_project(client, "EXPORT-FLUID-1", pipeline_type="bp",
                             business_plan_enabled=True, business_plan_year=2030)
    resp = client.put(f"/api/projects/{bp_pid}/formations",
                      json={"phase": "final", "rows": [{"formation": "SARH", "fluid": "Oil"}]})
    assert resp.status_code == 200, resp.get_json()

    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws_export = workbook["Portfolio Export"]
    header = [cell.value for cell in ws_export[4]]
    matching_rows = [row for row in ws_export.iter_rows(min_row=5, max_row=ws_export.max_row, values_only=True)
                     if row[header.index("Well Name")] == "EXPORT-FLUID-1"]
    assert len(matching_rows) == 1
    assert matching_rows[0][header.index("Status")] == "Oil"


# ---------------------------------------------------------------------------
# Response compression (stdlib gzip after_request)
# ---------------------------------------------------------------------------

def test_json_gzipped_only_for_accepting_clients(client):
    """A large JSON response is gzipped exactly when the client advertises
    gzip support; the decompressed bytes are the identical JSON either way,
    and Vary: Accept-Encoding is set on both so caches can't cross the two."""
    import gzip as gz

    for i in range(10):
        create_project(client, f"GZIP-{i}")

    plain = client.get("/api/projects")
    assert plain.headers.get("Content-Encoding") is None
    assert "Accept-Encoding" in (plain.headers.get("Vary") or "")

    zipped = client.get("/api/projects", headers={"Accept-Encoding": "gzip"})
    assert zipped.headers.get("Content-Encoding") == "gzip"
    assert "Accept-Encoding" in (zipped.headers.get("Vary") or "")
    assert gz.decompress(zipped.data) == plain.data
    assert len(zipped.data) < len(plain.data)


def test_small_json_and_file_downloads_stay_uncompressed(client):
    """Bodies under the size floor and non-JSON downloads are left alone --
    the Excel file is already a compressed container and streams via
    direct_passthrough."""
    small = client.get("/api/health", headers={"Accept-Encoding": "gzip"})
    assert small.headers.get("Content-Encoding") is None

    export = client.get("/api/export/excel", headers={"Accept-Encoding": "gzip"})
    assert export.status_code == 200
    assert export.headers.get("Content-Encoding") is None
    # Still a valid xlsx (zip magic), not double-compressed.
    assert export.data[:2] == b"PK"
