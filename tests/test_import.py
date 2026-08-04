"""Tests for the Excel importer (import_excel.py).

Workbooks are built in-memory with openpyxl (written to tmp_path .xlsx files),
covering headers at Excel row 1 (hand-made sheets) AND row 4 (our own export's
startrow=3 layout) to exercise the autodetect. Every test drives the pure
functions (parse_workbook / classify_row / import_rows) against a session on the
per-test sqlite DB the `client` fixture stands up.

Two behaviors are owned by a parallel workstream and are RELIED ON here:
  (a) get_projects(pipeline_filter='bp') excludes completed BP wells (all BP
      steps Approved), and
  (b) the business-plan year guard accepts 1990-2040 (so a historical well with
      year 2019 promotes) -- imports pass allow_historical_year=True, the
      escape hatch that skips the promotion-only current-year floor.
Where a case depends on those, it is noted in a comment.
"""
from __future__ import annotations

import json

import openpyxl
import pytest

import portfolio_export
from portfolio_export import PORTFOLIO_EXPORT_COLUMNS


# ---------------------------------------------------------------------------
# Workbook / session helpers
# ---------------------------------------------------------------------------

def _write_sheet(path, rows, header_row=1):
    """Write ``rows`` (list of {column: value} dicts) under the full
    PORTFOLIO_EXPORT_COLUMNS header placed at Excel ``header_row``."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for col_index, col in enumerate(PORTFOLIO_EXPORT_COLUMNS, start=1):
        sheet.cell(row=header_row, column=col_index, value=col)
    for row_offset, record in enumerate(rows, start=header_row + 1):
        for col_index, col in enumerate(PORTFOLIO_EXPORT_COLUMNS, start=1):
            value = record.get(col)
            if value is not None and value != "":
                sheet.cell(row=row_offset, column=col_index, value=value)
    workbook.save(str(path))
    return str(path)


def _session(app_modules):
    _main, db = app_modules
    return db.new_session()


def _pid(session, name):
    import db as db_module
    row = db_module.fetch_one(session, "SELECT project_id FROM projects WHERE project_name = :n",
                              {"n": name})
    return row["project_id"] if row else None


def _tasks_by_name(session, name):
    import workflow
    pid = _pid(session, name)
    return {t["task_name"]: t for t in workflow.get_project_tasks(session, pid)}


def _export_by_name(session):
    return {r["Well Name"]: r for r in portfolio_export.get_portfolio_export_rows(session)}


# ---------------------------------------------------------------------------
# Case 1: one row per record type -> correct pipeline placement
# ---------------------------------------------------------------------------

def test_four_record_types_placed_correctly(client, app_modules, tmp_path):
    import import_excel
    import reporting
    import workflow

    rows = [
        # historical: BP Year < 2026 (relies on the 1990-2040 year guard).
        {"Well Name": "HIST-1", "BP Year": 2019, "P90 Area (km2)": 2, "P10 Area (km2)": 6},
        # bp well with a fluid status.
        {"Well Name": "BPWL-2", "BP Year": 2027, "Status": "Gas", "Classification": "Exploration",
         "OGIP Mean (BCF)": 10, "Reservoir CoS (%)": 45, "Trap CoS (%)": 50,
         "P50 Pay Thickness (ft)": 40, "Gas Rate (MMSCFD)": 8},
        # mature lead: Staked, no BP Year.
        {"Well Name": "MATR-3", "Status": "Staked", "P90 Area (km2)": 3},
        # proposed lead: partial data only.
        {"Well Name": "PROP-4", "Status": "Proposed", "P90 Area (km2)": 4, "Reservoir CoS (%)": 30},
    ]
    _write_sheet(tmp_path / "s.xlsx", rows, header_row=1)

    session = _session(app_modules)
    try:
        parsed = import_excel.parse_workbook(str(tmp_path / "s.xlsx"))
        report = import_excel.import_rows(session, parsed)

        outcomes = {r.well_name: r.outcome for r in report.results}
        assert outcomes == {"HIST-1": "created", "BPWL-2": "created",
                            "MATR-3": "created", "PROP-4": "created"}, [(r.well_name, r.outcome, r.reason) for r in report.results]

        # proposed lead -> on the Prospect board, its data-bearing steps Approved.
        prospect_names = {p["project_name"] for p in workflow.get_projects(session, pipeline_filter="prospect")}
        assert "PROP-4" in prospect_names
        prop_tasks = _tasks_by_name(session, "PROP-4")
        assert prop_tasks["Lead Assessment"]["status"] == "Approved"
        assert prop_tasks["Reservoir CoS"]["status"] == "Approved"
        # A step with no imported data stays open (not fully matured).
        assert prop_tasks["Trap and Seal CoS"]["status"] != "Approved"

        # mature lead -> off the Prospect board, in the Portfolio as 'Staked'.
        assert "MATR-3" not in prospect_names
        portfolio = {r["well_name"]: r for r in reporting.get_portfolio_rows(session)["rows"]}
        assert "MATR-3" in portfolio
        assert portfolio["MATR-3"]["status"] == "Staked"

        # bp well -> on the BP board and in the Portfolio with fluid-derived status.
        bp_names = {p["project_name"] for p in workflow.get_projects(session, pipeline_filter="bp")}
        assert "BPWL-2" in bp_names
        assert portfolio["BPWL-2"]["status"] == "Gas"

        # historical -> in the Portfolio, and OFF the BP board (all BP steps
        # Approved -> completed; relies on the completed-wells-exit rule).
        assert "HIST-1" in portfolio
        assert "HIST-1" not in bp_names

        # The escape hatch: HIST-1's 2019 year is well before today, yet the
        # import path (allow_historical_year=True) still enabled it -- a
        # promotion through the UI/API would be rejected for the same year.
        import db as db_module
        hist_project = db_module.fetch_one(session,
                                           "SELECT business_plan_enabled, business_plan_year FROM projects WHERE project_name = 'HIST-1'",
                                           {})
        assert int(hist_project["business_plan_enabled"]) == 1
        assert int(hist_project["business_plan_year"]) == 2019
    finally:
        session.close()


def test_lead_columns_share_the_consolidated_assessment_task_and_export(client, app_modules, tmp_path):
    """Area, thickness and lead PIIP keys now have one active EAV owner.

    The export is intentionally field-key based and retired-inclusive, so the
    task merge changes no column names or values on the round trip.
    """
    import import_excel
    import workflow

    row = {
        "Well Name": "ONE-LA-1", "Status": "Proposed",
        "P90 Area (km2)": 4, "P10 Area (km2)": 12,
        "SARH Formation Thickness (ft)": 85,
        "OGIP P90 (BCF)": 6, "OGIP Mean (BCF)": 10, "OGIP P10 (BCF)": 18,
    }
    _write_sheet(tmp_path / "one-la.xlsx", [row], header_row=1)

    session = _session(app_modules)
    try:
        result = import_excel.import_rows(
            session, import_excel.parse_workbook(str(tmp_path / "one-la.xlsx"))).results[0]
        assert result.outcome == "created", result.reason

        tasks = _tasks_by_name(session, "ONE-LA-1")
        assert "Lead Assessment" in tasks
        assert not ({"Area Definition", "Thickness Estimation", "GRV Inputs", "Resource Assessment"}
                    & set(tasks)), "retired checkpoint labels are not runnable tasks"
        fields = workflow.get_task_dynamic_fields(session, tasks["Lead Assessment"]["task_id"])
        assert fields["p90_area_km2"] == "4"
        assert fields["p10_area_km2"] == "12"
        assert fields["formation_thickness_ft"] == "85"
        assert fields["lead_piip_gas_mean"] == "10"

        exported = _export_by_name(session)["ONE-LA-1"]
        assert float(exported["P90 Area (km2)"]) == 4
        assert float(exported["P10 Area (km2)"]) == 12
        assert float(exported["SARH Formation Thickness (ft)"]) == 85
        assert float(exported["OGIP Mean (BCF)"]) == 10
    finally:
        session.close()


def test_export_folds_v7_retired_lead_rows_then_prefers_active_values(client, app_modules):
    """Pre-v7 EAV remains export-visible without letting history beat new data."""
    import db as db_module
    import workflow

    pid = create_project_for_import_test(client, "V7-EXPORT-LEGACY")
    session = _session(app_modules)
    try:
        with db_module.write_transaction(session):
            legacy_groups = [
                (1, "Area Definition", {"p90_area_km2": "4", "p10_area_km2": "12"}),
                (2, "Thickness Estimation", {"reservoir_thickness_ft": "70"}),
                (4, "Resource Assessment", {"lead_piip_gas_mean": "10"}),
            ]
            for sequence, task_name, fields in legacy_groups:
                retired = db_module.execute(session, """
                    INSERT INTO project_tasks
                        (project_id, sequence_no, task_name, stage_group, status, priority, is_active)
                    VALUES (:project_id, :sequence, :task_name, 'Lead Assessment',
                            'Approved', 'Low', 0)
                """, {"project_id": pid, "sequence": sequence,
                       "task_name": task_name}).lastrowid
                db_module.execute_many(session, """
                    INSERT INTO task_dynamic_fields (task_id, field_key, field_value)
                    VALUES (:task_id, :field_key, :field_value)
                """, [
                    {"task_id": retired, "field_key": key, "field_value": value}
                    for key, value in fields.items()
                ])

        row = _export_by_name(session)["V7-EXPORT-LEGACY"]
        assert row["P90 Area (km2)"] == "4"
        assert row["P10 Area (km2)"] == "12"
        assert row["P50 Pay Thickness (ft)"] == "70"
        assert row["OGIP Mean (BCF)"] == "10"

        active = _tasks_by_name(session, "V7-EXPORT-LEGACY")["Lead Assessment"]
        workflow.save_task_dynamic_fields(session, active["task_id"], {
            "p90_area_km2": "6", "p10_area_km2": "", "lead_piip_gas_mean": "15",
        }, changed_by="Test")
        row = _export_by_name(session)["V7-EXPORT-LEGACY"]
        assert row["P90 Area (km2)"] == "6", "active nonblank wins"
        assert row["P10 Area (km2)"] == "12", "active blank cannot erase history"
        assert row["P50 Pay Thickness (ft)"] == "70"
        assert row["OGIP Mean (BCF)"] == "15"
    finally:
        session.close()


def create_project_for_import_test(client, name):
    response = client.post("/api/projects", json={"project_name": name})
    assert response.status_code == 201, response.get_json()
    return response.get_json()["project_id"]


# ---------------------------------------------------------------------------
# Case 2: round-trip through portfolio_export (headers at row 4)
# ---------------------------------------------------------------------------

def test_round_trip_reproduces_key_cells(client, app_modules, tmp_path):
    import import_excel
    import workflow

    row = {
        "X": 512000, "Y": 2875000,
        "Well Name": "MDFT-7", "BP Year": 2027, "Classification": "Exploration",
        "Seismic Block": "Block A", "AR Number": 2525, "Status": "Gas",
        "Dynamic Mean (BCF)": 30, "Booked": "Yes",
        "OGIP P90 (BCF)": 5, "OGIP Mean (BCF)": 10, "OGIP P10 (BCF)": 18,
        "Condensate P90 (MMSTB)": 1, "Condensate Mean (MMSTB)": 2, "Condensate P10 (MMSTB)": 3,
        "Pull-up": "Semi", "Amplitude Ratio": 0.5, "BTS": 0.4, "Reservoir CoS (%)": 45,
        "P90 Area (km2)": 3, "P10 Area (km2)": 8, "SARH Formation Thickness (ft)": 100,
        "P50 Pay Thickness (ft)": 45, "P50 Porosity (%)": 12, "Water Saturation (%)": 30,
        "SARH-QWRH Thickness (ft)": 200, "Trap CoS (%)": 60,
        "Most Recent Age of Fault": 0.5, "Dip": 0.5, "Azimuth vs SHmax": 0.5,
        "Fault LoC": 0.5, "FPPM": 0.5,
        # Recompute from these inputs yields 25; the sheet pins 60 instead.
        "Seal CoS (%)": 60, "Pore Pressure Gradient (psi/ft)": 0.45,
        "Gas Rate (MMSCFD)": 8, "Water Rate (BWPD)": 200, "Condensate Rate (BPD)": 120,
        "Choke Size (in)": 0.5, "WHP (psi)": 3000,
    }
    _write_sheet(tmp_path / "exp.xlsx", [row], header_row=4)  # our export's layout

    session = _session(app_modules)
    try:
        parsed = import_excel.parse_workbook(str(tmp_path / "exp.xlsx"))
        report = import_excel.import_rows(session, parsed)
        assert [r.outcome for r in report.results] == ["created"], report.format()
        # The seal discrepancy is surfaced as a note.
        assert any("Seal CoS" in note for note in report.results[0].notes)

        exported = _export_by_name(session)["MDFT-7"]
        assert exported["BP Year"] == 2027
        assert exported["Status"] == "Gas"
        # X/Y land on projects.lead_x/lead_y and lead the exported row.
        assert float(exported["X"]) == 512000
        assert float(exported["Y"]) == 2875000
        # OGIP trio lands under the with-fluid prefix (resource_update).
        assert float(exported["OGIP P90 (BCF)"]) == 5
        assert float(exported["OGIP Mean (BCF)"]) == 10
        assert float(exported["OGIP P10 (BCF)"]) == 18
        assert float(exported["Condensate Mean (MMSTB)"]) == 2
        # Reservoir row values.
        assert exported["Pull-up"] == "Semi"
        assert float(exported["Amplitude Ratio"]) == 0.5
        assert float(exported["BTS"]) == 0.4
        assert float(exported["Reservoir CoS (%)"]) == 45
        # Seal CoS pinned to the sheet value (60), not the recomputed 25.
        assert float(exported["Seal CoS (%)"]) == 60
        # Flowback Gas + Condensate Rates from stage #1.
        assert float(exported["Gas Rate (MMSCFD)"]) == 8
        assert float(exported["Condensate Rate (BPD)"]) == 120
        # P50 trio from the SARH 'final' formation row.
        assert float(exported["P50 Pay Thickness (ft)"]) == 45
        assert float(exported["P50 Porosity (%)"]) == 12
        assert float(exported["Water Saturation (%)"]) == 30

        # AR Number round-trips, and the export DERIVES Seismic Block from it
        # (2525 -> Block A in the shipped map, agreeing with the sheet's cell).
        assert exported["AR Number"] == "2525"
        assert exported["Seismic Block"] == "Block A"
        reservoir_fields = workflow.get_task_dynamic_fields(
            session, _tasks_by_name(session, "MDFT-7")["Reservoir CoS"]["task_id"])
        stored_rows = json.loads(reservoir_fields["reservoir_cos_rows"])
        assert stored_rows[0]["seismic_volume_ar_number"] == "2525"
        assert stored_rows[0]["seismic_block"] == "Block A"
    finally:
        session.close()


# ---------------------------------------------------------------------------
# Case 3: duplicate skip, and --update merge semantics
# ---------------------------------------------------------------------------

def test_duplicate_skip_and_update_merge(client, app_modules, tmp_path):
    import import_excel
    import workflow

    base = {"Well Name": "UPD-1", "BP Year": 2027, "Status": "Gas", "Classification": "Appraisal",
            "OGIP Mean (BCF)": 10, "Trap CoS (%)": 50, "Reservoir CoS (%)": 40}
    _write_sheet(tmp_path / "base.xlsx", [base], header_row=1)

    session = _session(app_modules)
    try:
        first = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "base.xlsx")))
        assert first.results[0].outcome == "created"

        # Re-import WITHOUT --update: skipped as existing.
        again = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "base.xlsx")))
        assert again.results[0].outcome == "skipped"

        # Confirm a data-bearing BP step reached Approved on the first import.
        tasks = _tasks_by_name(session, "UPD-1")
        assert tasks["SAD Update"]["status"] == "Approved"

        # --update: change OGIP Mean, leave Trap CoS blank (must not erase).
        upd = {"Well Name": "UPD-1", "BP Year": 2027, "Status": "Gas", "OGIP Mean (BCF)": 20}
        _write_sheet(tmp_path / "upd.xlsx", [upd], header_row=1)
        updated = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "upd.xlsx")),
                                           update=True)
        assert updated.results[0].outcome == "updated"

        exported = _export_by_name(session)["UPD-1"]
        assert float(exported["OGIP Mean (BCF)"]) == 20       # changed
        assert float(exported["Trap CoS (%)"]) == 50          # blank cell did not erase

        # An Approved step stays Approved after the additive update.
        tasks = _tasks_by_name(session, "UPD-1")
        assert tasks["SAD Update"]["status"] == "Approved"
    finally:
        session.close()


# ---------------------------------------------------------------------------
# Case 4: validation (row errors + a cell warning + sheet-internal duplicate)
# ---------------------------------------------------------------------------

def test_validation_errors_and_cell_warning(client, app_modules, tmp_path):
    import import_excel

    rows = [
        {"Well Name": "V1", "Status": "Gas"},                                  # fluid, no BP Year -> error
        {"Well Name": "V2", "Status": "Sludge"},                               # unknown token -> error
        {"Well Name": "V3", "Status": "Proposed",
         "P90 Area (km2)": "abc", "P10 Area (km2)": 5},                        # non-numeric cell -> warning
        {"Well Name": "DUP-9", "Status": "Proposed", "P90 Area (km2)": 1},     # first DUP-9 -> created
        {"Well Name": "DUP-9", "Status": "Proposed", "P90 Area (km2)": 2},     # second DUP-9 -> error
    ]
    _write_sheet(tmp_path / "v.xlsx", rows, header_row=1)

    session = _session(app_modules)
    try:
        report = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "v.xlsx")))
        results = report.results

        # classify_row exposes the same row-local errors for V1/V2.
        assert classify_helper(import_excel, {"Well Name": "V1", "Status": "Gas"})[0] is None

        by_name = {}
        for res in results:
            by_name.setdefault(res.well_name, []).append(res)

        assert by_name["V1"][0].outcome == "error"
        assert "BP Year" in by_name["V1"][0].reason

        assert by_name["V2"][0].outcome == "error"
        assert "unknown Status token" in by_name["V2"][0].reason

        # V3: row lands despite the bad P90 cell; the warning is reported.
        v3 = by_name["V3"][0]
        assert v3.outcome == "created"
        assert any("P90 Area" in w for w in v3.warnings)

        # Sheet-internal duplicate: first created, second errored.
        dup_outcomes = [r.outcome for r in by_name["DUP-9"]]
        assert dup_outcomes == ["created", "error"]
        assert "duplicate name within the sheet" in by_name["DUP-9"][1].reason
    finally:
        session.close()


def classify_helper(import_excel, row):
    return import_excel.classify_row(row)


# ---------------------------------------------------------------------------
# Review fixes: year floor, partial seal inputs, update-merge, promotions
# ---------------------------------------------------------------------------

def test_year_before_1990_is_row_error_with_no_project(client, app_modules, tmp_path):
    """M3a: an out-of-range year fails at classification, BEFORE any write --
    no orphan half-record."""
    import import_excel

    _write_sheet(tmp_path / "old.xlsx", [{"Well Name": "OLD-1", "BP Year": 1980,
                                          "P90 Area (km2)": 2}], header_row=1)
    session = _session(app_modules)
    try:
        # classify_row surfaces the same row-local error.
        record_type, errors = import_excel.classify_row({"Well Name": "OLD-1", "BP Year": "1980"})
        assert record_type is None
        assert any("before 1990" in e for e in errors)

        report = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "old.xlsx")))
        assert report.results[0].outcome == "error"
        assert "before 1990" in report.results[0].reason
        assert _pid(session, "OLD-1") is None  # nothing was created
    finally:
        session.close()


def test_partial_seal_inputs_warn_and_pct_pins(client, app_modules, tmp_path):
    """M3b: an incomplete seal input set (only Dip) must not abort the row --
    the inputs are skipped with a warning and the sheet's pct still lands."""
    import import_excel
    import workflow

    row = {"Well Name": "SEAL-1", "Status": "Proposed", "Dip": 0.5, "Seal CoS (%)": 55,
           "Pore Pressure Gradient (psi/ft)": 0.45, "P90 Area (km2)": 3}
    _write_sheet(tmp_path / "seal.xlsx", [row], header_row=1)

    session = _session(app_modules)
    try:
        report = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "seal.xlsx")))
        result = report.results[0]
        assert result.outcome == "created", result.reason
        assert any("incomplete Seal CoS inputs" in w and "seal_dip" in w for w in result.warnings)

        fields = workflow.get_task_dynamic_fields(
            session, _tasks_by_name(session, "SEAL-1")["Trap and Seal CoS"]["task_id"])
        assert "seal_dip" not in fields                    # incomplete inputs skipped
        assert fields.get("seal_cos_pct") == "55"          # sheet pct pinned
        assert fields.get("seal_pore_pressure_gradient_psi_ft") == "0.45"  # rider kept
    finally:
        session.close()


def test_update_merges_json_minisheets_preserving_siblings(client, app_modules, tmp_path):
    """M1: --update merges the sheet's contribution into the PRIMARY row/stage
    of a stored JSON blob -- sibling rows/stages and in-blob keys survive."""
    import import_excel
    import workflow

    base = {"Well Name": "MRG-1", "BP Year": 2027, "Status": "Gas",
            "Reservoir CoS (%)": 40, "Seismic Block": "Block A", "Amplitude Ratio": 0.5,
            "Gas Rate (MMSCFD)": 8, "Water Rate (BWPD)": 200}
    _write_sheet(tmp_path / "m1a.xlsx", [base], header_row=1)

    session = _session(app_modules)
    try:
        assert import_excel.import_rows(
            session, import_excel.parse_workbook(str(tmp_path / "m1a.xlsx"))).results[0].outcome == "created"

        tasks = _tasks_by_name(session, "MRG-1")
        # Simulate a user adding a SECOND reservoir row (real AR) and a second
        # flowback stage after the import.
        stored = json.loads(workflow.get_task_dynamic_fields(
            session, tasks["Reservoir CoS"]["task_id"])["reservoir_cos_rows"])
        stored.append({"seismic_volume_ar_number": "2525", "reservoir_cos_pct": "70"})
        workflow.save_task_dynamic_fields(session, tasks["Reservoir CoS"]["task_id"],
                                          {"reservoir_cos_rows": json.dumps(stored)},
                                          changed_by="Test")
        stages = json.loads(workflow.get_task_dynamic_fields(
            session, tasks["Flowback Results"]["task_id"])["flowback_stages_rows"])
        # Fresh import writes the SARH default in-row (per-stage), not step-level.
        assert len(stages) == 1
        assert stages[0]["flowback_formation"] == "SARH"
        stages.append({"flowback_gas_rate_mmscfd": "3", "flowback_formation": "KHUFF"})
        workflow.save_task_dynamic_fields(session, tasks["Flowback Results"]["task_id"],
                                          {"flowback_stages_rows": json.dumps(stages)},
                                          changed_by="Test")

        # Update sheet contributes only BTS + Choke Size (everything else blank).
        upd = {"Well Name": "MRG-1", "BP Year": 2027, "Status": "Gas",
               "BTS": 0.9, "Choke Size (in)": 0.5}
        _write_sheet(tmp_path / "m1b.xlsx", [upd], header_row=1)
        report = import_excel.import_rows(
            session, import_excel.parse_workbook(str(tmp_path / "m1b.xlsx")), update=True)
        assert report.results[0].outcome == "updated", report.results[0].reason

        merged = json.loads(workflow.get_task_dynamic_fields(
            session, tasks["Reservoir CoS"]["task_id"])["reservoir_cos_rows"])
        assert len(merged) == 2                              # second row preserved
        assert merged[1]["seismic_volume_ar_number"] == "2525"
        assert merged[1]["reservoir_cos_pct"] == "70"
        primary = merged[0]
        assert primary["base_tight_sarah"] == "0.9"          # sheet cell landed
        assert primary["reservoir_cos_pct"] == "40"          # blank cells didn't erase
        assert primary["seismic_block"] == "Block A"
        assert primary["amplitude_ratio"] == "0.5"

        merged_stages = json.loads(workflow.get_task_dynamic_fields(
            session, tasks["Flowback Results"]["task_id"])["flowback_stages_rows"])
        assert len(merged_stages) == 2                       # second stage preserved
        assert merged_stages[0]["flowback_choke_size_in"] == "0.5"
        assert str(merged_stages[0]["flowback_gas_rate_mmscfd"]) == "8"   # sibling keys intact
        assert str(merged_stages[0]["flowback_water_rate_bwpd"]) == "200"
        # The update merge never sets flowback_formation, so the in-row values
        # (primary's SARH default + the user-chosen KHUFF on stage 2) survive.
        assert merged_stages[0]["flowback_formation"] == "SARH"
        assert merged_stages[1]["flowback_formation"] == "KHUFF"
    finally:
        session.close()


def test_update_applies_changed_bp_year(client, app_modules, tmp_path):
    """M2: --update on an already-promoted well moves a changed BP Year."""
    import import_excel

    _write_sheet(tmp_path / "y1.xlsx", [{"Well Name": "YRC-1", "BP Year": 2027,
                                         "Classification": "Appraisal"}], header_row=1)
    session = _session(app_modules)
    try:
        assert import_excel.import_rows(
            session, import_excel.parse_workbook(str(tmp_path / "y1.xlsx"))).results[0].outcome == "created"

        _write_sheet(tmp_path / "y2.xlsx", [{"Well Name": "YRC-1", "BP Year": 2031}], header_row=1)
        report = import_excel.import_rows(
            session, import_excel.parse_workbook(str(tmp_path / "y2.xlsx")), update=True)
        assert report.results[0].outcome == "updated", report.results[0].reason

        import db as db_module
        project = db_module.fetch_one(session,
                                      "SELECT business_plan_year, business_plan_enabled FROM projects WHERE project_name = 'YRC-1'",
                                      {})
        assert int(project["business_plan_year"]) == 2031
        assert int(project["business_plan_enabled"]) == 1
    finally:
        session.close()


def test_fluidless_bp_well_trio_lands_in_pre_drill(client, app_modules, tmp_path):
    """A BP well WITHOUT a fluid status routes its OGIP trio through
    pre_drill_piip_* and round-trips it through the export."""
    import import_excel
    import workflow

    row = {"Well Name": "NOF-1", "BP Year": 2028,
           "OGIP P90 (BCF)": 5, "OGIP Mean (BCF)": 10, "OGIP P10 (BCF)": 18}
    _write_sheet(tmp_path / "nof.xlsx", [row], header_row=1)

    session = _session(app_modules)
    try:
        report = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "nof.xlsx")))
        assert report.results[0].outcome == "created", report.results[0].reason

        fields = workflow.get_task_dynamic_fields(
            session, _tasks_by_name(session, "NOF-1")["Pre-Drilling GeoX Assessment"]["task_id"])
        assert fields.get("pre_drill_piip_gas_mean") == "10"

        exported = _export_by_name(session)["NOF-1"]
        assert float(exported["OGIP P90 (BCF)"]) == 5
        assert float(exported["OGIP Mean (BCF)"]) == 10
        assert float(exported["OGIP P10 (BCF)"]) == 18
    finally:
        session.close()


def test_update_promotes_existing_lead_when_sheet_gains_year(client, app_modules, tmp_path):
    """--update promotes a stored proposed lead to BP when the sheet now
    carries a BP Year (never the reverse)."""
    import import_excel
    import workflow

    _write_sheet(tmp_path / "p1.xlsx", [{"Well Name": "PRM-1", "Status": "Proposed",
                                         "P90 Area (km2)": 3}], header_row=1)
    session = _session(app_modules)
    try:
        assert import_excel.import_rows(
            session, import_excel.parse_workbook(str(tmp_path / "p1.xlsx"))).results[0].outcome == "created"

        _write_sheet(tmp_path / "p2.xlsx", [{"Well Name": "PRM-1", "BP Year": 2027}], header_row=1)
        report = import_excel.import_rows(
            session, import_excel.parse_workbook(str(tmp_path / "p2.xlsx")), update=True)
        assert report.results[0].outcome == "updated", report.results[0].reason
        assert report.results[0].record_type == "bp"

        bp_names = {p["project_name"] for p in workflow.get_projects(session, pipeline_filter="bp")}
        assert "PRM-1" in bp_names
        import db as db_module
        project = db_module.fetch_one(session,
                                      "SELECT business_plan_enabled, business_plan_year FROM projects WHERE project_name = 'PRM-1'",
                                      {})
        assert int(project["business_plan_enabled"]) == 1
        assert int(project["business_plan_year"]) == 2027
    finally:
        session.close()


def test_meanless_ogip_trio_warns(client, app_modules, tmp_path):
    """OGIP P90/P10 without a Mean: stored, but the export's assessment scan
    keys on the mean -- the operator gets a warning."""
    import import_excel

    row = {"Well Name": "MNL-1", "Status": "Proposed",
           "OGIP P90 (BCF)": 5, "OGIP P10 (BCF)": 12}
    _write_sheet(tmp_path / "mnl.xlsx", [row], header_row=1)

    session = _session(app_modules)
    try:
        report = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "mnl.xlsx")))
        result = report.results[0]
        assert result.outcome == "created", result.reason
        assert any("without 'OGIP Mean (BCF)'" in w for w in result.warnings)
    finally:
        session.close()


def test_multisheet_workbook_picks_best_matching_sheet(tmp_path):
    """Our own export is multi-sheet, and an EARLIER tab ("Wells Overview")
    also carries a "Well Name" header with a handful of contract columns.
    parse_workbook must score every sheet against PORTFOLIO_EXPORT_COLUMNS and
    pick the best match -- not the first sheet containing "Well Name" (that
    regression classified every export row as a data-less proposed lead)."""
    import import_excel

    workbook = openpyxl.Workbook()
    overview = workbook.active
    overview.title = "Wells Overview"
    # A decoy header: Well Name plus one contract column and one foreign one.
    for col_index, col in enumerate(["Well Name", "Classification", "Health"], start=1):
        overview.cell(row=1, column=col_index, value=col)
    overview.cell(row=2, column=1, value="DECOY-1")

    export = workbook.create_sheet("Portfolio Export")
    for col_index, col in enumerate(PORTFOLIO_EXPORT_COLUMNS, start=1):
        export.cell(row=4, column=col_index, value=col)  # export layout: startrow=3
    export.cell(row=5, column=PORTFOLIO_EXPORT_COLUMNS.index("Well Name") + 1, value="REAL-1")
    export.cell(row=5, column=PORTFOLIO_EXPORT_COLUMNS.index("BP Year") + 1, value=2027)
    workbook.save(str(tmp_path / "multi.xlsx"))

    rows = import_excel.parse_workbook(str(tmp_path / "multi.xlsx"))
    assert [r["Well Name"] for r in rows] == ["REAL-1"]
    assert rows[0]["BP Year"] == "2027"


def test_tight_and_slash_status_aliases(client, app_modules, tmp_path):
    """External sheets write "Tight" (a dry well) and may spell Gas over Water
    with a slash; both alias to the canonical fluid instead of erroring the
    row, and the canonical value is what stores/exports/shows as the status.
    The pre-v10 labels alias the same way -- FORWARD onto their replacements --
    so an old sheet never reintroduces retired vocabulary."""
    import import_excel
    import reporting

    # Tight is a FLUID status: classification demands a BP Year like any
    # drilled well, proving the token was recognized (not "unknown Status").
    record_type, errors = import_excel.classify_row({"Well Name": "T", "Status": "Tight"})
    assert record_type is None
    assert any("BP Year" in e for e in errors)

    rows = [{"Well Name": "TGT-1", "BP Year": 2027, "Status": "Tight"},
            {"Well Name": "GOW-1", "BP Year": 2027, "Status": "gas/water"},
            {"Well Name": "DRY-1", "BP Year": 2027, "Status": "Dry"},
            {"Well Name": "WAT-1", "BP Year": 2027, "Status": "water"},
            {"Well Name": "CND-1", "BP Year": 2027, "Status": "Condensate"},
            {"Well Name": "LIQ-1", "BP Year": 2027, "Status": "Liquid"}]
    _write_sheet(tmp_path / "tgt.xlsx", rows, header_row=1)

    session = _session(app_modules)
    try:
        report = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "tgt.xlsx")))
        outcomes = {r.well_name: r.outcome for r in report.results}
        assert outcomes == {name: "created" for name in
                            ("TGT-1", "GOW-1", "DRY-1", "WAT-1", "CND-1", "LIQ-1")}, report.format()

        expected = {"TGT-1": "Dry Hole", "GOW-1": "Gas over Water", "DRY-1": "Dry Hole",
                    "WAT-1": "Water Bearing", "CND-1": "Oil over Gas", "LIQ-1": "Oil"}
        portfolio = {r["well_name"]: r for r in reporting.get_portfolio_rows(session)["rows"]}
        assert {name: portfolio[name]["status"] for name in expected} == expected

        exported = _export_by_name(session)
        assert {name: exported[name]["Status"] for name in expected} == expected
    finally:
        session.close()


def test_commit_failure_mid_batch_recovers_and_continues(client, app_modules, tmp_path):
    """A commit that dies mid-row (e.g. SQLite 'database is locked' at COMMIT
    time) must not strand the session in the 'prepared' state and kill the
    batch: the failing row reports an error (its partial project cleaned up)
    and every later row still imports. Regression for the two-layer fix --
    db.write_transaction now rolls back a failed commit, and import_rows
    rolls back before its cleanup/next row."""
    import import_excel

    rows = [{"Well Name": "CF-1", "Status": "Proposed", "P90 Area (km2)": 1},
            {"Well Name": "CF-2", "Status": "Proposed", "P90 Area (km2)": 2}]
    _write_sheet(tmp_path / "cf.xlsx", rows, header_row=1)

    session = _session(app_modules)
    real_commit = session.commit
    calls = {"count": 0}

    def flaky_commit():
        calls["count"] += 1
        # Call #1 seeds the import user, #2 creates CF-1; #3 is a field save
        # inside CF-1's record -- failing THERE leaves a committed partial
        # project for the cleanup path to remove.
        if calls["count"] == 3:
            raise RuntimeError("simulated commit failure")
        real_commit()

    session.commit = flaky_commit
    try:
        report = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "cf.xlsx")))
        outcomes = {r.well_name: (r.outcome, r.reason) for r in report.results}
        assert outcomes["CF-1"][0] == "error"
        assert "simulated commit failure" in outcomes["CF-1"][1]
        assert "partially created project removed" in outcomes["CF-1"][1]
        assert _pid(session, "CF-1") is None          # cleanup delete succeeded
        assert outcomes["CF-2"] == ("created", "")    # the batch kept going
        assert _pid(session, "CF-2") is not None
    finally:
        session.commit = real_commit
        session.close()


def test_xy_coordinates_create_and_update(client, app_modules, tmp_path):
    """X/Y land on projects.lead_x/lead_y at create; --update overwrites them
    with non-blank cells and a blank cell never erases a stored coordinate."""
    import import_excel

    _write_sheet(tmp_path / "xy1.xlsx", [{"Well Name": "XY-1", "Status": "Proposed",
                                          "X": 512000, "Y": 2875000}], header_row=1)
    session = _session(app_modules)
    try:
        assert import_excel.import_rows(
            session, import_excel.parse_workbook(str(tmp_path / "xy1.xlsx"))).results[0].outcome == "created"
        import db as db_module
        project = db_module.fetch_one(
            session, "SELECT lead_x, lead_y FROM projects WHERE project_name = 'XY-1'", {})
        assert float(project["lead_x"]) == 512000
        assert float(project["lead_y"]) == 2875000

        # --update moves X only; the blank Y cell must keep the stored value.
        _write_sheet(tmp_path / "xy2.xlsx", [{"Well Name": "XY-1", "Status": "Proposed",
                                              "X": 513500}], header_row=1)
        report = import_excel.import_rows(
            session, import_excel.parse_workbook(str(tmp_path / "xy2.xlsx")), update=True)
        assert report.results[0].outcome == "updated", report.results[0].reason
        project = db_module.fetch_one(
            session, "SELECT lead_x, lead_y FROM projects WHERE project_name = 'XY-1'", {})
        assert float(project["lead_x"]) == 513500
        assert float(project["lead_y"]) == 2875000
    finally:
        session.close()


def test_ar_number_imports_and_block_mismatch_warns(client, app_modules, tmp_path):
    """A sheet AR lands in the stored reservoir row; a Seismic Block cell that
    contradicts the AR's mapped block draws a warning, because the export
    derives its Seismic Block column from the AR (the AR wins on round-trip)."""
    import import_excel

    row = {"Well Name": "ARX-1", "Status": "Proposed", "AR Number": 2525,
           "Seismic Block": "Block D", "Reservoir CoS (%)": 30}
    _write_sheet(tmp_path / "arx.xlsx", [row], header_row=1)

    session = _session(app_modules)
    try:
        report = import_excel.import_rows(session, import_excel.parse_workbook(str(tmp_path / "arx.xlsx")))
        result = report.results[0]
        assert result.outcome == "created", result.reason
        assert any("does not match AR 2525" in w for w in result.warnings)

        exported = _export_by_name(session)["ARX-1"]
        assert exported["AR Number"] == "2525"
        assert exported["Seismic Block"] == "Block A"  # derived from the AR, not the cell
    finally:
        session.close()
