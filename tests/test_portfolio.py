"""Tests for the Portfolio rework (reporting.get_portfolio_rows).

The Portfolio is the analysis surface for every non-archived record -- BP
wells PLUS leads at every maturity stage. Column sources each get pinned here
-- gas-field derivation from the project name, seismic AR -> block-name
mapping (config.AR_TO_SEISMIC_BLOCK, FIRST non-empty AR) with raw fallback,
the well-fluid ladder (resolve_well_fluid: SARH 'final'-phase formation fluid
-> legacy final_fluid_type -> resource_update -> post_drill -> SARH
'quicklook'-phase formation fluid -> legacy quicklook_fluid_type), status
(fluid -> Staked when Approval to Stake approved -> Proposed default),
mean-OGIP precedence (post-drill -> pre-drill -> lead), classification (BP
Execution Gate beats legacy GHEER), and lead/well membership (is_lead: 1 for
every non-BP-enabled record, regardless of maturity).
"""
from __future__ import annotations

import json

import pytest

from conftest import create_project, get_task_by_name, get_tasks

BP_KWARGS = {"business_plan_enabled": True, "business_plan_year": 2027}
PROSPECT_STAGES = {"Lead Assessment", "Risk Analysis", "Pre-Well Delivery"}
BP_STAGES = {"Well Delivery", "Post-Drilling", "Post-Testing"}


def _approve_all_prospect_tasks(client, pid):
    """Complete all checkpoints and approve every prospect lifecycle task."""
    lead_assessment = get_task_by_name(client, pid, "Lead Assessment")
    resp = client.patch(f"/api/tasks/{lead_assessment['task_id']}/dynamic-fields", json={
        "fields": {
            "p90_area_km2": "1", "p10_area_km2": "2",
            "reservoir_thickness_ft": "1", "formation_thickness_ft": "2",
            "grv_p90_thousand_acre_ft": "1", "grv_p10_thousand_acre_ft": "2",
            "polygons_surfaces_loaded": "1", "lead_piip_gas_mean": "1",
        }
    })
    assert resp.status_code == 200, resp.get_json()
    for task in get_tasks(client, pid):
        if task["stage_group"] in PROSPECT_STAGES and task["status"] != "Approved":
            resp = client.patch(f"/api/tasks/{task['task_id']}", json={
                "status": "Approved", "revision": task["revision"],
            })
            assert resp.status_code == 200, resp.get_json()


def _approve_all_bp_tasks(client, pid):
    """Approve every BP-stage task so the well fully matures (overall Completed).

    Mirrors _approve_all_prospect_tasks; used for the Excel import tool's
    "historical well" case (a BP well drilled in the past, every step Approved).
    """
    for task in get_tasks(client, pid):
        if task["stage_group"] in BP_STAGES and task["status"] != "Approved":
            resp = client.patch(f"/api/tasks/{task['task_id']}", json={
                "status": "Approved", "revision": task["revision"],
            })
            assert resp.status_code == 200, resp.get_json()


def _rows(client, **query):
    qs = "&".join(f"{k}={v}" for k, v in query.items())
    resp = client.get("/api/portfolio/rows" + (f"?{qs}" if qs else ""))
    assert resp.status_code == 200, resp.get_json()
    return resp.get_json()


def _row_for(client, pid):
    return next(r for r in _rows(client)["rows"] if r["project_id"] == pid)


def _save_fields(client, pid, task_name, fields):
    task = get_task_by_name(client, pid, task_name)
    resp = client.patch(f"/api/tasks/{task['task_id']}/dynamic-fields", json={"fields": fields})
    assert resp.status_code == 200, resp.get_json()


def _put_sarh_fluid(client, pid, phase, fluid):
    """Upsert a SARH formation row carrying ``fluid`` at ``phase`` (the well
    inherits SARH's per-formation fluid, replacing the old step-level select)."""
    resp = client.put(f"/api/projects/{pid}/formations",
                      json={"phase": phase, "rows": [{"formation": "SARH", "fluid": fluid}]})
    assert resp.status_code == 200, resp.get_json()


# ---------------------------------------------------------------------------
# Gas field derivation
# ---------------------------------------------------------------------------

def test_gas_field_is_name_prefix_before_first_hyphen(client):
    pid = create_project(client, "JOHN-4", **BP_KWARGS)
    row = _row_for(client, pid)
    assert row["well_name"] == "JOHN-4"
    assert row["gas_field"] == "JOHN"


def test_gas_field_whole_name_when_no_hyphen(client):
    pid = create_project(client, "SOLO", **BP_KWARGS)
    assert _row_for(client, pid)["gas_field"] == "SOLO"


# ---------------------------------------------------------------------------
# Seismic block
# ---------------------------------------------------------------------------

def test_seismic_block_maps_first_nonempty_ar_number(client, monkeypatch):
    import config
    monkeypatch.setattr(config, "AR_TO_SEISMIC_BLOCK", {"2525": "Block A"})

    pid = create_project(client, "SEISMIC-1", **BP_KWARGS)
    # The mapped AR '2525' is the FIRST non-empty AR; the later AR-1111111 must
    # be ignored under the first-non-empty rule.
    _save_fields(client, pid, "Reservoir CoS", {"reservoir_cos_rows": json.dumps([
        {"seismic_volume_ar_number": "", "reservoir_cos_pct": "50"},
        {"seismic_volume_ar_number": "2525", "reservoir_cos_pct": ""},
        {"seismic_volume_ar_number": "AR-1111111", "reservoir_cos_pct": "40"},
    ])})
    assert _row_for(client, pid)["seismic_block"] == "Block A"


def test_seismic_block_falls_back_to_raw_first_ar_number(client, monkeypatch):
    import config
    monkeypatch.setattr(config, "AR_TO_SEISMIC_BLOCK", {"2525": "Block A"})

    pid = create_project(client, "SEISMIC-2", **BP_KWARGS)
    # First non-empty AR is unmapped -> raw AR shows; the later mapped '2525'
    # is not consulted.
    _save_fields(client, pid, "Reservoir CoS", {"reservoir_cos_rows": json.dumps([
        {"seismic_volume_ar_number": "AR-9999999"},
        {"seismic_volume_ar_number": "2525"},
    ])})
    assert _row_for(client, pid)["seismic_block"] == "AR-9999999"


def test_seismic_block_empty_when_no_rows(client):
    pid = create_project(client, "SEISMIC-3", **BP_KWARGS)
    assert _row_for(client, pid)["seismic_block"] == ""


def test_seismic_block_key_survives_dynamic_fields_patch_round_trip(client):
    """A 'seismic_block' key inside a reservoir_cos_rows row is opaque to the
    storage path (PATCH /api/tasks/<id>/dynamic-fields -> save_task_dynamic_fields,
    the same route api.saveFields uses), so it must survive untouched. (Reservoir
    CoS recompute -- cos.calculate_reservoir_cos_rows -- only fires on the
    full PATCH /api/tasks/<id> save_task path; it does `dict(item)` plus one
    added key, so unknown keys survive there too, but this test pins the
    dynamic-fields route the seismic_block/AR dropdowns actually save through.)"""
    pid = create_project(client, "SEISMIC-4", **BP_KWARGS)
    task = get_task_by_name(client, pid, "Reservoir CoS")
    resp = client.patch(f"/api/tasks/{task['task_id']}/dynamic-fields", json={"fields": {
        "reservoir_cos_rows": json.dumps([
            {"seismic_block": "Block A", "seismic_volume_ar_number": "2525",
             "pull_up": "Yes", "amplitude_ratio": 0.5, "base_tight_sarah": 0.5,
             "reservoir_cos_pct": "40"},
        ]),
    }})
    assert resp.status_code == 200, resp.get_json()

    detail = client.get(f"/api/tasks/{task['task_id']}/dynamic-fields")
    assert detail.status_code == 200, detail.get_json()
    stored_rows = json.loads(detail.get_json()["reservoir_cos_rows"])
    assert stored_rows[0]["seismic_block"] == "Block A"
    assert stored_rows[0]["seismic_volume_ar_number"] == "2525"
    assert stored_rows[0]["reservoir_cos_pct"] == "40"

    # Also pin the OTHER save route, PATCH /api/tasks/<id> (save_task), which
    # DOES run the row set through cos.calculate_reservoir_cos_rows
    # (workflow/lifecycle.py's task_name == "Reservoir CoS" branch). That
    # function does `row = dict(item or {})` then adds only
    # `reservoir_cos_pct`, so unknown keys must survive this recompute too --
    # confirming no fix to cos.py was needed for seismic_block to round-trip.
    resp = client.patch(f"/api/tasks/{task['task_id']}", json={"fields": {
        "reservoir_cos_rows": json.dumps([
            {"seismic_block": "Block B", "seismic_volume_ar_number": "1201",
             "pull_up": "Yes", "amplitude_ratio": 0.5, "base_tight_sarah": 0.5},
        ]),
    }})
    assert resp.status_code == 200, resp.get_json()
    detail = client.get(f"/api/tasks/{task['task_id']}/dynamic-fields")
    recomputed_rows = json.loads(detail.get_json()["reservoir_cos_rows"])
    assert recomputed_rows[0]["seismic_block"] == "Block B"
    assert recomputed_rows[0]["seismic_volume_ar_number"] == "1201"
    assert "reservoir_cos_pct" in recomputed_rows[0]  # recomputed by the RF model


# ---------------------------------------------------------------------------
# config seismic-block map inversion (unit, no HTTP)
# ---------------------------------------------------------------------------

def test_ar_to_seismic_block_is_correct_inversion_of_shipped_map():
    """AR_TO_SEISMIC_BLOCK, as loaded from the real seismic_blocks.json at
    import time, must be the exact reverse index of SEISMIC_BLOCK_AR_MAP."""
    import config
    expected = {}
    for block, ars in config.SEISMIC_BLOCK_AR_MAP.items():
        for ar in ars:
            expected.setdefault(ar, block)
    assert config.AR_TO_SEISMIC_BLOCK == expected
    # Sanity: the shipped placeholder file actually has content to invert.
    assert config.SEISMIC_BLOCK_AR_MAP
    assert config.AR_TO_SEISMIC_BLOCK


# ---------------------------------------------------------------------------
# Status (Proposed / Staked / fluid) + raw fluid precedence
# ---------------------------------------------------------------------------

def test_status_defaults_to_proposed(client):
    pid = create_project(client, "STATUS-1", **BP_KWARGS)
    row = _row_for(client, pid)
    assert row["status"] == "Proposed"
    # The old 'Not Drilled Yet' fluid fallback is gone: raw fluid is ''.
    assert row["fluid"] == ""


def test_status_staked_when_approval_to_stake_approved(client):
    pid = create_project(client, "STATUS-2", **BP_KWARGS)
    task = get_task_by_name(client, pid, "Approval to Stake")
    resp = client.patch(f"/api/tasks/{task['task_id']}", json={
        "status": "Approved", "revision": task["revision"],
    })
    assert resp.status_code == 200, resp.get_json()
    row = _row_for(client, pid)
    assert row["status"] == "Staked"
    assert row["fluid"] == ""  # no fluid recorded yet


def test_status_staked_is_driven_by_card_4b_s_two_checkboxes(client):
    """Card 4B moved 'Approval to Stake' onto the field-completion engine, and
    THIS column is what that step's status has always meant.

    The Portfolio read is unchanged (reporting._approval_to_stake_map still
    looks for status == 'Approved'); what changed is who puts it there. So the
    contract to pin is the whole chain: ticking BOTH confirmations on the
    Staking Letters page -> the engine approves the step -> the record reads
    Staked. Ticking only the letter must NOT, or a lead with no well record
    would show up staked.
    """
    pid = create_project(client, "STATUS-4B", **BP_KWARGS)
    task = get_task_by_name(client, pid, "Approval to Stake")

    # The letter alone: still Proposed.
    resp = client.patch(f"/api/tasks/{task['task_id']}", json={
        "fields": {"approval_stake_letter_loaded": "1"}, "revision": task["revision"],
    })
    assert resp.status_code == 200, resp.get_json()
    assert resp.get_json()["task"]["status"] != "Approved"
    assert _row_for(client, pid)["status"] == "Proposed"

    # Both boxes: the engine approves the step and the record reads Staked.
    task = get_task_by_name(client, pid, "Approval to Stake")
    resp = client.patch(f"/api/tasks/{task['task_id']}", json={
        "fields": {"staking_well_created": "1", "approval_stake_letter_loaded": "1"},
        "revision": task["revision"],
    })
    assert resp.status_code == 200, resp.get_json()
    assert resp.get_json()["task"]["status"] == "Approved"
    assert _row_for(client, pid)["status"] == "Staked"

    # And unticking reopens the step, so the record falls back to Proposed --
    # the Portfolio can never show a stake that was withdrawn.
    task = get_task_by_name(client, pid, "Approval to Stake")
    resp = client.patch(f"/api/tasks/{task['task_id']}", json={
        "fields": {"staking_well_created": "1", "approval_stake_letter_loaded": ""},
        "revision": task["revision"],
    })
    assert resp.status_code == 200, resp.get_json()
    assert _row_for(client, pid)["status"] == "Proposed"


def test_fluid_wins_status_and_final_beats_quicklook(client):
    pid = create_project(client, "STATUS-3", **BP_KWARGS)
    _save_fields(client, pid, "Quicklook Logs", {"quicklook_fluid_type": "Gas"})
    row = _row_for(client, pid)
    assert row["fluid"] == "Gas"
    assert row["status"] == "Gas"  # a recorded fluid outranks Proposed/Staked

    _save_fields(client, pid, "Final Log Analysis", {"final_fluid_type": "Dry"})
    row = _row_for(client, pid)
    assert row["fluid"] == "Dry"
    assert row["status"] == "Dry"


def test_fluid_precedence_legacy_eav_ladder_no_sarh_row(client):
    """Old-well fallback: with NO SARH formation row, the legacy step-level EAV
    keys still resolve down final -> resource_update -> post_drill -> quicklook,
    driven through the API so record_status resolves the same ladder the raw
    fluid column does. (Nothing writes final_/quicklook_fluid_type anymore -- the
    step selects are gone -- but wells written before still populate this way.)"""
    pid = create_project(client, "STATUS-4", **BP_KWARGS)
    _save_fields(client, pid, "Quicklook Logs", {"quicklook_fluid_type": "Gas"})
    _save_fields(client, pid, "SAD Model",
                 {"post_drill_fluid_type": "Gas Condensate"})
    row = _row_for(client, pid)
    assert row["fluid"] == "Gas Condensate"  # post_drill beats quicklook
    assert row["status"] == "Gas Condensate"

    # A SAD Update (resource_update) revision beats post_drill and quicklook...
    _save_fields(client, pid, "SAD Update", {"resource_update_fluid_type": "Wet"})
    row = _row_for(client, pid)
    assert row["fluid"] == "Wet"
    assert row["status"] == "Wet"

    # ...but loses to the legacy final petrophysical read.
    _save_fields(client, pid, "Final Log Analysis", {"final_fluid_type": "Dry"})
    row = _row_for(client, pid)
    assert row["fluid"] == "Dry"
    assert row["status"] == "Dry"


def test_sarh_final_phase_fluid_beats_everything(client):
    """Top of the ladder: the SARH 'final'-phase formation fluid outranks every
    lower rung, including the legacy final_fluid_type EAV."""
    pid = create_project(client, "FLUID-SARH-1", **BP_KWARGS)
    _save_fields(client, pid, "Quicklook Logs", {"quicklook_fluid_type": "Gas"})
    _save_fields(client, pid, "SAD Model", {"post_drill_fluid_type": "Wet"})
    _save_fields(client, pid, "SAD Update", {"resource_update_fluid_type": "Gas Condensate"})
    _save_fields(client, pid, "Final Log Analysis", {"final_fluid_type": "Dry"})
    _put_sarh_fluid(client, pid, "quicklook", "Water")
    _put_sarh_fluid(client, pid, "final", "Oil")

    row = _row_for(client, pid)
    assert row["fluid"] == "Oil"    # SARH final beats the legacy final EAV
    assert row["status"] == "Oil"


def test_resource_update_and_post_drill_slot_between_sarh_phases(client):
    """Rungs 3-4 sit between the two SARH phases: with only a SARH 'quicklook'
    formation fluid present, the post_drill step fluid beats it, and a
    resource_update revision beats post_drill -- none reaching the SARH final rung
    (no final-phase row exists)."""
    pid = create_project(client, "FLUID-SARH-2", **BP_KWARGS)
    _put_sarh_fluid(client, pid, "quicklook", "Water")
    _save_fields(client, pid, "Quicklook Logs", {"quicklook_fluid_type": "Gas"})
    row = _row_for(client, pid)
    assert row["fluid"] == "Water"  # SARH quicklook beats legacy quicklook EAV

    _save_fields(client, pid, "SAD Model", {"post_drill_fluid_type": "Wet"})
    assert _row_for(client, pid)["fluid"] == "Wet"  # post_drill beats SARH quicklook

    _save_fields(client, pid, "SAD Update", {"resource_update_fluid_type": "Gas Condensate"})
    row = _row_for(client, pid)
    assert row["fluid"] == "Gas Condensate"  # resource_update beats post_drill
    assert row["status"] == "Gas Condensate"


def test_sarh_quicklook_fluid_beats_legacy_quicklook_eav(client):
    """Bottom two rungs: the SARH 'quicklook'-phase formation fluid outranks the
    legacy quicklook_fluid_type EAV."""
    pid = create_project(client, "FLUID-SARH-3", **BP_KWARGS)
    _save_fields(client, pid, "Quicklook Logs", {"quicklook_fluid_type": "Gas"})
    assert _row_for(client, pid)["fluid"] == "Gas"  # only the legacy EAV so far

    _put_sarh_fluid(client, pid, "quicklook", "Water")
    row = _row_for(client, pid)
    assert row["fluid"] == "Water"
    assert row["status"] == "Water"


# ---------------------------------------------------------------------------
# Mean OGIP precedence
# ---------------------------------------------------------------------------

def test_mean_ogip_precedence_post_beats_pre_beats_lead(client):
    pid = create_project(client, "OGIP-1", **BP_KWARGS)
    _save_fields(client, pid, "Lead Assessment", {"lead_piip_gas_mean": "5.0"})
    assert _row_for(client, pid)["mean_ogip"] == "5.0"

    _save_fields(client, pid, "Pre-Drilling GeoX Assessment", {"pre_drill_piip_gas_mean": "7.5"})
    assert _row_for(client, pid)["mean_ogip"] == "7.5"

    _save_fields(client, pid, "SAD Model", {"post_drill_piip_gas_mean": "9.25"})
    assert _row_for(client, pid)["mean_ogip"] == "9.25"


def test_summary_cumulative_ogip_sums_mean_ogip(client):
    pid_a = create_project(client, "SUM-A", **BP_KWARGS)
    pid_b = create_project(client, "SUM-B", **BP_KWARGS)
    _save_fields(client, pid_a, "Lead Assessment", {"lead_piip_gas_mean": "4.0"})
    _save_fields(client, pid_b, "SAD Model", {"post_drill_piip_gas_mean": "6.5"})

    payload = _rows(client)
    assert payload["summary"]["business_plan_wells"] == 2
    assert payload["summary"]["cumulative_ogip"] == 10.5


# ---------------------------------------------------------------------------
# Classification (BP Execution Gate, with legacy GHEER fallback)
# ---------------------------------------------------------------------------

def test_classification_falls_back_to_gheer_save(client):
    pid = create_project(client, "CLASS-1", **BP_KWARGS)
    assert _row_for(client, pid)["classification"] == ""

    _save_fields(client, pid, "GHEER", {"gheer_classification": "Appraisal"})
    assert _row_for(client, pid)["classification"] == "Appraisal"


def test_classification_bp_gate_beats_gheer(client):
    pid = create_project(client, "CLASS-2", **BP_KWARGS)
    _save_fields(client, pid, "GHEER", {"gheer_classification": "Exploration"})
    assert _row_for(client, pid)["classification"] == "Exploration"

    # The BP Execution Gate value is the primary source and wins over GHEER.
    _save_fields(client, pid, "BP Execution Gate", {"bp_gate_classification": "Development"})
    assert _row_for(client, pid)["classification"] == "Development"


# ---------------------------------------------------------------------------
# NUCD Area (projects.nucd_area) -- the column that replaced Classification in
# the Portfolio table. Record-level, not a step input: nothing in the UI writes
# it, so these drive the domain setter the importer calls.
# ---------------------------------------------------------------------------

def _set_area(app_modules, project_id, value, changed_by="External Import"):
    import workflow

    _main, db = app_modules
    session = db.new_session()
    try:
        return workflow.set_nucd_area(session, project_id, value, changed_by=changed_by)
    finally:
        session.close()


def test_nucd_area_is_blank_until_something_states_one(client):
    """No default, and nothing derives it from the field or seismic block: an
    area the app was never told is reported as unknown, not guessed."""
    pid = create_project(client, "AREA-1", **BP_KWARGS)
    row = _row_for(client, pid)
    assert row["nucd_area"] == ""
    # Classification still travels on the row (the export reads it); it simply
    # stopped being a column of the table.
    assert "classification" in row


def test_nucd_area_reaches_the_portfolio_row_trimmed(client, app_modules):
    pid = create_project(client, "AREA-2", **BP_KWARGS)
    assert _set_area(app_modules, pid, "  North   Jafurah ") == "North Jafurah"
    assert _row_for(client, pid)["nucd_area"] == "North Jafurah"

    # A lead is a record too -- the property is not BP-only.
    lead = create_project(client, "AREA-LEAD-1")
    _set_area(app_modules, lead, "South Ghawar")
    assert _row_for(client, lead)["nucd_area"] == "South Ghawar"


def test_nucd_area_is_audited_once_per_real_change(client, app_modules):
    pid = create_project(client, "AREA-3", **BP_KWARGS)
    assert _history(client, pid, "NUCD Area Changed") == []

    _set_area(app_modules, pid, "North Jafurah")
    _set_area(app_modules, pid, "North Jafurah")  # unchanged -- no second event
    events = _history(client, pid, "NUCD Area Changed")
    assert len(events) == 1
    assert (events[0]["old_status"], events[0]["new_status"]) == ("", "North Jafurah")
    assert events[0]["changed_by"] == "External Import"

    # Blank CLEARS the area (stored NULL, read back as ''), and says so.
    assert _set_area(app_modules, pid, "") == ""
    assert _row_for(client, pid)["nucd_area"] == ""
    events = _history(client, pid, "NUCD Area Changed")
    assert len(events) == 2
    assert (events[1]["old_status"], events[1]["new_status"]) == ("North Jafurah", "")
    assert events[1]["comment"] == "NUCD Area cleared."


def test_an_unknown_nucd_area_is_stored_not_rejected(client, app_modules):
    """No vocabulary has been supplied, so the setter validates length only.
    Refusing unrecognized areas would silently drop real ones."""
    pid = create_project(client, "AREA-4", **BP_KWARGS)
    assert _set_area(app_modules, pid, "Somewhere Nobody Listed") == "Somewhere Nobody Listed"

    with pytest.raises(ValueError, match="120 characters"):
        _set_area(app_modules, pid, "x" * 121)
    assert _row_for(client, pid)["nucd_area"] == "Somewhere Nobody Listed"


# ---------------------------------------------------------------------------
# Scope / filters keep working with the new row shape
# ---------------------------------------------------------------------------

def test_portfolio_includes_all_non_archived_records(client):
    bare_pid = create_project(client, "BARE-LEAD")
    pid = create_project(client, "IN-PORTFOLIO", **BP_KWARGS)
    rows = _rows(client)["rows"]
    assert {r["project_id"] for r in rows} == {bare_pid, pid}
    bare_row = next(r for r in rows if r["project_id"] == bare_pid)
    assert bare_row["status"] == "Proposed"


def test_portfolio_membership_includes_bp_wells_and_leads_at_every_stage(client):
    fresh_pid = create_project(client, "FRESH-PROSPECT")            # not mature
    bp_pid = create_project(client, "BP-WELL", pipeline_type="bp", **BP_KWARGS)
    lead_pid = create_project(client, "MATURE-LEAD")
    _approve_all_prospect_tasks(client, lead_pid)

    payload = _rows(client)
    by_id = {r["project_id"]: r for r in payload["rows"]}
    assert set(by_id) == {bp_pid, lead_pid, fresh_pid}                # every non-archived record

    assert by_id[bp_pid]["is_lead"] == 0
    assert by_id[bp_pid]["pipeline_type"] == "bp"
    assert by_id[lead_pid]["is_lead"] == 1
    assert by_id[lead_pid]["pipeline_type"] == "prospect"
    assert by_id[lead_pid]["status"] == "Staked"
    assert by_id[fresh_pid]["is_lead"] == 1
    assert by_id[fresh_pid]["status"] == "Proposed"

    assert payload["summary"]["business_plan_wells"] == 1
    assert payload["summary"]["leads"] == 2

    # A fully-matured lead leaves the prospect board; the fresh one stays.
    prospect_ids = [r["project_id"] for r in client.get("/api/projects?pipeline_filter=prospect").get_json()]
    assert lead_pid not in prospect_ids
    assert fresh_pid in prospect_ids


# ---------------------------------------------------------------------------
# risking_passed (Segmentation Slides approved) -- gates the promote "+" in
# the Portfolio UI (static/js/views/portfolio.js yearCellMarkup), independent
# of Approval to Stake.
# ---------------------------------------------------------------------------

def test_risking_passed_true_when_only_segmentation_slides_approved(client):
    pid = create_project(client, "RISK-PASSED-1")
    task = get_task_by_name(client, pid, "Segmentation Slides")
    resp = client.patch(f"/api/tasks/{task['task_id']}", json={
        "status": "Approved", "revision": task["revision"],
    })
    assert resp.status_code == 200, resp.get_json()
    assert _row_for(client, pid)["risking_passed"] == 1


def test_risking_passed_false_for_a_fresh_lead(client):
    pid = create_project(client, "RISK-PASSED-2")
    assert _row_for(client, pid)["risking_passed"] == 0


def test_portfolio_year_and_activity_filters(client):
    pid_2027 = create_project(client, "FILTER-2027", **BP_KWARGS)
    pid_2030 = create_project(client, "FILTER-2030", business_plan_enabled=True, business_plan_year=2030)
    client.patch(f"/api/projects/{pid_2030}/flags", json={"active_well_enabled": True})

    rows = _rows(client, year=2027)["rows"]
    assert [r["project_id"] for r in rows] == [pid_2027]

    rows = _rows(client, activity="Active")["rows"]
    assert [r["project_id"] for r in rows] == [pid_2030]

    rows = _rows(client, activity="Non-Active")["rows"]
    assert [r["project_id"] for r in rows] == [pid_2027]


# ---------------------------------------------------------------------------
# Historical BP wells (Excel import): fully-approved BP wells leave the BP
# board but stay in the Portfolio and the Excel export.
# ---------------------------------------------------------------------------

def test_completed_bp_well_leaves_bp_board_but_stays_in_portfolio_and_export(client):
    """The Excel importer creates "historical wells": BP wells drilled in the
    past (business_plan_year < 2026) whose steps are all Approved. Such a well
    must drop off the BP execution board the same way a matured lead drops off
    the prospect board (workflow/projects.py get_projects), while remaining
    visible in the Portfolio (reporting.get_portfolio_rows) and the Excel
    export (portfolio_export.get_portfolio_export_rows) -- those are separate
    readers, untouched by the board's row-skip."""
    import db as dbmod
    import portfolio_export
    import workflow

    # A historical business_plan_year (2019, pre-2026) is what the real
    # importer lands: POST /api/projects now gates business_plan_enabled at
    # creation to the same current-year..2035 promotion window (see
    # main.create_project), so a historical BP well is created plain and
    # promoted in-process with allow_historical_year=True, exactly like
    # import_excel.py does -- never through the HTTP route.
    pid = create_project(client, "HIST-1")
    session = dbmod.new_session()
    try:
        workflow.update_project_flags(session, pid, business_plan_enabled=True,
                                      business_plan_year=2019, changed_by="Import",
                                      allow_historical_year=True)
    finally:
        session.close()
    _approve_all_bp_tasks(client, pid)

    bp_ids = [r["project_id"] for r in client.get("/api/projects?pipeline_filter=bp").get_json()]
    assert pid not in bp_ids

    portfolio_ids = [r["project_id"] for r in _rows(client)["rows"]]
    assert pid in portfolio_ids

    session = dbmod.new_session()
    try:
        export_rows = portfolio_export.get_portfolio_export_rows(session)
    finally:
        session.close()
    assert "HIST-1" in [row["Well Name"] for row in export_rows]


# ---------------------------------------------------------------------------
# Business-plan year windows: promotion targets current-year..2035, while
# year edits on an already-enabled well keep the wide 1990 floor so imported
# historical wells stay editable.
# ---------------------------------------------------------------------------

def test_year_floor_is_strict_on_promotion_but_wide_on_edits(client):
    from datetime import date

    pid = create_project(client, "HIST-YEAR-1")
    resp = client.patch(f"/api/projects/{pid}/flags", json={
        "business_plan_enabled": True, "business_plan_year": 2019,
    })
    assert resp.status_code == 400

    resp = client.patch(f"/api/projects/{pid}/flags", json={
        "business_plan_enabled": True, "business_plan_year": date.today().year,
    })
    assert resp.status_code == 200, resp.get_json()

    # Already enabled: a year-only edit may land in history (import parity)...
    resp = client.patch(f"/api/projects/{pid}/flags", json={
        "business_plan_enabled": True, "business_plan_year": 2019,
    })
    assert resp.status_code == 200, resp.get_json()

    # ...but never below the absolute 1990 floor.
    resp = client.patch(f"/api/projects/{pid}/flags", json={
        "business_plan_enabled": True, "business_plan_year": 1989,
    })
    assert resp.status_code == 400


# ---------------------------------------------------------------------------
# Staked well name -- what a record is CALLED, and where
# ---------------------------------------------------------------------------
# Staking does not REWRITE projects.project_name -- that stays the lead name and
# the stable key, and only the explicit rename endpoint writes it. What staking
# does is decide the name the record is KNOWN BY. Card 3V makes that name
# canonical EVERYWHERE once staking is confirmed, Segment Maturation included;
# the lead name travels alongside so the pairing stays recoverable.
#
# CONFIRMED is an event, not a keystroke: the Well Site Location letter filed
# AND its coordinates recorded (the step's own completion predicate), plus a
# stored name. _confirm_staking below is that whole act.

def _confirm_staking(client, project_id, well_name):
    _save_fields(client, project_id, "Well Site Location", {
        "staked_well_name": well_name,
        "wellsite_letter_loaded": "1",
        "staked_x": "512340",
        "staked_y": "2765410",
    })

def test_staking_sets_the_portfolio_well_name_without_renaming_the_lead(client):
    pid = create_project(client, "STAKE-MAP-1", **BP_KWARGS)
    row = _row_for(client, pid)
    assert row["well_name"] == "STAKE-MAP-1", "falls back to the lead name while unstaked"
    assert row["staked_well_name"] == ""
    assert row["lead_name"] == "STAKE-MAP-1"

    # A name TYPED but not confirmed renames nothing.
    _save_fields(client, pid, "Well Site Location", {"staked_well_name": "STAKE-MAP-1ST2"})
    row = _row_for(client, pid)
    assert row["well_name"] == "STAKE-MAP-1", "an unconfirmed staking name is not canonical"
    assert row["staked_well_name"] == "STAKE-MAP-1ST2", "but it is still carried"

    _confirm_staking(client, pid, "STAKE-MAP-1ST2")
    row = _row_for(client, pid)
    # The Portfolio now calls it by its well name...
    assert row["well_name"] == "STAKE-MAP-1ST2"
    assert row["staked_well_name"] == "STAKE-MAP-1ST2"
    # ...while the lead name is still carried.
    assert row["lead_name"] == "STAKE-MAP-1"

    # Card 3V: Segment Maturation calls it by the same canonical name -- ONE
    # name source, so no surface disagrees about what a record is called. The
    # lead name rides alongside, and the STORED column is untouched, which is
    # what keeps relations and historical audit rows anchored.
    project = client.get(f"/api/projects/{pid}/detail").get_json()["project"]
    assert project["project_name"] == "STAKE-MAP-1ST2"
    assert project["lead_name"] == "STAKE-MAP-1"
    from conftest import raw_sqlite_connect
    conn = raw_sqlite_connect(client.db_path)
    try:
        stored = conn.execute("SELECT project_name FROM projects WHERE project_id = ?",
                              (pid,)).fetchone()["project_name"]
    finally:
        conn.close()
    assert stored == "STAKE-MAP-1", "staking never rewrites the stable name"


def test_business_plan_execution_calls_a_staked_record_by_its_well_name(client):
    pid = create_project(client, "STAKE-BPE-1", **BP_KWARGS)
    _confirm_staking(client, pid, "STAKE-BPE-1ST1")

    # BP_KWARGS puts the well in 2027; the dashboard defaults to the current year.
    wells = client.get("/api/business-plan/dashboard?year=2027").get_json()["wells"]
    well = next(w for w in wells if w["project_id"] == pid)
    assert well["project_name"] == "STAKE-BPE-1ST1"
    assert well["lead_name"] == "STAKE-BPE-1"
    # The FIELD still comes from the lead name: the field is where the segment
    # is, and a staked name need not carry the same prefix.
    assert well["field"] == "STAKE"

    detail = client.get(f"/api/business-plan/wells/{pid}/steps/business-plan-gate").get_json()
    assert detail["project"]["project_name"] == "STAKE-BPE-1ST1"
    assert detail["project"]["lead_name"] == "STAKE-BPE-1"


def test_export_names_a_staked_record_by_its_well_name_and_keeps_the_lead(client):
    import io

    import openpyxl

    import portfolio_export

    lead = create_project(client, "STAKE-XL-1")
    _confirm_staking(client, lead, "STAKE-XL-1ST1")

    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.data))

    portfolio = workbook["Portfolio Export"]
    header = [cell.value for cell in portfolio[4]]
    rows = [r for r in portfolio.iter_rows(min_row=5, max_row=portfolio.max_row, values_only=True)
            if r[header.index("Lead Name")] == "STAKE-XL-1"]
    assert len(rows) == 1
    # Known by its well name, paired with the lead it came from.
    assert rows[0][header.index("Well Name")] == "STAKE-XL-1ST1"

    staking = workbook["Staking Options"]
    staking_header = [cell.value for cell in staking[4]]
    staking_rows = [r for r in staking.iter_rows(min_row=5, max_row=staking.max_row, values_only=True)
                    if r[staking_header.index("Lead Name")] == "STAKE-XL-1"]
    assert len(staking_rows) == 1
    assert staking_rows[0][staking_header.index("Staked Well Name")] == "STAKE-XL-1ST1"

    # The Portfolio sheet's historical column POSITIONS are a contract for
    # external consumers, so a new column is appended, never inserted: every
    # column that shipped before keeps the index it had.
    assert portfolio_export.PORTFOLIO_EXPORT_COLUMNS[:3] == ["X", "Y", "Well Name"]
    assert portfolio_export.PORTFOLIO_EXPORT_COLUMNS[-2:] == ["Lead Name", "NUCD Area"]


# ---------------------------------------------------------------------------
# Card 3V -- the canonical name, guarded and audited
# ---------------------------------------------------------------------------

def _history(client, project_id, action_type):
    from conftest import raw_sqlite_connect
    conn = raw_sqlite_connect(client.db_path)
    try:
        return [dict(row) for row in conn.execute(
            "SELECT old_status, new_status, changed_by, comment FROM task_history "
            "WHERE project_id = ? AND action_type = ? ORDER BY history_id",
            (project_id, action_type))]
    finally:
        conn.close()


def test_confirming_staking_writes_exactly_one_canonical_name_event(client):
    pid = create_project(client, "AUDIT-1", **BP_KWARGS)
    assert _history(client, pid, "Canonical Name Set") == []

    _confirm_staking(client, pid, "AUDIT-1ST1")
    events = _history(client, pid, "Canonical Name Set")
    assert len(events) == 1
    assert (events[0]["old_status"], events[0]["new_status"]) == ("AUDIT-1", "AUDIT-1ST1")
    assert "AUDIT-1" in events[0]["comment"]

    # Replaying the same save, and editing the step again afterwards, must not
    # claim the record was renamed twice.
    _confirm_staking(client, pid, "AUDIT-1ST1")
    _save_fields(client, pid, "Well Site Location", {"staked_x": "512341"})
    assert len(_history(client, pid, "Canonical Name Set")) == 1


def test_an_unconfirmed_name_writes_no_event(client):
    """The event records a thing that HAPPENED. Typing a name is not that."""
    pid = create_project(client, "AUDIT-2", **BP_KWARGS)
    _save_fields(client, pid, "Well Site Location", {"staked_well_name": "AUDIT-2ST1"})
    assert _history(client, pid, "Canonical Name Set") == []


def test_a_staking_name_another_record_already_answers_to_is_refused(client):
    """A canonical name has to identify ONE record, or every surface showing it
    is ambiguous. Nothing is numbered, merged or silently altered."""
    other = create_project(client, "CLASH-TARGET-1", **BP_KWARGS)
    pid = create_project(client, "CLASH-1", **BP_KWARGS)

    task = get_task_by_name(client, pid, "Well Site Location")
    resp = client.patch(f"/api/tasks/{task['task_id']}/dynamic-fields",
                        json={"fields": {"staked_well_name": "CLASH-TARGET-1"}})
    assert resp.status_code == 400
    assert "already the name of another record" in resp.get_json()["detail"]

    # And against another record's STAKED name, which is equally a name people
    # are already using.
    _confirm_staking(client, other, "CLASH-TARGET-1ST9")
    resp = client.patch(f"/api/tasks/{task['task_id']}/dynamic-fields",
                        json={"fields": {"staked_well_name": "clash-target-1st9"}})
    assert resp.status_code == 400
    assert "already the staked well name" in resp.get_json()["detail"]

    # The record kept the name it had -- the save was rejected before any write.
    assert _row_for(client, pid)["staked_well_name"] == ""


def test_the_canonical_name_reaches_every_surface_at_once(client):
    """Card 3V's point: ONE name source, so no tab disagrees with another."""
    pid = create_project(client, "CROSS-1", **BP_KWARGS)
    _confirm_staking(client, pid, "CROSS-1ST3")

    # Segment Maturation board and detail (the surfaces that used to carve
    # themselves out of this rule).
    board = client.get("/api/projects").get_json()
    row = next(r for r in board if r["project_id"] == pid)
    assert (row["project_name"], row["lead_name"]) == ("CROSS-1ST3", "CROSS-1")

    detail = client.get(f"/api/projects/{pid}/detail").get_json()["project"]
    assert (detail["project_name"], detail["lead_name"]) == ("CROSS-1ST3", "CROSS-1")

    # Portfolio.
    assert _row_for(client, pid)["well_name"] == "CROSS-1ST3"

    # Business Plan Execution.
    wells = client.get("/api/business-plan/dashboard?year=2027").get_json()["wells"]
    assert next(w for w in wells if w["project_id"] == pid)["project_name"] == "CROSS-1ST3"

    # And the folder destinations resolve under the name the well carries.
    folder = client.get(
        f"/api/business-plan/wells/{pid}/steps/quicklook-logs").get_json()["folder"]
    assert folder["unc_path"].endswith(r"WELLS\CROSS\CROSS-1ST3\LOGS\QUICKLOOK_LOGS")
